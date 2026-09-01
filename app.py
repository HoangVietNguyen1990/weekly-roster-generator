import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import os
import sys
import copy
import json
import re
import math
from datetime import datetime, timedelta

# --- FIREBASE AUTHENTICATION & CLOUD FIRESTORE STORAGE ENGINE ---
FIREBASE_INITIALIZED = False
FIREBASE_DB = None
FIREBASE_AUTH = None
FIREBASE_WEB_API_KEY = ""

def get_firebase_db():
    global FIREBASE_INITIALIZED, FIREBASE_DB, FIREBASE_AUTH, FIREBASE_WEB_API_KEY
    if FIREBASE_INITIALIZED:
        return FIREBASE_DB
        
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore, auth
        
        firebase_config = None
        if hasattr(st, "secrets"):
            if "firebase" in st.secrets:
                firebase_config = dict(st.secrets["firebase"])
            elif "FIREBASE" in st.secrets:
                firebase_config = dict(st.secrets["FIREBASE"])
                
        if not firebase_config:
            secrets_path = os.path.join(".streamlit", "secrets.toml")
            if os.path.exists(secrets_path):
                try:
                    if sys.version_info >= (3, 11):
                        import tomllib
                        with open(secrets_path, "rb") as f:
                            toml_data = tomllib.load(f)
                    else:
                        import toml
                        toml_data = toml.load(secrets_path)
                    firebase_config = toml_data.get("firebase") or toml_data.get("FIREBASE")
                except Exception:
                    pass

        if firebase_config:
            if "private_key" in firebase_config and isinstance(firebase_config["private_key"], str):
                firebase_config["private_key"] = firebase_config["private_key"].replace("\\n", "\n")
                
            FIREBASE_WEB_API_KEY = firebase_config.get("web_api_key", "")
            
            if not firebase_admin._apps:
                cred = credentials.Certificate(firebase_config)
                firebase_admin.initialize_app(cred)
                
            FIREBASE_DB = firestore.client()
            FIREBASE_AUTH = auth
            FIREBASE_INITIALIZED = True
            return FIREBASE_DB
    except Exception:
        FIREBASE_INITIALIZED = False
        FIREBASE_DB = None
        
    return None

def is_firebase_active():
    return get_firebase_db() is not None

def firestore_save_df(collection_name, df):
    db = get_firebase_db()
    if db is None or df is None:
        return False
    try:
        records = df.astype(str).to_dict(orient="records")
        doc_ref = db.collection(collection_name).document("master_list")
        doc_ref.set({"records": records, "updated_at": datetime.now().isoformat()})
        return True
    except Exception:
        return False

def firestore_load_df(collection_name):
    db = get_firebase_db()
    if db is None:
        return None
    try:
        doc_ref = db.collection(collection_name).document("master_list")
        doc = doc_ref.get()
        if doc.exists:
            data = doc.to_dict()
            records = data.get("records", [])
            if records:
                return pd.DataFrame(records)
    except Exception:
        pass
    return None

def firestore_save_profiles(profiles):
    db = get_firebase_db()
    if db is None or not profiles:
        return False
    try:
        doc_ref = db.collection("system").document("user_profiles")
        doc_ref.set({"profiles": profiles, "updated_at": datetime.now().isoformat()})
        return True
    except Exception:
        return False

def firestore_load_profiles():
    db = get_firebase_db()
    if db is None:
        return None
    try:
        doc_ref = db.collection("system").document("user_profiles")
        doc = doc_ref.get()
        if doc.exists:
            return doc.to_dict().get("profiles")
    except Exception:
        pass
    return None

def migrate_all_local_files_to_firebase():
    """One-click migration tool: uploads all CSVs, user_profiles.json, and config to Firebase Cloud Firestore."""
    db = get_firebase_db()
    if db is None:
        return False, "⚠️ Firebase is not configured yet. Please add Firebase secrets to `.streamlit/secrets.toml` or Streamlit Cloud Secrets dashboard first."
        
    uploaded_files = []
    try:
        # 1. Employees CSV
        emp_path = os.path.join(DATA_DIR, "employees.csv")
        if os.path.exists(emp_path):
            emp_df = pd.read_csv(emp_path, dtype=str, keep_default_na=False)
            if emp_df is not None:
                firestore_save_df("employees", emp_df)
                uploaded_files.append("employees.csv")
            
        # 2. Unavailability CSV
        unavail_path = os.path.join(DATA_DIR, "unavailability.csv")
        if os.path.exists(unavail_path):
            unavail_df = pd.read_csv(unavail_path, dtype=str, keep_default_na=False)
            if unavail_df is not None:
                firestore_save_df("unavailability", unavail_df)
                uploaded_files.append("unavailability.csv")
            
        # 3. Requirements CSV
        req_path = os.path.join(DATA_DIR, "requirements.csv")
        if os.path.exists(req_path):
            req_df = pd.read_csv(req_path, dtype=str, keep_default_na=False)
            if req_df is not None:
                firestore_save_df("requirements", req_df)
                uploaded_files.append("requirements.csv")
            
        # 4. Fixed shifts CSV
        fixed_path = os.path.join(DATA_DIR, "fixed.csv")
        if os.path.exists(fixed_path):
            fixed_df = pd.read_csv(fixed_path, dtype=str, keep_default_na=False)
            if fixed_df is not None:
                firestore_save_df("fixed", fixed_df)
                uploaded_files.append("fixed.csv")
            
        # 5. User Profiles JSON
        if os.path.exists(USER_PROFILES_FILE):
            with open(USER_PROFILES_FILE, "r", encoding="utf-8") as f:
                profiles = json.load(f)
                if profiles:
                    firestore_save_profiles(profiles)
                    uploaded_files.append("user_profiles.json")
            
        return True, f"🎉 Successfully uploaded {len(uploaded_files)} database resource(s) to Firebase Cloud Firestore: {', '.join(uploaded_files)}!"
    except Exception as e:
        return False, f"Migration error: {e}"

# Set page configuration with a modern design
st.set_page_config(
    page_title="Brumby's Bakery Roster Creator",
    page_icon="🥐",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom high-contrast Emerald & Golden Wheat Bakery styling
st.markdown("""
<style>
    /* HIDE TOP STREAMLIT HEADER, TOOLBAR (Fork, GitHub, 3-Dots Menu) & FOOTER */
    header[data-testid="stHeader"],
    [data-testid="stHeader"],
    .stAppHeader,
    #MainMenu,
    footer,
    [data-testid="stFooter"],
    [data-testid="stToolbar"],
    div[data-testid="stDecoration"] {
        visibility: hidden !important;
        display: none !important;
        height: 0px !important;
    }

    /* REDUCE TOP CONTAINER PADDING FOR CLEAN FULL-SCREEN LAYOUT */
    .main .block-container,
    div[data-testid="stMainBlockContainer"],
    section.main > div {
        padding-top: 0.8rem !important;
    }

    /* Main App Background - Deep Emerald & Forest Teal Gradient */
    .stApp {
        background: linear-gradient(135deg, #0e2b26 0%, #16443c 50%, #1f574d 100%);
        color: #ffffff;
    }

    /* GLOBAL TEXT OVERRIDE FOR ALL LABELS & WIDGET TITLES */
    label, 
    label p, 
    label span, 
    div[data-testid="stWidgetLabel"], 
    div[data-testid="stWidgetLabel"] *, 
    .stMarkdown, 
    .stMarkdown p, 
    .stMarkdown span {
        color: #ffffff !important;
        font-weight: 600 !important;
        opacity: 1 !important;
    }
    
    /* Code tag override for dark emerald/gold theme */
    code, .stMarkdown code {
        background-color: rgba(9, 38, 33, 0.9) !important;
        color: #f7d594 !important;
        border: 1px solid rgba(229, 169, 60, 0.4) !important;
        border-radius: 6px !important;
        padding: 2px 8px !important;
        font-family: inherit !important;
        font-weight: 700 !important;
    }
    
    /* Header Banner Styling - Ultra Standout Title */
    .header-style {
        background: linear-gradient(135deg, #ffffff 0%, #f7d594 40%, #e5a93c 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-weight: 900;
        font-size: 3.2rem;
        margin-bottom: 6px;
        letter-spacing: -0.8px;
        filter: drop-shadow(0 2px 8px rgba(0,0,0,0.5));
    }
    .sub-header-style {
        color: #c8e6e0 !important;
        font-size: 1.2rem;
        margin-bottom: 25px;
        font-weight: 400;
    }

    /* Subheadings */
    h1, h2, h3, h4, h5, h6 {
        color: #ffffff !important;
        font-weight: 700 !important;
    }

    /* TABLE HEADERS: GREEN BACKGROUND (#081D19) WITH BOLD WHITE TEXT (#FFFFFF) ACROSS ALL TABS */
    div[data-testid="stDataEditor"], div[data-testid="stDataFrame"] {
        border-radius: 0 0 14px 14px !important;
        overflow: hidden !important;
        box-shadow: 0 8px 28px rgba(0, 0, 0, 0.45) !important;
        border: 2px solid #e5a93c !important;
        background-color: #ffffff !important;
        
        /* Glide Data Grid Header Canvas CSS Variables */
        --gdg-bg-header: #081d19 !important;
        --gdg-bg-header-has-focus: #133b34 !important;
        --gdg-text-header: #ffffff !important;
        --gdg-text-header-selected: #ffffff !important;
        --gdg-font-family: 'Inter', sans-serif !important;
    }
    
    /* Fallback DOM Header Element Styling - Green Background & White Text */
    div[data-testid="stDataEditor"] th, 
    div[data-testid="stDataFrame"] th, 
    div[class*="header"], 
    div[class*="Header"],
    div[role="columnheader"],
    div[class*="gdg-header"] {
        background-color: #081d19 !important;
        background: linear-gradient(135deg, #081d19 0%, #16443c 100%) !important;
        color: #ffffff !important;
        font-weight: 900 !important;
        font-size: 1.05rem !important;
        border-bottom: 2px solid #e5a93c !important;
    }
    div[role="columnheader"] *, div[role="columnheader"] p, div[role="columnheader"] span {
        color: #ffffff !important;
        font-weight: 900 !important;
    }

    /* TAB BAR CUSTOMIZATION - ULTRA-WIDE ROUNDED PILL BUTTONS */
    /* TACTILE 3D TABS STYLING */
    .stTabs [data-baseweb="tab-list"] {
        gap: 16px !important;
        background-color: rgba(6, 24, 20, 0.95) !important;
        padding: 16px 24px !important;
        border-radius: 40px !important;
        border: 2px solid rgba(229, 169, 60, 0.45) !important;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.6), inset 0 2px 4px rgba(0,0,0,0.5) !important;
    }
    .stTabs button[role="tab"],
    .stTabs [data-baseweb="tab"] {
        height: 54px !important;
        padding-left: 36px !important;
        padding-right: 36px !important;
        border-radius: 28px !important;
        font-size: 1.05rem !important;
        transition: all 0.15s ease-in-out !important;
        margin: 0 4px !important;
        cursor: pointer !important;
    }
    .stTabs [aria-selected="false"] {
        background: linear-gradient(180deg, #1d574c 0%, #134038 50%, #0b2923 100%) !important;
        border: 2px solid #1f5c50 !important;
        box-shadow: 0 5px 0 #061814, 0 6px 14px rgba(0, 0, 0, 0.4), inset 0 1px 0 rgba(255, 255, 255, 0.2) !important;
    }
    .stTabs [aria-selected="false"]:hover {
        background: linear-gradient(180deg, #24685b 0%, #184c42 50%, #0d332c 100%) !important;
        border-color: #e5a93c !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 7px 0 #061814, 0 8px 18px rgba(0, 0, 0, 0.5), inset 0 1px 0 rgba(255, 255, 255, 0.3) !important;
    }
    .stTabs [aria-selected="false"] * {
        color: #ffffff !important;
        font-weight: 700 !important;
        text-shadow: 0 1px 2px rgba(0,0,0,0.8) !important;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(180deg, #fce4b3 0%, #e5a93c 45%, #b87b1c 100%) !important;
        border: 2px solid #ffe8be !important;
        box-shadow: 0 6px 0 #734c0e, 0 10px 24px rgba(229, 169, 60, 0.6), inset 0 1px 0 rgba(255, 255, 255, 0.7) !important;
        transform: translateY(-2px) !important;
    }
    .stTabs [aria-selected="true"] * {
        color: #081d19 !important;
        font-weight: 900 !important;
        text-shadow: 0 1px 1px rgba(255, 255, 255, 0.4) !important;
    }

    /* UNIFIED HIGH-CONTRAST 3D BUTTON STYLING FOR ALL BUTTONS ACROSS THE ENTIRE APP */
    button,
    .stButton > button,
    .stFormSubmitButton > button,
    button[data-testid="stBaseButton-secondary"],
    button[data-testid="stBaseButton-primary"],
    button[data-testid="baseButton-secondary"],
    button[data-testid="baseButton-primary"],
    section[data-testid="stSidebar"] button,
    div[data-testid="stSidebar"] button,
    div[data-testid="stSidebar"] .stButton > button,
    button[kind="secondary"],
    button[kind="primary"],
    button[kind="header"] {
        background: linear-gradient(180deg, #fce4b3 0%, #e5a93c 45%, #b87b1c 100%) !important;
        border: 2px solid #ffe8be !important;
        border-radius: 14px !important;
        color: #081d19 !important;
        font-weight: 900 !important;
        font-size: 1.02rem !important;
        padding: 10px 22px !important;
        box-shadow: 0 5px 0 #734c0e, 0 8px 20px rgba(0, 0, 0, 0.4), inset 0 1px 0 rgba(255, 255, 255, 0.7) !important;
        transition: all 0.15s ease-in-out !important;
        cursor: pointer !important;
        text-shadow: 0 1px 1px rgba(255, 255, 255, 0.4) !important;
    }
    
    /* ALL BUTTON TEXT & SPAN OVERRIDE TO DEEP EMERALD BLACK */
    button *,
    .stButton > button *,
    .stFormSubmitButton > button *,
    button[data-testid="stBaseButton-secondary"] *,
    button[data-testid="stBaseButton-primary"] *,
    button[data-testid="baseButton-secondary"] *,
    button[data-testid="baseButton-primary"] *,
    section[data-testid="stSidebar"] button *,
    div[data-testid="stSidebar"] button * {
        color: #081d19 !important;
        font-weight: 900 !important;
    }

    button:hover,
    .stButton > button:hover,
    .stFormSubmitButton > button:hover,
    button[data-testid="stBaseButton-secondary"]:hover,
    button[data-testid="stBaseButton-primary"]:hover,
    section[data-testid="stSidebar"] button:hover,
    div[data-testid="stSidebar"] button:hover {
        background: linear-gradient(180deg, #fff0d4 0%, #f0b548 45%, #c78822 100%) !important;
        box-shadow: 0 6px 0 #734c0e, 0 12px 24px rgba(229, 169, 60, 0.5), inset 0 1px 0 rgba(255, 255, 255, 0.8) !important;
        transform: translateY(-2px) !important;
    }

    button:active,
    .stButton > button:active,
    .stFormSubmitButton > button:active,
    button[data-testid="stBaseButton-secondary"]:active,
    button[data-testid="stBaseButton-primary"]:active,
    section[data-testid="stSidebar"] button:active,
    div[data-testid="stSidebar"] button:active {
        box-shadow: 0 2px 0 #734c0e, 0 4px 10px rgba(0, 0, 0, 0.4), inset 0 1px 0 rgba(0, 0, 0, 0.2) !important;
        transform: translateY(3px) !important;
    }

    /* TACTILE 3D EXPANDER BUTTON STYLING */
    .stExpander, 
    div[data-testid="stExpander"], 
    details {
        background-color: transparent !important;
        border: none !important;
        box-shadow: none !important;
        margin-bottom: 14px !important;
    }
    .stExpander summary, 
    div[data-testid="stExpander"] summary,
    details summary {
        background: linear-gradient(180deg, #1d574c 0%, #144038 50%, #0b2923 100%) !important;
        border: 2px solid #e5a93c !important;
        border-radius: 12px !important;
        color: #f7d594 !important;
        padding: 12px 18px !important;
        box-shadow: 0 5px 0 #061814, 0 8px 20px rgba(0, 0, 0, 0.5), inset 0 1px 0 rgba(255, 255, 255, 0.25) !important;
        transition: all 0.15s ease-in-out !important;
        cursor: pointer !important;
    }
    .stExpander summary:hover,
    div[data-testid="stExpander"] summary:hover,
    details summary:hover {
        background: linear-gradient(180deg, #23665a 0%, #184c42 50%, #0e332c 100%) !important;
        border-color: #f7d594 !important;
        box-shadow: 0 6px 0 #061814, 0 10px 24px rgba(0, 0, 0, 0.6), inset 0 1px 0 rgba(255, 255, 255, 0.35) !important;
        transform: translateY(-1px) !important;
    }
    .stExpander summary:active,
    div[data-testid="stExpander"] summary:active,
    details summary:active {
        box-shadow: 0 2px 0 #061814, 0 4px 10px rgba(0, 0, 0, 0.4), inset 0 1px 0 rgba(0, 0, 0, 0.2) !important;
        transform: translateY(3px) !important;
    }
    .stExpander summary *, 
    div[data-testid="stExpander"] summary *, 
    div[data-testid="stExpander"] summary p, 
    div[data-testid="stExpander"] summary span, 
    details summary p, 
    details summary span,
    [data-testid="stExpanderToggleIcon"],
    svg[data-testid="stExpanderToggleIcon"] {
        color: #f7d594 !important;
        font-weight: 800 !important;
        font-size: 1.05rem !important;
        fill: #f7d594 !important;
        text-shadow: 0 1px 3px rgba(0, 0, 0, 0.8) !important;
    }
    div[data-testid="stExpanderDetails"],
    details[open] > div {
        background: #0f2e29 !important;
        border: 2px solid #e5a93c !important;
        border-top: none !important;
        border-radius: 0 0 12px 12px !important;
        padding: 16px !important;
        margin-top: -6px !important;
        box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4) !important;
    }
    .stTabs [data-baseweb="tab-highlight-container"] {
        display: none !important;
    }

    /* Radio Buttons High Contrast */
    .stRadio label, .stRadio div, .stRadio span, .stRadio p, .stRadio * {
        color: #ffffff !important;
        font-size: 1rem !important;
        font-weight: 600 !important;
        opacity: 1 !important;
    }

    /* File Uploader Dropzone & Instruction Text High Contrast */
    [data-testid="stFileUploaderDropzone"] {
        background-color: rgba(9, 32, 28, 0.85) !important;
        border: 2px dashed rgba(229, 169, 60, 0.6) !important;
        border-radius: 12px !important;
        padding: 12px !important;
    }
    [data-testid="stFileUploaderDropzone"]:hover {
        border-color: #e5a93c !important;
        background-color: rgba(9, 32, 28, 0.95) !important;
    }
    [data-testid="stFileUploaderDropzone"] *,
    [data-testid="stFileUploaderInstructions"] *,
    section[data-testid="stFileUploader"] * {
        color: #ffffff !important;
        opacity: 1 !important;
    }

    /* FILE UPLOADER BUTTON SPECIFIC - CLEAN SINGLE GOLDEN BUTTON NO NESTING */
    [data-testid="stFileUploaderDropzone"] button {
        background-color: #e5a93c !important;
        background: linear-gradient(135deg, #e5a93c 0%, #d48827 100%) !important;
        color: #0e2b26 !important;
        border: none !important;
        border-radius: 8px !important;
        box-shadow: none !important;
        padding: 8px 18px !important;
    }
    [data-testid="stFileUploaderDropzone"] button * {
        background: transparent !important;
        color: #0e2b26 !important;
        border: none !important;
        box-shadow: none !important;
        font-weight: 800 !important;
    }

    /* Hero Generate Button Styling - Tight Under Date Picker */
    .hero-generate-btn {
        margin-top: 5px !important;
    }
    .hero-generate-btn button {
        width: 100% !important;
        background: linear-gradient(135deg, #e5a93c 0%, #d48827 100%) !important;
        color: #0e2b26 !important;
        border: none !important;
        padding: 14px 28px !important;
        border-radius: 12px !important;
        font-weight: 900 !important;
        font-size: 1.2rem !important;
        letter-spacing: 0.5px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 6px 20px rgba(229, 169, 60, 0.45) !important;
        margin-top: 0px !important;
        margin-bottom: 5px !important;
    }
    .hero-generate-btn button p, .hero-generate-btn button span {
        color: #0e2b26 !important;
        font-weight: 900 !important;
        font-size: 1.2rem !important;
    }
    .hero-generate-btn button:hover {
        transform: translateY(-2px) scale(1.01) !important;
        box-shadow: 0 10px 28px rgba(229, 169, 60, 0.65) !important;
        background: linear-gradient(135deg, #f7d594 0%, #e5a93c 100%) !important;
    }

    /* EXPLICIT LOGOUT BUTTON STYLING OVERRIDE FOR HEADER & SIDEBAR */
    button[key*="logout"],
    button[key="btn_logout"],
    button[key="btn_logout_header"],
    div[data-testid="stSidebar"] button[key="btn_logout"] {
        background: linear-gradient(180deg, #fce4b3 0%, #e5a93c 45%, #b87b1c 100%) !important;
        border: 2px solid #ffe8be !important;
        border-radius: 12px !important;
        color: #081d19 !important;
        font-weight: 900 !important;
        box-shadow: 0 4px 0 #734c0e, 0 6px 16px rgba(0, 0, 0, 0.4) !important;
    }
    button[key*="logout"] *,
    button[key="btn_logout"] *,
    button[key="btn_logout_header"] *,
    div[data-testid="stSidebar"] button[key="btn_logout"] * {
        color: #081d19 !important;
        font-weight: 900 !important;
    }

    /* Download Button Specific Accent */
    .stDownloadButton>button,
    button[data-testid="stDownloadButton"] {
        background: linear-gradient(180deg, #3dbd98 0%, #1f8567 45%, #10523f 100%) !important;
        border: 2px solid #5ce4bc !important;
        color: #ffffff !important;
        border-radius: 12px !important;
        font-weight: 900 !important;
        box-shadow: 0 5px 0 #083327, 0 8px 20px rgba(31, 133, 103, 0.4) !important;
    }
    .stDownloadButton>button *,
    button[data-testid="stDownloadButton"] * {
        color: #ffffff !important;
        font-weight: 900 !important;
    }

    /* Sidebar Styling High Contrast */
    section[data-testid="stSidebar"] {
        background-color: #09201c;
        border-right: 1px solid rgba(229, 169, 60, 0.2);
    }
    section[data-testid="section-sidebar"] * {
        color: #ffffff !important;
    }

    /* GLOBAL INPUT & WIDGET TEXT CONTRAST ENHANCEMENTS */
    div[data-baseweb="input"] input,
    div[data-baseweb="base-input"] input,
    div[data-baseweb="textarea"] textarea,
    div[data-baseweb="select"] div,
    div[data-baseweb="select"] input,
    div[data-baseweb="select"] span,
    .stTextInput input, 
    .stNumberInput input, 
    .stDateInput input, 
    .stSelectbox div[role="combobox"], 
    .stTextArea textarea {
        background-color: #0c2b25 !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: 1.5px solid #e5a93c !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        opacity: 1 !important;
        border-radius: 8px !important;
    }

    /* ULTRA HIGH-CONTRAST PLACEHOLDER TEXT FOR ALL INPUT FIELDS */
    input::placeholder,
    textarea::placeholder,
    div[data-baseweb="input"] input::placeholder,
    div[data-baseweb="base-input"] input::placeholder,
    div[data-baseweb="textarea"] textarea::placeholder,
    .stTextInput input::placeholder,
    .stNumberInput input::placeholder,
    .stDateInput input::placeholder,
    .stTextArea textarea::placeholder {
        color: #f7d594 !important;
        -webkit-text-fill-color: #f7d594 !important;
        opacity: 0.95 !important;
        font-weight: 600 !important;
    }

    /* DROPDOWN OPTIONS & MENU LIST CONTRAST */
    div[data-baseweb="popover"] *,
    div[data-baseweb="menu"] *,
    div[role="option"] {
        background-color: #081d19 !important;
        color: #ffffff !important;
        font-weight: 600 !important;
    }
    div[role="option"]:hover, 
    div[role="option"][aria-selected="true"] {
        background-color: #16443c !important;
        color: #f7d594 !important;
    }

    /* CHECKBOX & RADIO LABELS HIGH CONTRAST */
    .stCheckbox label, .stCheckbox span, .stCheckbox p,
    .stRadio label, .stRadio span, .stRadio p {
        color: #ffffff !important;
        font-weight: 700 !important;
        font-size: 0.98rem !important;
        opacity: 1 !important;
    }

    /* STREAMLIT ALERT / NOTIFICATION BOX CONTRAST */
    .stAlert, div[data-testid="stAlert"] {
        background-color: #0c2b25 !important;
        border: 1px solid #e5a93c !important;
        border-radius: 12px !important;
        color: #ffffff !important;
    }
    .stAlert *, div[data-testid="stAlert"] * {
        color: #ffffff !important;
        font-weight: 700 !important;
    }

    /* RESPONSIVE MOBILE OPTIMIZATION (@media max-width 768px) */
    @media screen and (max-width: 768px) {
        .header-style {
            font-size: 1.85rem !important;
            letter-spacing: -0.5px !important;
        }
        .sub-header-style {
            font-size: 0.95rem !important;
            margin-bottom: 16px !important;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px !important;
            padding: 10px 14px !important;
            border-radius: 24px !important;
            overflow-x: auto !important;
            flex-wrap: nowrap !important;
            white-space: nowrap !important;
            -webkit-overflow-scrolling: touch !important;
        }
        .stTabs button[role="tab"],
        .stTabs [data-baseweb="tab"] {
            height: 42px !important;
            padding-left: 18px !important;
            padding-right: 18px !important;
            font-size: 0.88rem !important;
            margin: 0 2px !important;
        }
        .stButton > button,
        .stFormSubmitButton > button,
        button[data-testid="baseButton-secondary"],
        button[data-testid="baseButton-primary"] {
            font-size: 0.92rem !important;
            padding: 10px 16px !important;
            border-radius: 10px !important;
        }
        div[data-testid="stDataEditor"], div[data-testid="stDataFrame"] {
            font-size: 0.85rem !important;
        }
        div[data-testid="stExpanderDetails"], details[open] > div {
            padding: 10px !important;
        }
    }
</style>
""", unsafe_allow_html=True)

import json

# --- DISK PERSISTENCE ENGINE ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)
FINALIZED_DIR = os.path.join(DATA_DIR, "finalized_rosters")
os.makedirs(FINALIZED_DIR, exist_ok=True)
TIMESHEETS_DIR = os.path.join(DATA_DIR, "timesheets")
os.makedirs(TIMESHEETS_DIR, exist_ok=True)
TIMECARDS_FILE = os.path.join(DATA_DIR, "timecards.csv")
ANNOUNCEMENTS_FILE = os.path.join(DATA_DIR, "announcements.json")

BAKERY_LAT = -38.063557
BAKERY_LON = 145.455262

def calculate_haversine_distance(lat1, lon1, lat2=BAKERY_LAT, lon2=BAKERY_LON):
    try:
        if lat1 is None or lon1 is None or str(lat1).strip() == "" or str(lon1).strip() == "":
            return 0.0
        lat1, lon1, lat2, lon2 = map(math.radians, [float(lat1), float(lon1), float(lat2), float(lon2)])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        r = 6371000
        return round(c * r, 1)
    except:
        return 0.0

# Helpers for parsing times
def parse_time_to_decimal(time_str):
    try:
        time_str = str(time_str).strip().lower().replace(" ", "")
        is_pm = "pm" in time_str
        time_str = time_str.replace("am", "").replace("pm", "")
        if ":" in time_str:
            parts = time_str.split(":")
            hours = int(parts[0])
            minutes = int(parts[1])
        else:
            hours = int(time_str)
            minutes = 0
        if is_pm and hours < 12:
            hours += 12
        elif not is_pm and hours == 12:
            hours = 0
        return hours + minutes / 60.0
    except:
        return 0.0

def parse_shift_range(shift_str):
    if not shift_str or "unavailable" in str(shift_str).strip().lower() or str(shift_str).strip().lower() in ["off", "nan", ""]:
        return None
    try:
        parts = str(shift_str).split("-")
        start = parse_time_to_decimal(parts[0])
        end = parse_time_to_decimal(parts[1])
        duration = end - start if end > start else (24 - start) + end
        return start, end, duration
    except:
        return None

def is_overlapping_unavailability(unavail_str, shift_start, shift_end):
    unavail_str = str(unavail_str).strip().lower()
    if "all day" in unavail_str or "anytime" in unavail_str:
        return True
    if "before" in unavail_str:
        time_part = unavail_str.replace("before", "").strip()
        t = parse_time_to_decimal(time_part)
        return shift_start < t
    if "after" in unavail_str:
        time_part = unavail_str.replace("after", "").strip()
        t = parse_time_to_decimal(time_part)
        return shift_end > t
    r = parse_shift_range(unavail_str)
    if r:
        u_start, u_end, _ = r
        return max(shift_start, u_start) < min(shift_end, u_end)
    return False

# Robust Name Matcher
def find_matching_employee(raw_name, name_map):
    if not name_map or not isinstance(name_map, dict):
        return str(raw_name).strip() if raw_name else None
    raw_name_clean = str(raw_name).strip().lower()
    if not raw_name_clean or raw_name_clean == "nan":
        return None
    # 1. Exact match
    if raw_name_clean in name_map:
        return name_map[raw_name_clean]
    # 2. Prefix match (e.g. "Ainsley" matches "Ainsley Mactier", "Ana" matches "anastasia")
    for norm_name, display_name in name_map.items():
        if norm_name.startswith(raw_name_clean) or raw_name_clean.startswith(norm_name):
            return display_name
    # 3. First name matches first name
    for norm_name, display_name in name_map.items():
        first_raw = raw_name_clean.split()[0]
        first_norm = norm_name.split()[0]
        if first_raw == first_norm:
            return display_name
    return None

def get_week_start_date_str(dt_obj=None):
    if dt_obj is None:
        dt_obj = datetime.now().date()
    if isinstance(dt_obj, datetime):
        dt_obj = dt_obj.date()
    mon_dt = dt_obj - timedelta(days=dt_obj.weekday())
    return mon_dt.strftime("%d.%m.%Y")

def load_persisted_timecards():
    df_cards = pd.DataFrame()
    if os.path.exists(TIMECARDS_FILE):
        try:
            df_cards = pd.read_csv(TIMECARDS_FILE, dtype=str, keep_default_na=False)
        except:
            df_cards = pd.DataFrame()
            
    archived_rows = []
    if os.path.exists(TIMESHEETS_DIR):
        for f in os.listdir(TIMESHEETS_DIR):
            if f.endswith(".json"):
                fpath = os.path.join(TIMESHEETS_DIR, f)
                try:
                    with open(fpath, "r", encoding="utf-8") as file:
                        records = json.load(file)
                        if isinstance(records, list):
                            archived_rows.extend(records)
                except:
                    pass
    if archived_rows:
        df_arch = pd.DataFrame(archived_rows)
        if df_cards is not None and not df_cards.empty:
            df_cards = pd.concat([df_arch, df_cards], ignore_index=True)
        else:
            df_cards = df_arch
        if not df_cards.empty and "Record ID" in df_cards.columns:
            df_cards = df_cards.drop_duplicates(subset=["Record ID"], keep="last").reset_index(drop=True)
            
    return df_cards

def save_timecard_records(df_cards):
    if df_cards is None or not isinstance(df_cards, pd.DataFrame):
        return
    df_cards = df_cards.copy().astype(str)
    try:
        df_cards.to_csv(TIMECARDS_FILE, index=False)
    except:
        pass
        
    if "Date" in df_cards.columns:
        weeks_map = {}
        for idx, row in df_cards.iterrows():
            d_str = row.get("Date", "")
            d_obj = parse_date_robust(d_str)
            if d_obj:
                w_str = get_week_start_date_str(d_obj)
                if w_str not in weeks_map:
                    weeks_map[w_str] = []
                weeks_map[w_str].append(row.to_dict())
                
        # Synchronize and clean stale weekly archives in data/timesheets/
        if os.path.exists(TIMESHEETS_DIR):
            for f in os.listdir(TIMESHEETS_DIR):
                if f.startswith("timesheet_week_"):
                    w_str = f.replace("timesheet_week_", "").replace(".json", "").replace(".csv", "")
                    if w_str not in weeks_map:
                        try:
                            os.remove(os.path.join(TIMESHEETS_DIR, f))
                        except:
                            pass

        for w_str, week_rows in weeks_map.items():
            week_json_file = os.path.join(TIMESHEETS_DIR, f"timesheet_week_{w_str}.json")
            week_csv_file = os.path.join(TIMESHEETS_DIR, f"timesheet_week_{w_str}.csv")
            try:
                with open(week_json_file, "w", encoding="utf-8") as f_out:
                    json.dump(week_rows, f_out, indent=2)
                pd.DataFrame(week_rows).astype(str).to_csv(week_csv_file, index=False)
            except:
                pass

USER_PROFILES_FILE = os.path.join(DATA_DIR, "user_profiles.json")

DEFAULT_PROFILES = {
  "admin": {
    "username": "admin", "password": "admin123", "role": "Manager", "employee_name": "Bakery Manager",
    "profile": { "full_name": "Bakery Manager", "address": "", "home_phone": "", "mobile": "", "email": "manager@brumbys.com.au", "dob": "", "gender": "", "tfn": "", "store": "Brumby's Pakenham", "classification": "Full-Time", "commencement_date": "", "employment_level": "Store Manager", "super_fund": "", "super_policy": "", "super_address": "", "super_contact": "", "super_abn": "", "bank_name": "", "bank_branch": "", "bank_bsb": "", "bank_account": "", "account_name": "" }
  },
  "ainsley.mactier": {
    "username": "ainsley.mactier", "password": "TempPass123!", "role": "Employee", "employee_name": "Ainsley Mactier",
    "profile": { "full_name": "Ainsley Brenda Mactier", "address": "8 Knapton Ave, Beaconsfield Upper, Vic 3808", "home_phone": "0359192106", "mobile": "0479122444", "email": "ainsley.mac@outlook.com", "dob": "14/08/2006", "gender": "Female", "tfn": "520700", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "04/10/2021", "employment_level": "Service Staff" }
  },
  "elizabeth": {
    "username": "elizabeth", "password": "TempPass123!", "role": "Employee", "employee_name": "Elizabeth",
    "profile": { "full_name": "Elizabeth", "address": "", "home_phone": "", "mobile": "", "email": "", "dob": "30/07/2004", "gender": "Female", "tfn": "", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "03/06/2024", "employment_level": "Service Staff" }
  },
  "stella": {
    "username": "stella", "password": "TempPass123!", "role": "Employee", "employee_name": "Stella",
    "profile": { "full_name": "Stella", "address": "", "home_phone": "", "mobile": "", "email": "", "dob": "03/07/2007", "gender": "Female", "tfn": "", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "09/01/2024", "employment_level": "Service Staff" }
  },
  "aimi": {
    "username": "aimi", "password": "TempPass123!", "role": "Employee", "employee_name": "Aimi",
    "profile": { "full_name": "Aimi", "address": "", "home_phone": "", "mobile": "", "email": "", "dob": "10/11/2006", "gender": "Female", "tfn": "", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "01/10/2023", "employment_level": "Service Staff" }
  },
  "jude": {
    "username": "jude", "password": "TempPass123!", "role": "Employee", "employee_name": "Jude",
    "profile": { "full_name": "Jude", "address": "", "home_phone": "", "mobile": "", "email": "", "dob": "28/04/2011", "gender": "Male", "tfn": "", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "27/07/2026", "employment_level": "Service Staff" }
  },
  "aroha": {
    "username": "aroha", "password": "TempPass123!", "role": "Employee", "employee_name": "Aroha",
    "profile": { "full_name": "Aroha", "address": "", "home_phone": "", "mobile": "", "email": "", "dob": "24/05/2005", "gender": "Female", "tfn": "", "store": "Brumby's Pakenham", "classification": "part time", "commencement_date": "27/09/2021", "employment_level": "baker assitant" }
  },
  "robert": {
    "username": "robert", "password": "TempPass123!", "role": "Employee", "employee_name": "Robert",
    "profile": { "full_name": "Robert", "address": "", "home_phone": "", "mobile": "", "email": "", "dob": "19/02/2004", "gender": "Male", "tfn": "", "store": "Brumby's Pakenham", "classification": "part time", "commencement_date": "22/01/2024", "employment_level": "baker" }
  },
  "anastasia": {
    "username": "anastasia", "password": "TempPass123!", "role": "Employee", "employee_name": "Anastasia",
    "profile": { "full_name": "Anastasia", "dob": "02/03/2000", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "04/10/2021", "employment_level": "Service Staff" }
  },
  "jack": {
    "username": "jack", "password": "TempPass123!", "role": "Employee", "employee_name": "Jack",
    "profile": { "full_name": "Jack", "dob": "28/04/2011", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "27/07/2026", "employment_level": "Service Staff" }
  },
  "jane": {
    "username": "jane", "password": "TempPass123!", "role": "Employee", "employee_name": "Jane",
    "profile": { "full_name": "Jane", "store": "Brumby's Pakenham", "classification": "owner", "employment_level": "Service Staff" }
  },
  "amy": {
    "username": "amy", "password": "TempPass123!", "role": "Employee", "employee_name": "Amy",
    "profile": { "full_name": "Amy", "dob": "27/02/2010", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "25/05/2026", "employment_level": "Service Staff" }
  },
  "viet": {
    "username": "viet", "password": "TempPass123!", "role": "Employee", "employee_name": "Viet",
    "profile": { "full_name": "Viet", "store": "Brumby's Pakenham", "classification": "owner", "employment_level": "baker" }
  },
  "esther.amataiti": {
    "username": "esther.amataiti", "password": "TempPass123!", "role": "Employee", "employee_name": "Esther Amataiti",
    "profile": { "full_name": "Esther Amataiti", "dob": "20/09/2001", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "20/09/2021", "employment_level": "Service Staff" }
  },
  "olivia": {
    "username": "olivia", "password": "TempPass123!", "role": "Employee", "employee_name": "Olivia",
    "profile": { "full_name": "Olivia", "dob": "15/01/2007", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "27/05/2024", "employment_level": "Service Staff" }
  },
  "violet": {
    "username": "violet", "password": "TempPass123!", "role": "Employee", "employee_name": "Violet",
    "profile": { "full_name": "Violet", "dob": "27/02/2010", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "25/05/2026", "employment_level": "Service Staff" }
  },
  "shaelyn": {
    "username": "shaelyn", "password": "TempPass123!", "role": "Employee", "employee_name": "Shaelyn",
    "profile": { "full_name": "Shaelyn", "dob": "01/08/2011", "store": "Brumby's Pakenham", "classification": "casual", "commencement_date": "01/08/2026", "employment_level": "Service Staff" }
  }
}

def load_user_profiles():
    fs_profiles = firestore_load_profiles()
    if fs_profiles and isinstance(fs_profiles, dict) and "admin" in fs_profiles:
        return fs_profiles

    profiles = {}
    if os.path.exists(USER_PROFILES_FILE):
        try:
            with open(USER_PROFILES_FILE, "r", encoding="utf-8") as f:
                profiles = json.load(f)
        except Exception:
            profiles = {}
            
    if not profiles or "admin" not in profiles:
        profiles = copy.deepcopy(DEFAULT_PROFILES)
        try:
            with open(USER_PROFILES_FILE, "w", encoding="utf-8") as f:
                json.dump(profiles, f, indent=2)
        except:
            pass
                
    # Always ensure default accounts exist
    for d_key, d_val in DEFAULT_PROFILES.items():
        if d_key not in profiles:
            profiles[d_key] = copy.deepcopy(d_val)
            
    return profiles

def get_active_user_profiles():
    return load_user_profiles()

def save_user_profiles(profiles):
    firestore_save_profiles(profiles)
    try:
        with open(USER_PROFILES_FILE, "w", encoding="utf-8") as f:
            json.dump(profiles, f, indent=2)
    except Exception as e:
        st.error(f"Error saving user profiles: {e}")

# --- SMTP EMAIL CONFIGURATION ENGINE ---
SMTP_CONFIG_FILE = os.path.join(DATA_DIR, "smtp_config.json")

def load_smtp_config():
    default_config = {
        "smtp_server": "smtp.gmail.com",
        "smtp_port": 587,
        "sender_email": "Brumby.pakenham@gmail.com",
        "sender_password": "",
        "portal_url": "https://weekly-roster-generator.streamlit.app",
        "sender_name": "Bakery Manager",
        "notification_recipients": "quietsong2006@yahoo.com, uyentrinhtran2309@gmail.com"
    }

    cfg = default_config.copy()
    if os.path.exists(SMTP_CONFIG_FILE):
        try:
            with open(SMTP_CONFIG_FILE, "r", encoding="utf-8") as f:
                disk_cfg = json.load(f)
                for k, v in disk_cfg.items():
                    if v or k not in cfg:
                        cfg[k] = v
        except:
            pass

    # Check Streamlit secrets or env vars fallback if password is empty
    if not cfg.get("sender_password"):
        try:
            if hasattr(st, "secrets") and "SENDER_PASSWORD" in st.secrets:
                cfg["sender_password"] = str(st.secrets["SENDER_PASSWORD"]).strip()
            elif hasattr(st, "secrets") and "SMTP_PASSWORD" in st.secrets:
                cfg["sender_password"] = str(st.secrets["SMTP_PASSWORD"]).strip()
            elif os.environ.get("SENDER_PASSWORD"):
                cfg["sender_password"] = os.environ.get("SENDER_PASSWORD").strip()
        except:
            pass

    # Check session memory shield
    if not cfg.get("sender_password") and st.session_state.get("memory_smtp_password"):
        cfg["sender_password"] = st.session_state["memory_smtp_password"]

    return cfg

def save_smtp_config(cfg):
    try:
        # Load existing disk config to prevent erasing existing password with an empty string
        existing_cfg = {}
        if os.path.exists(SMTP_CONFIG_FILE):
            try:
                with open(SMTP_CONFIG_FILE, "r", encoding="utf-8") as f_in:
                    existing_cfg = json.load(f_in)
            except:
                pass
        
        # If new password is empty string, preserve existing password from disk or session memory
        new_pass = str(cfg.get("sender_password", "")).strip()
        if not new_pass:
            old_pass = str(existing_cfg.get("sender_password", "")).strip()
            if not old_pass:
                old_pass = str(st.session_state.get("memory_smtp_password", "")).strip()
            if old_pass:
                cfg["sender_password"] = old_pass
        else:
            st.session_state["memory_smtp_password"] = new_pass

        with open(SMTP_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2)
    except:
        pass

def build_welcome_email_content(emp_name, username, temp_password, portal_url="", sender_name=""):
    cfg = load_smtp_config()
    if not portal_url:
        portal_url = cfg.get("portal_url", "https://weekly-roster-generator.streamlit.app")
    if not sender_name:
        sender_name = cfg.get("sender_name", "Bakery Manager")
        
    subject = "Welcome to Brumby's Pakenham! 🥐 — Your Account Setup & Portal Login"
    body = f"""Subject: {subject}

Hi {emp_name},

Welcome to the team at Brumby's Pakenham! 🥐

We use an online portal for managing shift rostering, availability, and confidential employee onboarding. Please set up your account by completing these quick steps:

1. Log In to the Portal

Website: {portal_url}
Username: {username}
Temporary Password: {temp_password}

2. Complete Your Setup Tasks
Once logged in, please complete the following tabs:

📋 Personal Information Form: Fill out your contact details, Tax File Number (TFN), bank account details (for payroll), and superannuation fund.
📅 My Availability & Constraints: Select the days/time windows you are available to work.
🔑 Change Password: Click Change My Password in the left sidebar to update your temporary password to a secure personal one.

If you run into any issues, please feel free to reach out.

Welcome aboard!
{sender_name}
Brumby's Pakenham"""
    return subject, body

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def send_welcome_email_smtp(recipient_email, emp_name, username, temp_password, portal_url="", sender_name=""):
    cfg = load_smtp_config()
    smtp_server = cfg.get("smtp_server", "smtp.gmail.com")
    smtp_port = int(cfg.get("smtp_port", 587))
    sender_email = cfg.get("sender_email", "").strip()
    sender_pass = cfg.get("sender_password", "").strip()
    
    if not portal_url:
        portal_url = cfg.get("portal_url", "https://weekly-roster-generator.streamlit.app")
    if not sender_name:
        sender_name = cfg.get("sender_name", "Bakery Manager")

    subject, body = build_welcome_email_content(emp_name, username, temp_password, portal_url, sender_name)
    
    if not recipient_email or not recipient_email.strip():
        return False, "Recipient email is blank."
        
    if not sender_email or not sender_pass:
        return False, "SMTP credentials not yet configured in Email Settings."

    try:
        msg = MIMEMultipart()
        msg["From"] = f"{sender_name} <{sender_email}>"
        msg["To"] = recipient_email.strip()
        msg["Subject"] = subject
        
        email_text_content = body.split("\n\n", 1)[-1] if "\n\n" in body else body
        msg.attach(MIMEText(email_text_content, "plain", "utf-8"))

        server = smtplib.SMTP(smtp_server, smtp_port, timeout=10)
        server.starttls()
        server.login(sender_email, sender_pass)
        server.send_message(msg)
        server.quit()
        return True, f"Welcome Email sent to {recipient_email}!"
    except smtplib.SMTPAuthenticationError:
        return False, "Gmail authentication failed (Error 535). Gmail requires a 16-character 'Google App Password' instead of your normal account password."
    except Exception as e:
        return False, f"SMTP Error: {e}"

def send_test_email_smtp(recipient_email):
    cfg = load_smtp_config()
    smtp_server = cfg.get("smtp_server", "smtp.gmail.com")
    smtp_port = int(cfg.get("smtp_port", 587))
    sender_email = cfg.get("sender_email", "").strip()
    sender_pass = cfg.get("sender_password", "").strip()
    sender_name = cfg.get("sender_name", "Bakery Manager")
    
    if not recipient_email or not recipient_email.strip():
        return False, "Sender / recipient email address is blank."
    if not sender_email or not sender_pass:
        return False, "Please enter both Sender Email Address and App Password."
        
    try:
        msg = MIMEMultipart()
        msg["From"] = f"{sender_name} <{sender_email}>"
        msg["To"] = recipient_email.strip()
        msg["Subject"] = "🧪 Test Email — Brumby's Bakery Portal"
        
        test_body = f"Hello!\n\nThis is a test email sent from Brumby's Bakery Portal to verify your SMTP configuration.\n\nSender: {sender_email}\nSMTP Host: {smtp_server}:{smtp_port}\n\nIf you received this message, your welcome email setup is working perfectly! 🎉"
        msg.attach(MIMEText(test_body, "plain", "utf-8"))

        server = smtplib.SMTP(smtp_server, smtp_port, timeout=10)
        server.starttls()
        server.login(sender_email, sender_pass)
        server.send_message(msg)
        server.quit()
        return True, f"✅ Test email delivered to {recipient_email}!"
    except smtplib.SMTPAuthenticationError:
        return False, "❌ Gmail Authentication Failed (Error 535). Gmail requires a 16-character 'Google App Password' generated in your Google Account Security settings."
    except Exception as e:
        return False, f"❌ Connection Error: {e}"

def send_availability_notification_email_smtp(emp_name, action_type, details_str):

    cfg = load_smtp_config()
    smtp_server = cfg.get("smtp_server", "smtp.gmail.com")
    smtp_port = int(cfg.get("smtp_port", 587))
    sender_email = cfg.get("sender_email", "").strip()
    sender_pass = cfg.get("sender_password", "").strip()
    sender_name = cfg.get("sender_name", "Bakery Manager")
    
    recipients_raw = cfg.get("notification_recipients", "quietsong2006@yahoo.com, uyentrinhtran2309@gmail.com")
    recipient_list = [r.strip() for r in re.split(r'[,;]', str(recipients_raw)) if r.strip()]
    if not recipient_list:
        recipient_list = ["quietsong2006@yahoo.com", "uyentrinhtran2309@gmail.com"]

    if not sender_email or not sender_pass:
        return False, "SMTP credentials not configured."

    subject = f"📅 Staff Availability Alert: {emp_name} — {action_type}"
    now_str = datetime.now().strftime("%d/%m/%Y at %I:%M %p")

    body = f"""Subject: {subject}

Staff Availability Notification Alert 🥐
Brumby's Pakenham

Employee Name: {emp_name}
Action: {action_type}
Timestamp: {now_str}

Details of Availability Change:
{details_str}

This is an automated notification sent from the Brumby's Bakery Portal to keep managers informed of staff availability updates.
"""

    try:
        msg = MIMEMultipart()
        msg["From"] = f"{sender_name} <{sender_email}>"
        msg["To"] = ", ".join(recipient_list)
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))

        server = smtplib.SMTP(smtp_server, smtp_port, timeout=10)
        server.starttls()
        server.login(sender_email, sender_pass)
        server.send_message(msg)
        server.quit()
        return True, f"Notification sent to {', '.join(recipient_list)}"
    except Exception as e:
        return False, f"SMTP Notification Error: {e}"

def load_announcements():
    default_announcements = [
        {
            "id": "ANN_1001",
            "title": "Welcome to Brumby's Bakery Pakenham Portal",
            "content": "Welcome team! Please review your weekly scheduled shifts, award entitlements, and use the mobile timeclock when on shift.",
            "author": "Viet (Store Owner)",
            "date": datetime.now().strftime("%d/%m/%Y"),
            "priority": "Normal"
        }
    ]
    if os.path.exists(ANNOUNCEMENTS_FILE):
        try:
            with open(ANNOUNCEMENTS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
        except:
            pass
    return default_announcements

def save_announcements(announcements):
    try:
        with open(ANNOUNCEMENTS_FILE, "w", encoding="utf-8") as f:
            json.dump(announcements, f, indent=2, ensure_ascii=False)
        return True
    except:
        return False

def send_announcement_broadcast_smtp(title, content, author="Store Management"):
    cfg = load_smtp_config()
    sender_email = cfg.get("sender_email")
    sender_password = cfg.get("sender_password")
    smtp_server = cfg.get("smtp_server", "smtp.gmail.com")
    smtp_port = int(cfg.get("smtp_port", 587))
    portal_url = cfg.get("portal_url", "https://weekly-roster-generator.streamlit.app")

    if not sender_email or not sender_password:
        return False, "SMTP Email credentials not configured."

    recipient_emails = set()
    for u_key, u_data in user_profiles.items():
        em = u_data.get("profile", {}).get("email", "").strip()
        if em and "@" in em:
            recipient_emails.add(em)

    notif_recips = cfg.get("notification_recipients", "")
    if notif_recips:
        for r_addr in notif_recips.split(","):
            r_clean = r_addr.strip()
            if r_clean and "@" in r_clean:
                recipient_emails.add(r_clean)

    if not recipient_emails:
        return False, "No employee recipient email addresses found."

    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; background-color: #f4f6f8; padding: 20px; margin: 0;">
        <div style="max-width: 600px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; border: 2px solid #e5a93c; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
            <div style="background: linear-gradient(135deg, #0e2b26 0%, #1a4d43 100%); padding: 20px; text-align: center; color: #ffffff;">
                <h2 style="color: #e5a93c; margin: 0; font-size: 1.5rem;">📢 Store Announcement Broadcast</h2>
                <p style="margin: 6px 0 0 0; color: #d0e6df; font-size: 0.95rem;">Brumby's Bakery Pakenham Roster & Attendance Portal</p>
            </div>
            <div style="padding: 24px; color: #2d3748; font-size: 1rem; line-height: 1.6;">
                <h3 style="color: #0e2b26; margin-top: 0;">{title}</h3>
                <p style="background: #f7fafc; border-left: 4px solid #e5a93c; padding: 12px 16px; margin: 15px 0; border-radius: 4px; font-size: 0.98rem; color: #2d3748;">
                    {content}
                </p>
                <p style="font-size: 0.85rem; color: #718096; margin-top: 20px;">
                    <b>Posted By:</b> {author}<br>
                    <b>Date:</b> {datetime.now().strftime("%d/%m/%Y %I:%M %p")}
                </p>
                <div style="text-align: center; margin-top: 25px;">
                    <a href="{portal_url}" style="background: linear-gradient(135deg, #e5a93c 0%, #c0841d 100%); color: #000000; text-decoration: none; padding: 12px 24px; font-weight: 800; border-radius: 8px; display: inline-block;">
                        🌐 Open Roster Portal
                    </a>
                </div>
            </div>
            <div style="background: #edf2f7; padding: 12px; text-align: center; font-size: 0.8rem; color: #718096;">
                Brumby's Bakery Pakenham • General Retail Industry Award 2020 Compliance
            </div>
        </div>
    </body>
    </html>
    """

    try:
        server = smtplib.SMTP(smtp_server, smtp_port, timeout=10)
        server.starttls()
        server.login(sender_email, sender_password)

        sent_count = 0
        for recip in recipient_emails:
            msg = MIMEText(html_body, "html", "utf-8")
            msg["Subject"] = f"📢 [Brumby's Bakery] Store Announcement: {title}"
            msg["From"] = f"Brumby's Bakery Pakenham <{sender_email}>"
            msg["To"] = recip
            server.sendmail(sender_email, [recip], msg.as_string())
            sent_count += 1

        server.quit()
        return True, f"Successfully broadcast announcement email to {sent_count} recipient(s)!"
    except Exception as e:
        return False, f"Failed to send email broadcast: {str(e)}"

def load_persisted_df(filename, default_df=None):
    collection_name = filename.replace(".csv", "")
    fs_df = firestore_load_df(collection_name)
    if fs_df is not None and not fs_df.empty:
        return fs_df

    path = os.path.join(DATA_DIR, filename)
    if os.path.exists(path):
        try:
            df = pd.read_csv(path, dtype=str, keep_default_na=False)
            if df is not None:
                return df
        except:
            pass
            
    # Auto-recovery fallback: Check if master Excel document exists in data/Document/
    doc_mapping = {
        "employees.csv": "EMPLOYEE LIST.xlsx",
        "unavailability.csv": "unavailability list.xlsx",
        "requirements.csv": "Daily Shift personel requirement.xlsx",
        "fixed.csv": "Roster fixed - dont change.xlsx"
    }
    excel_name = doc_mapping.get(filename)
    if excel_name:
        excel_path = os.path.join(DATA_DIR, "Document", excel_name)
        if os.path.exists(excel_path):
            try:
                excel_df = read_excel_robust(excel_path)
                if excel_df is not None and not excel_df.empty:
                    try:
                        excel_df.astype(str).to_csv(path, index=False)
                    except:
                        pass
                    return excel_df
            except:
                pass
                
    return default_df

def clear_unavailability_widget_cache():
    for k in list(st.session_state.keys()):
        if k.startswith("edit_unavail_month_"):
            del st.session_state[k]

def save_persisted_df(df, filename):
    collection_name = filename.replace(".csv", "")
    if df is not None:
        firestore_save_df(collection_name, df)
    path = os.path.join(DATA_DIR, filename)
    try:
        df.astype(str).to_csv(path, index=False)
        if filename == "unavailability.csv":
            clear_unavailability_widget_cache()
    except:
        pass

def build_roster_excel_bytes(edited_final_df, start_date):
    if edited_final_df is not None and not edited_final_df.empty:
        edited_final_df = strip_daily_gross_row(edited_final_df)
        edited_final_df = sort_dataframe_by_team_and_age(edited_final_df)
        
    if isinstance(start_date, str):
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        except:
            start_date = datetime.now().date()
            
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Roster"

    # Header styling
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "BRUMBY'S PAKENHAM - WEEKLY STAFF ROSTER"
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.fill = title_fill
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    sub_cell = ws.cell(row=2, column=1)
    end_date = start_date + timedelta(days=6)
    sub_cell.value = f"Week Period: Monday {start_date.strftime('%d/%m/%Y')} to Sunday {end_date.strftime('%d/%m/%Y')}"
    sub_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    sub_cell.font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 10

    tbl_hdr_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    tbl_hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(edited_final_df.columns, 1):
        c = ws.cell(row=4, column=col_idx)
        c.value = col_name
        c.fill = tbl_hdr_fill
        c.font = tbl_hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
            
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # Write roster data rows
    row_start = 5
    for r_idx, row_data in enumerate(edited_final_df.itertuples(index=False), start=row_start):
        for c_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=r_idx, column=c_idx)
            val_str = "" if pd.isna(val) else str(val).strip()
            
            # Leave unavailable, off, and none cells completely clean & blank in exported Excel file
            if not val_str or val_str.lower() in ["off", "none", "nan", "null", "unavailable", " unavailable"] or val_str.lower().startswith("unavail"):
                val_str = ""
            
            c.value = val_str
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            
            if val_str:
                if c_idx == 1:
                    c.font = Font(name="Calibri", size=11, bold=True)
                    c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.font = Font(name="Calibri", size=10)

    last_roster_row = row_start + len(edited_final_df) - 1

    # --- EMPLOYEE SHIFT BREAK ENTITLEMENTS GUIDE ---
    guide_row_1 = last_roster_row + 3
    
    ws.merge_cells(start_row=guide_row_1, start_column=1, end_row=guide_row_1, end_column=8)
    g_title = ws.cell(row=guide_row_1, column=1)
    g_title.value = "☕ EMPLOYEE SHIFT BREAK ENTITLEMENTS GUIDE (General Retail Industry Award 2020)"
    g_title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    g_title.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    g_title.alignment = Alignment(horizontal="center", vertical="center")

    hdr_row = guide_row_1 + 1
    headers_break = [
        (1, 2, "Shift Duration"),
        (3, 4, "Paid Rest Break (10 min)"),
        (5, 6, "Unpaid Meal Break (30 min)"),
        (7, 8, "Total Break Entitlement")
    ]
    for start_col, end_col, text in headers_break:
        ws.merge_cells(start_row=hdr_row, start_column=start_col, end_row=hdr_row, end_column=end_col)
        hc = ws.cell(row=hdr_row, column=start_col)
        hc.value = text
        hc.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        hc.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        hc.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(start_col, end_col + 1):
            ws.cell(row=hdr_row, column=col).border = thin_border

    break_data = [
        ("Less than 4 hours", "No break", "No break", "No breaks required"),
        ("4 hours up to 5 hours", "1 x 10-minute rest break", "No meal break", "10 min Paid break"),
        ("5 hours up to 7 hours", "1 x 10-minute rest break", "1 x 30-minute meal break", "10m Paid + 30m Unpaid meal"),
        ("7 hours up to 10 hours", "2 x 10-minute rest breaks", "1 x 30-minute meal break", "20m Paid + 30m Unpaid meal"),
        ("10 hours or more", "2 x 10-minute rest breaks", "2 x 30-minute meal breaks", "20m Paid + 60m Unpaid meals"),
    ]

    curr_b_row = hdr_row + 1
    for d1, d2, d3, d4 in break_data:
        row_cols = [(1, 2, d1), (3, 4, d2), (5, 6, d3), (7, 8, d4)]
        for start_col, end_col, text in row_cols:
            ws.merge_cells(start_row=curr_b_row, start_column=start_col, end_row=curr_b_row, end_column=end_col)
            cell = ws.cell(row=curr_b_row, column=start_col)
            cell.value = text
            cell.font = Font(name="Calibri", size=9.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(start_col, end_col + 1):
                ws.cell(row=curr_b_row, column=col).border = thin_border
        curr_b_row += 1

    note_row = curr_b_row + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row+2, end_column=8)
    n_cell = ws.cell(row=note_row, column=1)
    n_cell.value = (
        "📌 Key Break Rules for Bakery Staff:\n"
        "• Rest breaks (10 mins) are paid. Meal breaks (30 mins) are unpaid.\n"
        "• Breaks cannot be taken in the first or last hour of work.\n"
        "• An unpaid meal break must be taken no later than after 5 hours of continuous work."
    )
    n_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    n_cell.font = Font(name="Calibri", size=9, italic=True, color="333333")
    n_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in range(note_row, note_row + 3):
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = thin_border

    # Dedicated 2nd Sheet Tab: Break Entitlements Guide
    ws2 = wb.create_sheet(title="Break Entitlements Guide")
    ws2.merge_cells("A1:D1")
    t2 = ws2.cell(row=1, column=1)
    t2.value = "BRUMBY'S PAKENHAM - EMPLOYEE BREAK ENTITLEMENTS GUIDE"
    t2.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    t2.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t2.alignment = Alignment(horizontal="center", vertical="center")

    headers_s2 = ["Shift Duration", "Paid Rest Break (10 min)", "Unpaid Meal Break (30 min)", "Total Break Entitlement"]
    for col_idx, h_text in enumerate(headers_s2, 1):
        c2 = ws2.cell(row=3, column=col_idx)
        c2.value = h_text
        c2.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        c2.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = thin_border

    for r_idx, (d1, d2, d3, d4) in enumerate(break_data, start=4):
        vals = [d1, d2, d3, d4]
        for c_idx, val in enumerate(vals, start=1):
            c = ws2.cell(row=r_idx, column=c_idx)
            c.value = val
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 32

    # Auto-adjust column widths for Sheet 1 (based only on roster table cells, ignoring merged title & note cells)
    for col_idx in range(1, len(edited_final_df.columns) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(4, last_roster_row + 1))
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = max(max_len + 5, 20)
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    # Dedicated 3rd Sheet Tab: Hour Rate Breakdown
    ws3 = wb.create_sheet(title="Hour Rate Breakdown")
    hb_df = calculate_weekly_hour_rate_breakdown(edited_final_df)
    if not hb_df.empty:
        last_col_letter = openpyxl.utils.get_column_letter(len(hb_df.columns))
        ws3.merge_cells(f"A1:{last_col_letter}1")
        t3 = ws3.cell(row=1, column=1)
        date_label_str = start_date.strftime("%d.%m.%Y") if hasattr(start_date, 'strftime') else str(start_date)
        t3.value = f"BRUMBY'S PAKENHAM - HOUR RATE BREAKDOWN ({date_label_str})"
        t3.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        t3.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        t3.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(hb_df.columns, 1):
            c3 = ws3.cell(row=3, column=col_idx)
            c3.value = col_name
            c3.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
            c3.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c3.alignment = Alignment(horizontal="center", vertical="center")
            c3.border = thin_border

        for r_idx, row_data in enumerate(hb_df.itertuples(index=False), start=4):
            is_summary = (r_idx == len(hb_df) + 3)
            for c_idx, val in enumerate(row_data, start=1):
                c = ws3.cell(row=r_idx, column=c_idx)
                c.value = val
                c.alignment = Alignment(horizontal="center", vertical="center") if c_idx > 1 else Alignment(horizontal="left", vertical="center")
                c.font = Font(name="Calibri", size=10, bold=is_summary)
                c.border = thin_border
                if is_summary:
                    c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_idx, col_name in enumerate(hb_df.columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_l = max(len(str(ws3.cell(row=r, column=col_idx).value or '')) for r in range(3, len(hb_df) + 4))
            ws3.column_dimensions[col_letter].width = max(max_l + 5, 18)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    return excel_buffer.getvalue()

def save_finalized_roster(df, start_date):
    if isinstance(start_date, str):
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        except:
            start_date = datetime.now().date()
            
    date_str = start_date.strftime("%Y-%m-%d")
    date_label = start_date.strftime("%d.%m.%Y")
    
    csv_filename = f"Roster_{date_str}.csv"
    xlsx_filename = f"Team_Roster_{date_label}.xlsx"
    
    excel_bytes = build_roster_excel_bytes(df, start_date)
    
    csv_path = os.path.join(FINALIZED_DIR, csv_filename)
    xlsx_path = os.path.join(FINALIZED_DIR, xlsx_filename)
    
    df.astype(str).to_csv(csv_path, index=False)
    with open(xlsx_path, "wb") as f:
        f.write(excel_bytes)
        
    return date_str, xlsx_filename, excel_bytes

def extract_date_from_filename(filename):
    if not filename:
        return None
    # Ignore sample files like 00.00.2025
    if "00.00." in filename or ("SAMPLE" in filename.upper() and "ROSTER" not in filename.upper()):
        return None

    month_map = {
        'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
        'apr': 4, 'april': 4, 'may': 5, 'june': 6, 'jun': 6,
        'jul': 7, 'july': 7, 'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9,
        'oct': 10, 'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12
    }
        
    # Match YYYY-MM-DD
    m = re.search(r'(20\d{2})[-._](\d{1,2})[-._](\d{1,2})', filename)
    if m:
        y, m_str, d = m.groups()
        try:
            dt = datetime(int(y), int(m_str), int(d)).date()
            if 2020 <= dt.year <= 2030:
                return dt
        except:
            pass

    # Match DD.MM.YYYY or DD-MM-YYYY or DD_MM_YYYY
    m = re.search(r'(\d{1,2})[-._](\d{1,2})[-._](20\d{2})', filename)
    if m:
        d, m_str, y = m.groups()
        try:
            dt = datetime(int(y), int(m_str), int(d)).date()
            if 2020 <= dt.year <= 2030:
                return dt
        except:
            pass

    # Match DD-MM-YY or DD.MM.YY
    m = re.search(r'(\d{1,2})[-._](\d{1,2})[-._](\d{2})', filename)
    if m:
        d, m_str, y_short = m.groups()
        y = 2000 + int(y_short)
        try:
            dt = datetime(y, int(m_str), int(d)).date()
            if 2020 <= dt.year <= 2030:
                return dt
        except:
            pass

    # Match DD Month YYYY or DD-Month-YYYY (e.g. 10 August 2026, 10-Aug-2026)
    m = re.search(r'(\d{1,2})[-._\s]+([a-zA-Z]{3,9})[-._\s]+(20\d{2})', filename)
    if m:
        d, m_name, y = m.groups()
        m_num = month_map.get(m_name.lower())
        if m_num:
            try:
                dt = datetime(int(y), m_num, int(d)).date()
                if 2020 <= dt.year <= 2030:
                    return dt
            except:
                pass

    # Match Month DD YYYY or Month-DD-YYYY (e.g. August 10 2026, Aug-10-2026)
    m = re.search(r'([a-zA-Z]{3,9})[-._\s]+(\d{1,2})[-._\s]+(20\d{2})', filename)
    if m:
        m_name, d, y = m.groups()
        m_num = month_map.get(m_name.lower())
        if m_num:
            try:
                dt = datetime(int(y), m_num, int(d)).date()
                if 2020 <= dt.year <= 2030:
                    return dt
            except:
                pass

    return None

def find_all_old_roster_files():
    found_files = []
    seen = set()
    
    known_candidates = [
        os.path.join(BASE_DIR, "Team Roster 10.08.2026.xlsx"),
        os.path.join(DATA_DIR, "demo doc", "reference roster", "Team_Roster_24.08.2026 (1).xlsx"),
        os.path.join(DATA_DIR, "demo doc", "reference roster", "Roster 27.07.2026 Pakenham payroll_.xlsx"),
        os.path.join(DATA_DIR, "demo doc", "reference roster", "Roster 20.07.2026 Pakenham payroll_.xlsx"),
        os.path.join(DATA_DIR, "demo doc", "reference roster", "Roster 13.07.2026 Pakenham payroll_.xlsx"),
        os.path.join(DATA_DIR, "demo doc", "reference roster", "Roster 29.06.2026 Pakenham payroll_.xlsx"),
        os.path.join(DATA_DIR, "demo doc", "weekly roster demo.xlsx"),
    ]
    for kp in known_candidates:
        if os.path.exists(kp) and kp not in seen:
            seen.add(kp)
            found_files.append(kp)

    search_dirs = [
        BASE_DIR,
        DATA_DIR,
        os.path.join(DATA_DIR, "demo doc"),
        os.path.join(DATA_DIR, "demo doc", "reference roster"),
        os.path.expanduser(r"~\Downloads"),
        os.path.expanduser(r"~\OneDrive\Desktop\wage"),
        os.path.expanduser(r"~\OneDrive\Desktop\Happy Roli\New folder\roster"),
        os.path.expanduser(r"~\OneDrive\Desktop\roli sugarhouse\roster"),
        os.path.expanduser(r"~\OneDrive\Desktop"),
    ]
    
    for sdir in search_dirs:
        if os.path.exists(sdir):
            try:
                for root, dirs, files in os.walk(sdir):
                    if "finalized_rosters" in root:
                        continue
                    for f in files:
                        ext = os.path.splitext(f)[1].lower()
                        if ext in [".xlsx", ".csv"] and not f.startswith("~$"):
                            f_lower = f.lower()
                            if any(k in f_lower for k in ["roster", "payroll", "pakenham", "team"]):
                                full_p = os.path.abspath(os.path.join(root, f))
                                if full_p not in seen:
                                    seen.add(full_p)
                                    found_files.append(full_p)
            except Exception:
                pass

    return found_files

def auto_import_reference_rosters(force_scan=False):
    if not os.path.exists(FINALIZED_DIR):
        os.makedirs(FINALIZED_DIR, exist_ok=True)
        
    imported_count = 0
    candidate_paths = find_all_old_roster_files()
    
    for file_path in candidate_paths:
        fname = os.path.basename(file_path)
        dt = extract_date_from_filename(fname)
        if dt is None:
            continue
            
        date_str = dt.strftime("%Y-%m-%d")
        csv_filename = f"Roster_{date_str}.csv"
        csv_path = os.path.join(FINALIZED_DIR, csv_filename)
        
        # Preserve existing user-saved rosters: only import if roster file does not exist or is empty
        if not os.path.exists(csv_path) or os.path.getsize(csv_path) == 0:
            try:
                if file_path.endswith(".csv"):
                    df = pd.read_csv(file_path, dtype=str, keep_default_na=False)
                else:
                    df = read_excel_robust(file_path)
                    
                if df is not None and not df.empty:
                    name_col = find_column(df, ["name", "employee", "staff"], "NAME")
                    if name_col in df.columns:
                        df = df.rename(columns={name_col: "NAME"})
                        
                    date_label = dt.strftime("%d.%m.%Y")
                    xlsx_filename = f"Team_Roster_{date_label}.xlsx"
                    xlsx_path = os.path.join(FINALIZED_DIR, xlsx_filename)
                    
                    df.astype(str).to_csv(csv_path, index=False)
                    excel_bytes = build_roster_excel_bytes(df, dt)
                    with open(xlsx_path, "wb") as f:
                        f.write(excel_bytes)
                    imported_count += 1
            except Exception:
                pass
    return imported_count

def list_finalized_rosters():
    if not os.path.exists(FINALIZED_DIR):
        os.makedirs(FINALIZED_DIR, exist_ok=True)
        
    auto_import_reference_rosters()
    
    files = [f for f in os.listdir(FINALIZED_DIR) if f.endswith(".csv") and f.startswith("Roster_")]
    files.sort(reverse=True)
    results = []
    for f in files:
        raw_date = f.replace("Roster_", "").replace(".csv", "")
        try:
            dt = datetime.strptime(raw_date, "%Y-%m-%d").date()
            end_dt = dt + timedelta(days=6)
            label = f"Week of {dt.strftime('%d/%m/%Y')} (Mon {dt.strftime('%d/%m')} - Sun {end_dt.strftime('%d/%m')})"
        except:
            dt = None
            label = f"Roster {raw_date}"
        results.append({"csv_filename": f, "date_str": raw_date, "start_date": dt, "label": label})
    return results

def clean_roster_unavailability_display(df):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df

    df_out = df.copy()
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    days_in_df = [d for d in days_of_week if d in df_out.columns]
    
    for idx, row in df_out.iterrows():
        for day in days_in_df:
            val = str(row.get(day, "")).strip()
            val_lower = val.lower()
            if not val or val_lower in ["off", "none", "nan", "null", "unavailable", " unavailable"] or val_lower.startswith("unavail") or val.startswith("🚫"):
                df_out.at[idx, day] = ""

    return df_out

def format_roster_with_unavailability_badges(df, unavail_data=None):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df

    df_out = df.copy()
    emp_col = find_column(df_out, ["employee", "name", "staff", "staff name"], df_out.columns[0])
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    days_in_df = [d for d in days_of_week if d in df_out.columns]
    
    if not days_in_df:
        return df_out

    if unavail_data is None or (isinstance(unavail_data, pd.DataFrame) and unavail_data.empty):
        unavail_data = st.session_state.get("manual_unavailability", None)
        if unavail_data is None or (isinstance(unavail_data, pd.DataFrame) and unavail_data.empty):
            unavail_path = os.path.join(DATA_DIR, "unavailability.csv")
            if os.path.exists(unavail_path):
                try:
                    unavail_data = pd.read_csv(unavail_path, dtype=str, keep_default_na=False)
                except:
                    unavail_data = pd.DataFrame()

    # Build name_map for robust employee name matching across table rows
    roster_emp_names = [str(r.get(emp_col, "")).strip() for _, r in df_out.iterrows() if str(r.get(emp_col, "")).strip()]
    name_map = {n.lower(): n for n in roster_emp_names}

    unavail_map = {}
    if isinstance(unavail_data, pd.DataFrame) and not unavail_data.empty:
        un_name_c = find_column(unavail_data, ["employee", "name", "staff", "person"])
        un_day_c = find_column(unavail_data, ["day", "date", "weekday"])
        un_win_c = find_column(unavail_data, ["time window", "window", "time", "unavailability", "reason"])
        
        if un_name_c and un_day_c and un_win_c:
            for _, u_row in unavail_data.iterrows():
                raw_u_emp = str(u_row.get(un_name_c, "")).strip()
                u_day = str(u_row.get(un_day_c, "")).strip().lower()
                u_win = str(u_row.get(un_win_c, "Full Day")).strip()
                
                matched_emp = find_matching_employee(raw_u_emp, name_map)
                if matched_emp and u_day:
                    key = (matched_emp.lower(), u_day)
                    if key not in unavail_map:
                        unavail_map[key] = []
                    unavail_map[key].append(u_win)

    for idx, row in df_out.iterrows():
        emp_name = str(row.get(emp_col, "")).strip()
        if not emp_name:
            continue
        emp_lower = emp_name.lower()

        for day in days_in_df:
            val = str(row.get(day, "")).strip()
            val_lower = val.lower()
            key = (emp_lower, day.lower())

            # Check if this cell is unassigned/off or already flagged as unavailable
            if not val or val_lower in ["off", "none", "nan", "unavailable", " unavailable"] or "unavail" in val_lower or val.startswith("🚫"):
                if key in unavail_map and unavail_map[key]:
                    clean_win = clean_win_display(unavail_map[key][0])
                    if "all day" in clean_win.lower():
                        clean_win = "Full Day"
                    df_out.at[idx, day] = f"🚫 Unavailable ({clean_win})"
                else:
                    # Clean cell for employees with no logged unavailability on this day
                    df_out.at[idx, day] = ""

    return df_out

def highlight_unavailability_dataframe(df):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df

    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    days_in_df = [c for c in days_of_week if c in df.columns]
    
    def color_cells(val):
        v = str(val).strip()
        if not v:
            return ''
        if "🚫" in v or "unavail" in v.lower():
            # Crimson alert background with bold text for admin visual assistance
            return 'background-color: rgba(184, 40, 40, 0.45); color: #ffe6e6; font-weight: 800; border: 1.5px solid #ff4d4d;'
        return 'background-color: rgba(14, 43, 38, 0.6); color: #ffffff; font-weight: 600;'

    return df.style.map(color_cells, subset=days_in_df)

def strip_daily_gross_row(df):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    emp_col = df.columns[0]
    return df[~df[emp_col].astype(str).str.contains("💰|Predicted|Summary|Total", case=False, na=False)].copy()

def attach_daily_gross_row(df, daily_gross_dict):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    df_clean = strip_daily_gross_row(df)
    emp_col = df_clean.columns[0]
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    summary_row = {c: "" for c in df_clean.columns}
    summary_row[emp_col] = "💰 PREDICTED DAILY GROSS"
    
    for day in days_of_week:
        if day in df_clean.columns:
            val = daily_gross_dict.get(day, 0.0) if isinstance(daily_gross_dict, dict) else 0.0
            summary_row[day] = f"${val:,.2f}" if val > 0 else "$0.00"
            
    df_out = pd.concat([df_clean, pd.DataFrame([summary_row])], ignore_index=True)
    return df_out

def load_finalized_roster(csv_filename):
    csv_path = os.path.join(FINALIZED_DIR, csv_filename)
    if os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
            return clean_roster_unavailability_display(df)
        except:
            return None
    return None

def delete_finalized_roster(date_str):
    if not os.path.exists(FINALIZED_DIR):
        return False
    deleted = False
    date_label = date_str
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d").date()
        date_label = dt.strftime("%d.%m.%Y")
    except:
        pass
        
    for f in os.listdir(FINALIZED_DIR):
        if date_str in f or date_label in f:
            try:
                os.remove(os.path.join(FINALIZED_DIR, f))
                deleted = True
            except:
                pass
    return deleted

def calculate_roster_wages(edited_df):
    if edited_df is None or edited_df.empty:
        return {
            "total_gross": 0.0,
            "total_tax": 0.0,
            "total_net": 0.0,
            "total_super": 0.0,
            "total_hours": 0.0,
            "avg_hourly_rate": 0.0,
            "breakdown_df": pd.DataFrame()
        }

    emp_meta = {}
    if "manual_employees" in st.session_state and isinstance(st.session_state.manual_employees, pd.DataFrame):
        e_df = st.session_state.manual_employees
        n_col = find_column(e_df, ["name", "employee", "staff"], "Name")
        a_col = find_column(e_df, ["age"], None)
        dob_col = find_column(e_df, ["dob", "date of birth", "birth date", "birthdate"], None)
        s_col = find_column(e_df, ["employment type", "status", "classification", "type"], "Employment Type")
        
        for _, r in e_df.iterrows():
            name = str(r.get(n_col, "")).strip()
            if name:
                age = None
                if dob_col and dob_col in e_df.columns:
                    dob_res = calculate_age_from_dob(r.get(dob_col))
                    if dob_res:
                        age = dob_res[0] if isinstance(dob_res, tuple) else dob_res
                if age is None and a_col and a_col in e_df.columns:
                    try:
                        age = int(float(r.get(a_col)))
                    except:
                        age = None
                if age is None:
                    age = 21
                status = str(r.get(s_col, "Casual")).strip().lower()
                emp_meta[name.lower()] = {"name": name, "age": age, "status": status}

    for u_key, u_data in user_profiles.items():
        emp_name = u_data.get("employee_name", u_key)
        prof = u_data.get("profile", {})
        status = str(prof.get("classification", "Casual")).strip().lower()
        dob_val = prof.get("dob", "")
        dob_res = calculate_age_from_dob(dob_val)
        age = (dob_res[0] if isinstance(dob_res, tuple) else dob_res) if dob_res else 21
        if emp_name.lower() not in emp_meta or emp_meta[emp_name.lower()]["age"] == 21:
            emp_meta[emp_name.lower()] = {"name": emp_name, "age": age, "status": status}

    emp_col = find_column(edited_df, ["employee", "name", "staff"], "Employee")
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    emp_summary = []
    daily_gross_totals = {day: 0.0 for day in days_of_week}

    for _, row in edited_df.iterrows():
        emp_raw_name = str(row.get(emp_col, "")).strip()
        if not emp_raw_name or emp_raw_name.lower() in ["none", "nan", "total", "summary"] or "💰" in emp_raw_name or "predicted" in emp_raw_name.lower():
            continue

        meta = emp_meta.get(emp_raw_name.lower(), {"name": emp_raw_name, "age": 21, "status": "casual"})
        age = meta["age"]
        status_clean = meta["status"].lower()
        emp_name_lower = emp_raw_name.lower()

        # Specific user classification rules:
        # 1. Viet & Jane calculate as Casual Level 1
        # 2. Robert is Baker Level 1
        is_viet_or_jane = any(k in emp_name_lower for k in ["viet", "jane"]) or "owner" in status_clean
        is_casual = is_viet_or_jane or "casual" in status_clean
        emp_type = "casual" if is_casual else "pt_ft"

        # Level assignment: Robert is Level 1, others with explicit level 3 remain level 3 if specified
        if "robert" in emp_name_lower:
            lvl = "level1"
        elif is_viet_or_jane:
            lvl = "level1"
        else:
            is_baker_lvl3 = "level 3" in status_clean
            lvl = "level3" if is_baker_lvl3 else "level1"

        age_key = "adult" if age >= 21 else (age if age in [15, 16, 17, 18, 19, 20] else (15 if age < 15 else "adult"))

        # Strict Fair Work Ombudsman Pay Guide - General Retail Industry Award [MA000004] (Effective 01/07/2026)
        PAY_GUIDE_MA000004 = {
            # Casual Level 1 (Service Staff & Casual Level 1)
            ("casual", "level1", "adult"): {"ord": 34.76, "eve": 41.72, "sat": 41.72, "sun": 48.67, "ph": 69.53},
            ("casual", "level1", 20): {"ord": 34.76, "eve": 41.72, "sat": 41.72, "sun": 48.67, "ph": 69.53},
            ("casual", "level1", 19): {"ord": 27.81, "eve": 33.38, "sat": 33.38, "sun": 38.94, "ph": 55.63},
            ("casual", "level1", 18): {"ord": 24.34, "eve": 29.21, "sat": 29.21, "sun": 34.07, "ph": 48.68},
            ("casual", "level1", 17): {"ord": 20.86, "eve": 25.04, "sat": 25.04, "sun": 29.21, "ph": 41.73},
            ("casual", "level1", 16): {"ord": 17.39, "eve": 20.87, "sat": 20.87, "sun": 24.34, "ph": 34.78},
            ("casual", "level1", 15): {"ord": 15.64, "eve": 18.77, "sat": 18.77, "sun": 21.89, "ph": 31.28},
            # Full-time / Part-time Level 1 (Service Staff & Baker Lvl 1)
            ("pt_ft", "level1", "adult"): {"ord": 27.81, "eve": 34.76, "sat": 34.76, "sun": 41.72, "ph": 62.57},
            ("pt_ft", "level1", 20): {"ord": 27.81, "eve": 34.76, "sat": 34.76, "sun": 41.72, "ph": 62.57},
            ("pt_ft", "level1", 19): {"ord": 22.25, "eve": 27.81, "sat": 27.81, "sun": 33.38, "ph": 50.06},
            ("pt_ft", "level1", 18): {"ord": 19.47, "eve": 24.34, "sat": 24.34, "sun": 29.21, "ph": 43.81},
            ("pt_ft", "level1", 17): {"ord": 16.69, "eve": 20.86, "sat": 20.86, "sun": 25.04, "ph": 37.55},
            ("pt_ft", "level1", 16): {"ord": 13.91, "eve": 17.39, "sat": 17.39, "sun": 20.87, "ph": 31.30},
            ("pt_ft", "level1", 15): {"ord": 12.51, "eve": 15.64, "sat": 15.64, "sun": 18.77, "ph": 28.15},
            # Casual Level 3
            ("casual", "level3", "adult"): {"ord": 36.11, "eve": 43.34, "sat": 43.34, "sun": 50.56, "ph": 72.23},
            ("casual", "level3", 20): {"ord": 36.11, "eve": 43.34, "sat": 43.34, "sun": 50.56, "ph": 72.23},
            ("casual", "level3", 19): {"ord": 28.89, "eve": 34.67, "sat": 34.67, "sun": 40.44, "ph": 57.78},
            ("casual", "level3", 18): {"ord": 25.28, "eve": 30.33, "sat": 30.33, "sun": 35.39, "ph": 50.55},
            ("casual", "level3", 17): {"ord": 21.66, "eve": 26.00, "sat": 26.00, "sun": 30.33, "ph": 43.33},
            ("casual", "level3", 16): {"ord": 18.05, "eve": 21.66, "sat": 21.66, "sun": 25.27, "ph": 36.10},
            ("casual", "level3", 15): {"ord": 16.25, "eve": 19.50, "sat": 19.50, "sun": 22.75, "ph": 32.50},
            # Full-time / Part-time Level 3
            ("pt_ft", "level3", "adult"): {"ord": 28.89, "eve": 36.11, "sat": 36.11, "sun": 43.34, "ph": 65.00},
            ("pt_ft", "level3", 20): {"ord": 28.89, "eve": 36.11, "sat": 36.11, "sun": 43.34, "ph": 65.00},
            ("pt_ft", "level3", 19): {"ord": 23.11, "eve": 28.89, "sat": 28.89, "sun": 34.67, "ph": 52.00},
            ("pt_ft", "level3", 18): {"ord": 20.22, "eve": 25.28, "sat": 25.28, "sun": 30.33, "ph": 45.50},
            ("pt_ft", "level3", 17): {"ord": 17.33, "eve": 21.66, "sat": 21.66, "sun": 26.00, "ph": 38.99},
            ("pt_ft", "level3", 16): {"ord": 14.44, "eve": 18.05, "sat": 18.05, "sun": 21.66, "ph": 32.49},
            ("pt_ft", "level3", 15): {"ord": 13.00, "eve": 16.25, "sat": 16.25, "sun": 19.50, "ph": 29.25},
        }
        rates = PAY_GUIDE_MA000004.get((emp_type, lvl, age_key), PAY_GUIDE_MA000004[("casual", "level1", "adult")])

        total_emp_hours = 0.0
        total_emp_gross = 0.0

        for day in days_of_week:
            if day in row:
                shift_val = str(row[day]).strip()
                parsed = parse_shift_range(shift_val)
                if parsed:
                    start_t, end_t, duration = parsed
                    paid_hrs = duration - 0.5 if duration >= 5.0 else duration
                    total_emp_hours += paid_hrs

                    day_shift_gross = 0.0
                    if day in ["Saturday"]:
                        day_shift_gross = paid_hrs * rates["sat"]
                    elif day in ["Sunday"]:
                        day_shift_gross = paid_hrs * rates["sun"]
                    else:
                        # Mon-Fri
                        if end_t <= 18.0:
                            day_shift_gross = paid_hrs * rates["ord"]
                        elif start_t >= 18.0:
                            day_shift_gross = paid_hrs * rates["eve"]
                        else:
                            pre_6_hrs = max(0.0, 18.0 - start_t)
                            post_6_hrs = max(0.0, end_t - 18.0)
                            if duration >= 5.0:
                                ratio = paid_hrs / duration
                                pre_paid = pre_6_hrs * ratio
                                post_paid = post_6_hrs * ratio
                            else:
                                pre_paid = pre_6_hrs
                                post_paid = post_6_hrs
                            day_shift_gross = pre_paid * rates["ord"] + post_paid * rates["eve"]

                    total_emp_gross += day_shift_gross
                    daily_gross_totals[day] += day_shift_gross

        g = total_emp_gross
        if g <= 359:
            tax = 0.0
        elif g <= 865:
            tax = (g - 359) * 0.19
        elif g <= 2500:
            tax = 96.14 + (g - 865) * 0.325
        else:
            tax = 627.51 + (g - 2500) * 0.37

        net_pay = max(0.0, g - tax)

        # ATO Superannuation Guarantee Rule: Under-18 employees working 30h or less in a week receive $0.00 super
        if age < 18 and total_emp_hours <= 30.0:
            super_sg = 0.0
        else:
            super_sg = g * 0.125

        if is_viet_or_jane:
            status_label = "Casual Lvl 1 (Owner)"
        elif "robert" in emp_name_lower:
            status_label = "Part-Time (Baker Lvl 1)"
        else:
            status_label = "Casual" if is_casual else ("Part-Time" if "part" in status_clean else "Full-Time")
            if age < 21:
                status_label += f" ({age}yo)"

        emp_team = get_employee_team(meta.get("position", ""))
        emp_summary.append({
            "Staff Member": emp_raw_name,
            "Team": emp_team,
            "Status": status_label,
            "Paid Hours": round(total_emp_hours, 1),
            "Gross Pay": round(total_emp_gross, 2),
            "Est. Tax": round(tax, 2),
            "Net Pay": round(net_pay, 2),
            "Super (12.5%)": round(super_sg, 2)
        })

    breakdown_df = pd.DataFrame(emp_summary)
    breakdown_df = sort_dataframe_by_team_and_age(breakdown_df)
    tot_gross = sum(x["Gross Pay"] for x in emp_summary)
    tot_tax = sum(x["Est. Tax"] for x in emp_summary)
    tot_net = sum(x["Net Pay"] for x in emp_summary)
    tot_super = sum(x["Super (12.5%)"] for x in emp_summary)
    tot_hrs = sum(x["Paid Hours"] for x in emp_summary)
    avg_rate = (tot_gross / tot_hrs) if tot_hrs > 0 else 0.0

    return {
        "total_gross": round(tot_gross, 2),
        "total_tax": round(tot_tax, 2),
        "total_net": round(tot_net, 2),
        "total_super": round(tot_super, 2),
        "total_hours": round(tot_hrs, 1),
        "avg_hourly_rate": round(avg_rate, 2),
        "daily_gross": {d: round(daily_gross_totals[d], 2) for d in days_of_week},
        "breakdown_df": breakdown_df
    }

def calculate_weekly_hour_rate_breakdown(roster_df):
    if roster_df is None or roster_df.empty:
        return pd.DataFrame(columns=[
            "staff name", "early hour", "late hour", "ordinary hour", "saturday", "sunday", "laundry allowance"
        ])

    clean_df = strip_daily_gross_row(roster_df)
    emp_col = find_column(clean_df, ["employee", "name", "staff"], "Employee")
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    records = []

    for _, row in clean_df.iterrows():
        emp_raw_name = str(row.get(emp_col, "")).strip()
        if not emp_raw_name or emp_raw_name.lower() in ["none", "nan", "total", "summary"] or "💰" in emp_raw_name or "predicted" in emp_raw_name.lower():
            continue

        early_hrs = 0.0
        late_hrs = 0.0
        ord_hrs = 0.0
        sat_hrs = 0.0
        sun_hrs = 0.0
        shifts_count = 0

        for day in days_of_week:
            if day in row:
                shift_val = str(row[day]).strip()
                parsed = parse_shift_range(shift_val)
                if parsed:
                    start_t, end_t, duration = parsed
                    # Full shift hours calculated without deducting 30min unpaid meal break
                    shifts_count += 1

                    if day == "Saturday":
                        sat_hrs += duration
                    elif day == "Sunday":
                        sun_hrs += duration
                    else:
                        # Mon-Fri
                        # Early: before 7:00am
                        early_raw = max(0.0, min(end_t, 7.0) - start_t)
                        # Late: after 6:00pm (18.0)
                        late_raw = max(0.0, end_t - max(start_t, 18.0))
                        # Ordinary: between 7:00am and 6:00pm
                        ord_start = max(start_t, 7.0)
                        ord_end = min(end_t, 18.0)
                        ord_raw = max(0.0, ord_end - ord_start) if ord_end > ord_start else 0.0

                        early_hrs += early_raw
                        late_hrs += late_raw
                        ord_hrs += ord_raw

        laundry_allowance = round(shifts_count * 1.28, 2)

        records.append({
            "staff name": emp_raw_name,
            "early hour": round(early_hrs, 2),
            "late hour": round(late_hrs, 2),
            "ordinary hour": round(ord_hrs, 2),
            "saturday": round(sat_hrs, 2),
            "sunday": round(sun_hrs, 2),
            "laundry allowance": f"${laundry_allowance:.2f}"
        })

    df_result = pd.DataFrame(records)
    if not df_result.empty:
        df_result = df_result.sort_values(by="staff name", ascending=True, key=lambda col: col.astype(str).str.lower()).reset_index(drop=True)
        tot_early = sum(r["early hour"] for r in records)
        tot_late = sum(r["late hour"] for r in records)
        tot_ord = sum(r["ordinary hour"] for r in records)
        tot_sat = sum(r["saturday"] for r in records)
        tot_sun = sum(r["sunday"] for r in records)
        tot_laundry = sum(float(r["laundry allowance"].replace("$", "")) for r in records)

        summary_row = {
            "staff name": "💰 TOTAL SUMMARY",
            "early hour": round(tot_early, 2),
            "late hour": round(tot_late, 2),
            "ordinary hour": round(tot_ord, 2),
            "saturday": round(tot_sat, 2),
            "sunday": round(tot_sun, 2),
            "laundry allowance": f"${tot_laundry:.2f}"
        }
        df_result = pd.concat([df_result, pd.DataFrame([summary_row])], ignore_index=True)

    return df_result

def generate_xero_timesheet_csv(breakdown_df):
    """
    Generates a Xero Payroll Timesheet CSV format ready for 1-click import into Xero Pay Run.
    """
    if breakdown_df is None or breakdown_df.empty:
        return ""
    
    rows = []
    for _, r in breakdown_df.iterrows():
        emp = str(r.get("staff name", "")).strip()
        if not emp or "TOTAL" in emp.upper() or "SUMMARY" in emp.upper():
            continue
        
        ord_h = float(r.get("ordinary hour", 0.0) or 0.0)
        sat_h = float(r.get("saturday", 0.0) or 0.0)
        sun_h = float(r.get("sunday", 0.0) or 0.0)
        early_h = float(r.get("early hour", 0.0) or 0.0)
        late_h = float(r.get("late hour", 0.0) or 0.0)
        laundry_str = str(r.get("laundry allowance", "$0.00")).replace("$", "").strip()
        laundry_val = float(laundry_str) if laundry_str else 0.0

        if ord_h > 0:
            rows.append({"Employee Name": emp, "Earnings Rate": "Ordinary Hours", "Hours": ord_h})
        if sat_h > 0:
            rows.append({"Employee Name": emp, "Earnings Rate": "Saturday Penalty", "Hours": sat_h})
        if sun_h > 0:
            rows.append({"Employee Name": emp, "Earnings Rate": "Sunday Penalty", "Hours": sun_h})
        if early_h > 0:
            rows.append({"Employee Name": emp, "Earnings Rate": "Early Morning Shift", "Hours": early_h})
        if late_h > 0:
            rows.append({"Employee Name": emp, "Earnings Rate": "Late Shift", "Hours": late_h})
        if laundry_val > 0:
            rows.append({"Employee Name": emp, "Earnings Rate": "Laundry Allowance", "Hours": laundry_val})

    df_xero = pd.DataFrame(rows)
    return df_xero.to_csv(index=False)

def generate_xero_autofill_js(breakdown_df):
    if breakdown_df is None or breakdown_df.empty:
        return "// No breakdown data available."
    data_records = []
    for _, row in breakdown_df.iterrows():
        emp_name = str(row.get("staff name", "")).strip()
        if not emp_name or "TOTAL" in emp_name.upper() or "SUMMARY" in emp_name.upper():
            continue
        data_records.append({
            "name": emp_name,
            "early": float(row.get("early hour", 0.0) or 0.0),
            "late": float(row.get("late hour", 0.0) or 0.0),
            "ordinary": float(row.get("ordinary hour", 0.0) or 0.0),
            "saturday": float(row.get("saturday", 0.0) or 0.0),
            "sunday": float(row.get("sunday", 0.0) or 0.0),
            "laundry": str(row.get("laundry allowance", "$0.00")).replace("$", "").strip()
        })
    json_str = json.dumps(data_records, indent=2)
    return f"""javascript:(function(){{
    const staffData = {json_str};
    console.log("🥐 Brumby's Bakery - Xero Payroll Auto-Fill Triggered");
    
    // Check if on Xero Pay Run Overview summary vs Detailed Payslip/Timesheet input page
    const inputsOnPage = document.querySelectorAll('input[type="text"], input[type="number"], input:not([type="hidden"])');
    
    if (inputsOnPage.length === 0) {{
        alert("ℹ️ You are currently on the Xero Pay Run Overview Summary table.\\n\\n💡 Summary tables display read-only totals.\\n\\n👉 Recommendation:\\n1. Use the '📥 Download Xero Timesheet CSV' button in your Bakery App for 1-click Pay Run import!\\n2. Or click into an individual employee's payslip edit screen in Xero, then run this auto-fill button.");
        return;
    }}

    let matchedCount = 0, skippedCount = 0;
    const allRows = document.querySelectorAll('tr, div[class*="employee"], div[class*="payrun-row"], div[class*="row"], div[class*="modal"]');
    
    staffData.forEach(item => {{
        const targetName = item.name.toLowerCase();
        let foundRow = null;
        for (let r of allRows) {{
            const text = (r.innerText || r.textContent || "").toLowerCase();
            if (text.includes(targetName) || targetName.split(' ').every(part => text.includes(part))) {{
                foundRow = r; break;
            }}
        }}
        if (!foundRow) {{ skippedCount++; return; }}
        matchedCount++;
        const inputs = Array.from(foundRow.querySelectorAll('input[type="text"], input[type="number"], input:not([type="hidden"])'));
        inputs.forEach(input => {{
            const combinedContext = `${{input.getAttribute('aria-label')||''}} ${{input.getAttribute('placeholder')||''}} ${{input.getAttribute('name')||''}} ${{input.parentElement ? input.parentElement.innerText : ''}}`.toLowerCase();
            function setVal(v) {{ if (!v || v === "0.00") return; input.value = v; input.dispatchEvent(new Event('input', {{ bubbles: true }})); input.dispatchEvent(new Event('change', {{ bubbles: true }})); }}
            if (combinedContext.includes('sat')) setVal(item.saturday);
            else if (combinedContext.includes('sun')) setVal(item.sunday);
            else if (combinedContext.includes('early') || combinedContext.includes('night')) setVal(item.early);
            else if (combinedContext.includes('late') || combinedContext.includes('evening')) setVal(item.late);
            else if (combinedContext.includes('ord') || combinedContext.includes('base') || combinedContext.includes('hours')) setVal(item.ordinary);
            else if (combinedContext.includes('laundry')) setVal(item.laundry);
        }});
    }});
    alert(`🥐 Brumby's Bakery Xero Auto-Fill Complete!\\n\\n✅ Matched: ${{matchedCount}} staff\\n⚠️ Skipped (Not found on page): ${{skippedCount}} staff`);
}})();"""

def run_xero_playwright_autofill(breakdown_df, target_url="https://payroll.xero.com/PayRun/PayRun/Details/74565804?CID=!cSWq8", headless=False):
    if breakdown_df is None or breakdown_df.empty:
        return False, "No data available in Current Week Hour Rate Breakdown."
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return False, "Playwright is not installed in this environment. Run `pip install playwright` to enable local automated browser bot."

    staff_records = []
    for _, row in breakdown_df.iterrows():
        emp_name = str(row.get("staff name", "")).strip()
        if not emp_name or "TOTAL" in emp_name.upper() or "SUMMARY" in emp_name.upper():
            continue
        staff_records.append({
            "name": emp_name,
            "early": float(row.get("early hour", 0.0) or 0.0),
            "late": float(row.get("late hour", 0.0) or 0.0),
            "ordinary": float(row.get("ordinary hour", 0.0) or 0.0),
            "saturday": float(row.get("saturday", 0.0) or 0.0),
            "sunday": float(row.get("sunday", 0.0) or 0.0),
            "laundry": str(row.get("laundry allowance", "$0.00")).replace("$", "").strip()
        })

    matched_list, skipped_list = [], []
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless, args=["--start-maximized"])
            context = browser.new_context(viewport={"width": 1400, "height": 900})
            page = context.new_page()
            page.goto(target_url, timeout=60000)
            page.wait_for_load_state("networkidle")
            if "login" in page.url.lower():
                page.wait_for_url("**/PayRun/**", timeout=120000)
            page.wait_for_selector("table, div[class*='payrun']", timeout=30000)
            rows = page.query_selector_all("tr, div[class*='row']")

            for item in staff_records:
                target_name = item["name"].strip().lower()
                found_row = None
                for r in rows:
                    row_text = (r.inner_text() or "").lower()
                    if target_name in row_text or all(part in row_text for part in target_name.split()):
                        found_row = r; break
                if not found_row:
                    skipped_list.append(item["name"]); continue
                matched_list.append(item["name"])
                inputs = found_row.query_selector_all("input[type='text'], input[type='number'], input:not([type='hidden'])")
                for inp in inputs:
                    context_text = f"{inp.get_attribute('aria-label') or ''} {inp.get_attribute('placeholder') or ''} {inp.get_attribute('name') or ''}".lower()
                    if "sat" in context_text and item["saturday"] > 0: inp.fill(str(item["saturday"]))
                    elif "sun" in context_text and item["sunday"] > 0: inp.fill(str(item["sunday"]))
                    elif ("early" in context_text or "night" in context_text) and item["early"] > 0: inp.fill(str(item["early"]))
                    elif ("late" in context_text or "evening" in context_text) and item["late"] > 0: inp.fill(str(item["late"]))
                    elif ("ord" in context_text or "base" in context_text or "hours" in context_text) and item["ordinary"] > 0: inp.fill(str(item["ordinary"]))
                    elif ("laundry" in context_text or "allowance" in context_text) and float(item["laundry"] or 0) > 0: inp.fill(str(item["laundry"]))
            browser.close()
            return True, f"✅ Xero Auto-Input Complete!\n\nMatched ({len(matched_list)}): {', '.join(matched_list)}\nSkipped ({len(skipped_list)}): {', '.join(skipped_list) if skipped_list else 'None'}"
    except Exception as e:
        return False, f"Xero Automation Error: {str(e)}"

def build_payroll_historical_trend():
    past_rosters = list_finalized_rosters()
    if not past_rosters:
        return pd.DataFrame()
    
    records = []
    for r in past_rosters:
        df = load_finalized_roster(r["csv_filename"])
        if df is not None and not df.empty:
            w_sum = calculate_roster_wages(df)
            try:
                dt = datetime.strptime(r["date_str"], "%Y-%m-%d").date()
                week_label = dt.strftime("%d/%m/%Y")
            except:
                dt = datetime.min.date()
                week_label = r["date_str"]
                
            records.append({
                "date": dt,
                "Roster Week": week_label,
                "Gross Payroll ($)": w_sum["total_gross"],
                "Est. PAYG Tax ($)": w_sum["total_tax"],
                "Net Take-Home ($)": w_sum["total_net"],
                "Super 12.5% ($)": w_sum["total_super"]
            })
            
    if not records:
        return pd.DataFrame()
        
    trend_df = pd.DataFrame(records).sort_values("date").reset_index(drop=True)
    return trend_df

def find_column(df, candidates, default=""):
    if df is None or not hasattr(df, "columns") or len(df.columns) == 0:
        return default
    # 1. Exact match
    for c in df.columns:
        if str(c).strip().lower() in [cand.lower() for cand in candidates]:
            return c
    # 2. Substring match
    for c in df.columns:
        c_clean = str(c).strip().lower()
        for cand in candidates:
            cand_clean = cand.lower()
            if cand_clean in c_clean or c_clean in cand_clean:
                return c
    return default

# Initialize Session States with Disk Cache
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'logged_in_user' not in st.session_state:
    st.session_state.logged_in_user = None
if 'user_role' not in st.session_state:
    st.session_state.user_role = None
if 'is_demo' not in st.session_state:
    st.session_state.is_demo = False

user_profiles = get_active_user_profiles()

# LOGIN PAGE IF NOT AUTHENTICATED
if not st.session_state.authenticated:
    st.markdown("""
    <style>
        .stApp {
            background-color: #051412 !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1.8, 1])
    with col2:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 35px 30px; border-radius: 20px; border: 2px solid #e5a93c; box-shadow: 0 12px 36px rgba(0,0,0,0.6); margin-top: 40px;">
            <h2 style="color: #e5a93c !important; margin-top: 0; text-align: center; font-size: 1.8rem; font-weight: 900;">🥐 Brumby's Pakenham — Portal</h2>
            <p style="color: #c8e6e0 !important; font-size: 0.95rem; text-align: center; margin-bottom: 25px;">Please log in with your username and password to access your bakery account.</p>
        """, unsafe_allow_html=True)
        
        login_user = st.text_input("Username", key="input_user")
        login_pass = st.text_input("Password", type="password", key="input_pass")
        
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🚀 Login to Portal", use_container_width=True, key="btn_login"):
            login_user_clean = login_user.strip().lower()
            current_profiles = load_user_profiles()
            if login_user_clean in current_profiles:
                account = current_profiles[login_user_clean]
                if account.get("password") == login_pass:
                    st.session_state.authenticated = True
                    st.session_state.logged_in_user = login_user_clean
                    st.session_state.user_role = account.get("role", "Employee")
                    st.session_state.is_demo = False
                    st.success(f"Welcome back, {account.get('employee_name', login_user_clean)}!")
                    st.rerun()
                else:
                    st.error("🔒 Invalid username or password. Please check your credentials.")
            else:
                st.error("🔒 Invalid username or password. Please check your credentials.")
                
        # Store Phone Timeclock Kiosk Launcher
        st.markdown("<hr style='border-color: rgba(229, 169, 60, 0.4); margin: 18px 0;'>", unsafe_allow_html=True)
        st.markdown("""
        <div style="text-align: center; margin-bottom: 10px;">
            <span style="color: #e5a93c; font-weight: 800; font-size: 1rem;">🏪 SHARED STORE PHONE KIOSK MODE</span><br>
            <span style="color: #c8e6e0; font-size: 0.85rem;">Set this device as the shared store terminal at Brumby's Bakery:</span>
        </div>
        """, unsafe_allow_html=True)
        if st.button("⏱️ Launch Store Timeclock Kiosk Terminal", use_container_width=True, key="btn_launch_kiosk"):
            st.session_state.authenticated = True
            st.session_state.logged_in_user = "store.kiosk"
            st.session_state.user_role = "Kiosk"
            st.session_state.is_kiosk_mode = True
            st.session_state.is_demo = False
            st.rerun()

        st.markdown("""
        </div>
        """, unsafe_allow_html=True)
    st.stop()

# --- SIDEBAR CONFIGURATION (AUTHENTICATED) ---
st.sidebar.image("https://img.icons8.com/fluency/96/bakery.png", width=80)
st.sidebar.title("🍞 Bakery Portal Controls")

curr_user_key = st.session_state.logged_in_user
curr_user_info = user_profiles.get(curr_user_key, {})
display_name = "🏪 Store Timeclock Kiosk" if st.session_state.get("is_kiosk_mode", False) else curr_user_info.get("employee_name", curr_user_key)
role_title = "Store Kiosk" if st.session_state.get("is_kiosk_mode", False) else curr_user_info.get("role", "Employee")

st.sidebar.markdown(f"""
<div style="background-color: rgba(229, 169, 60, 0.15); padding: 12px 16px; border-radius: 12px; border: 1.5px solid #e5a93c; margin-bottom: 15px;">
    <div style="color: #e5a93c; font-size: 0.85rem; font-weight: 700;">LOGGED IN ACCOUNT</div>
    <div style="color: #ffffff; font-size: 1.1rem; font-weight: 900;">👤 {display_name}</div>
    <div style="color: #c8e6e0; font-size: 0.85rem;">Role: <b>{role_title}</b></div>
</div>
""", unsafe_allow_html=True)

# Helper function to render Change Password form
def render_change_password_form(user_key, is_admin=False):
    user_data = user_profiles.get(user_key, {})
    emp_name = user_data.get("employee_name", user_key)
    
    with st.expander(f"🔑 Change Password for {emp_name}", expanded=False):
        with st.form(key=f"form_change_pw_{user_key}"):
            if not is_admin:
                curr_pw = st.text_input("Current Password", type="password", key=f"cp_curr_pw_{user_key}")
            else:
                curr_pw = None
                
            new_pw = st.text_input("New Password", type="password", key=f"cp_new_pw_{user_key}")
            confirm_pw = st.text_input("Confirm New Password", type="password", key=f"cp_conf_pw_{user_key}")
            
            submit_pw = st.form_submit_button("🔒 Save New Password")
            
            if submit_pw:
                actual_pw = user_data.get("password", "")
                if not is_admin and curr_pw != actual_pw:
                    st.error("❌ Incorrect current password.")
                elif not new_pw.strip():
                    st.error("❌ New password cannot be empty.")
                elif new_pw != confirm_pw:
                    st.error("❌ New passwords do not match.")
                else:
                    user_profiles[user_key]["password"] = new_pw.strip()
                    save_user_profiles(user_profiles)
                    st.success(f"✅ Password for {emp_name} updated successfully!")

# Sidebar Change Password Expander (Hidden in Store Kiosk Mode)
if not st.session_state.get("is_kiosk_mode", False) and curr_user_key != "store.kiosk":
    with st.sidebar.expander("🔑 Change My Password", expanded=False):
        with st.form(key=f"form_sidebar_pw_{curr_user_key}"):
            sb_curr_pw = st.text_input("Current Password", type="password", key=f"sb_cp_curr_{curr_user_key}")
            sb_new_pw = st.text_input("New Password", type="password", key=f"sb_cp_new_{curr_user_key}")
            sb_conf_pw = st.text_input("Confirm New Password", type="password", key=f"sb_cp_conf_{curr_user_key}")
            
            sb_submit_pw = st.form_submit_button("🔒 Update Password")
            
            if sb_submit_pw:
                actual_pw = curr_user_info.get("password", "")
                if sb_curr_pw != actual_pw:
                    st.error("❌ Incorrect current password.")
                elif not sb_new_pw.strip():
                    st.error("❌ New password cannot be empty.")
                elif sb_new_pw != sb_conf_pw:
                    st.error("❌ New passwords do not match.")
                else:
                    user_profiles[curr_user_key]["password"] = sb_new_pw.strip()
                    save_user_profiles(user_profiles)
                    st.success("✅ Your password has been updated successfully!")

if role_title == "Manager":
    with st.sidebar.expander("📧 Email Settings (SMTP & Portal Link)", expanded=False):
        smtp_cfg = load_smtp_config()
        with st.form(key="form_smtp_config"):
            st.markdown("##### 🌐 Default Portal Web Link")
            cfg_url = st.text_input("Streamlit Cloud URL", value=smtp_cfg.get("portal_url", "https://weekly-roster-generator.streamlit.app"))
            cfg_sname = st.text_input("Sender Display Name", value=smtp_cfg.get("sender_name", "Bakery Manager"))
            
            st.markdown("##### 📮 SMTP Server Credentials")
            cfg_semail = st.text_input("Sender Email Address", value=smtp_cfg.get("sender_email", "Brumby.pakenham@gmail.com"), placeholder="e.g. Brumby.pakenham@gmail.com")
            cfg_spass = st.text_input("Sender App Password", value=smtp_cfg.get("sender_password", ""), type="password")
            cfg_host = st.text_input("SMTP Host", value=smtp_cfg.get("smtp_server", "smtp.gmail.com"))
            cfg_port = st.number_input("SMTP Port", value=int(smtp_cfg.get("smtp_port", 587)))
            
            st.markdown("##### 🔔 Manager Notification Recipients")
            cfg_recipients = st.text_input("Notification Emails (comma-separated)", value=smtp_cfg.get("notification_recipients", "quietsong2006@yahoo.com, uyentrinhtran2309@gmail.com"))
            
            c_save, c_test = st.columns([1.2, 1])
            with c_save:
                btn_save_smtp = st.form_submit_button("💾 Save Settings")
            with c_test:
                btn_test_smtp = st.form_submit_button("🧪 Test Email")

            if btn_save_smtp:
                saved_pass = cfg_spass.strip() if cfg_spass.strip() else smtp_cfg.get("sender_password", "")
                new_cfg = {
                    "portal_url": cfg_url.strip(),
                    "sender_name": cfg_sname.strip(),
                    "sender_email": cfg_semail.strip(),
                    "sender_password": saved_pass,
                    "smtp_server": cfg_host.strip(),
                    "smtp_port": int(cfg_port),
                    "notification_recipients": cfg_recipients.strip()
                }
                save_smtp_config(new_cfg)
                if saved_pass:
                    st.success("✅ Email settings saved successfully!")
                else:
                    st.warning("⚠️ Settings saved, but **Sender App Password** is currently blank. Email notifications will not send until a 16-character Google App Password is entered.")

            if btn_test_smtp:
                saved_pass = cfg_spass.strip() if cfg_spass.strip() else smtp_cfg.get("sender_password", "")
                new_cfg = {
                    "portal_url": cfg_url.strip(),
                    "sender_name": cfg_sname.strip(),
                    "sender_email": cfg_semail.strip(),
                    "sender_password": saved_pass,
                    "smtp_server": cfg_host.strip(),
                    "smtp_port": int(cfg_port),
                    "notification_recipients": cfg_recipients.strip()
                }
                save_smtp_config(new_cfg)
                ok, msg = send_test_email_smtp(cfg_semail.strip())
                if ok:
                    st.success(msg)
                else:
                    st.error(msg)

    with st.sidebar.expander("🔥 Firebase Cloud Sync & Security", expanded=False):
        if is_firebase_active():
            st.success("🟢 **Firebase Cloud Sync Active**\nAll employee accounts, profiles, rosters, & shift data are continuously synced to Google Cloud.")
        else:
            st.warning("🟡 **Local Backup Mode**\nRunning on local `.csv` / `.json` files. Add your Firebase credentials to `.streamlit/secrets.toml` or Streamlit Cloud Secrets to enable 24/7 cloud sync.")
            
        if st.button("🚀 Upload Local Files to Firebase Cloud", use_container_width=True, key="btn_sync_firebase_now"):
            ok, msg = migrate_all_local_files_to_firebase()
            if ok:
                st.success(msg)
            else:
                st.error(msg)

def logout_user():
    for k in ["authenticated", "logged_in_user", "user_role", "is_demo", "is_kiosk_mode", "demo_user_profiles", "demo_state_initialized", "manual_employees", "manual_unavailability", "manual_requirements", "manual_fixed", "final_roster_df", "edit_employees", "edit_unavailability_v4", "edit_requirements", "edit_fixed"]:
        st.session_state.pop(k, None)
    st.session_state.authenticated = False
    st.rerun()

if st.sidebar.button("🚪 Logout", key="btn_logout"):
    logout_user()

if not st.session_state.get("is_kiosk_mode", False):
    if st.sidebar.button("🏪 Switch to Store Kiosk Mode", key="btn_sb_switch_kiosk", use_container_width=True):
        st.session_state.logged_in_user = "store.kiosk"
        st.session_state.user_role = "Kiosk"
        st.session_state.is_kiosk_mode = True
        st.rerun()

col_head1, col_head2 = st.columns([3.5, 1])
with col_head1:
    st.markdown('<h1 class="header-style">🥐 Brumby\'s Pakenham Portal</h1>', unsafe_allow_html=True)
    st.markdown(f'<p class="sub-header-style">Welcome back, <b>{display_name}</b> ({role_title})</p>', unsafe_allow_html=True)
with col_head2:
    st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
    if st.button("🚪 Logout", key="btn_logout_header", use_container_width=True):
        logout_user()

# Helper to read excel sheets robustly, converting everything to strings for easy editing
def read_excel_robust(uploaded_file):
    if uploaded_file is None:
        return None
    try:
        # Load without headers first to scan rows
        df_raw = pd.read_excel(uploaded_file, header=None)
        
        # Look for the header row containing key columns
        keywords = ["name", "employee", "staff", "role", "age", "dob", "commence", "start", "shift", "day", "monday", "tuesday", "unavailability", "fixed", "status", "type"]
        header_row_idx = 0
        
        for idx, row in df_raw.iterrows():
            row_vals = [str(x).strip().lower() for x in row if pd.notna(x)]
            if len(row_vals) >= 2:
                matches = sum(1 for val in row_vals for kw in keywords if kw in val)
                if matches >= 2:
                    header_row_idx = idx
                    break
                
        # Re-read from that header row index
        df = pd.read_excel(uploaded_file, header=header_row_idx)
        df.columns = [str(c).strip() for c in df.columns]
        
        # Convert all columns to strings to make them fully editable in st.data_editor
        df = clean_roster_dataframe(df)
        return df
    except Exception as e:
        st.error(f"Error parsing Excel file structure: {e}")
        return None

def clean_roster_dataframe(df):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
        
    df = df.copy()
    emp_col = find_column(df, ["employee", "name", "staff", "staff name", "employee name"], df.columns[0])
    
    junk_keywords = [
        "hours or more", "hours up to", "less than", "shift duration", "break",
        "total", "legend", "notes", "rationale", "brumby", "unnamed", "system.xml",
        "week", "employee shift", "paid rest", "unpaid meal", "entitlement",
        "new trainer", "personnel", "coverage", "requirements"
    ]
    
    clean_indices = []
    for idx, row in df.iterrows():
        val = str(row.get(emp_col, "")).strip()
        val_lower = val.lower()
        
        if not val or val_lower in ["nan", "none", "nat", "null", ""]:
            continue
            
        if any(kw in val_lower for kw in junk_keywords):
            continue
            
        clean_indices.append(idx)
        
    return df.loc[clean_indices].reset_index(drop=True)

import re

def parse_date_robust(date_str):
    if not date_str or str(date_str).strip().lower() in ["nan", "none", "nat", ""]:
        return None
    date_str = str(date_str).strip()
    try:
        dt = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
        if pd.notna(dt):
            return dt.date()
    except:
        pass
    return None

def extract_date_range(text):
    if not text or str(text).strip().lower() in ["nan", "none", "nat", ""]:
        return None, None
    text = str(text)
    dates_found = []
    date_matches = re.findall(r'\b\d{1,4}[-/\.]\d{1,2}[-/\.]\d{2,4}\b', text)
    for m in date_matches:
        d = parse_date_robust(m)
        if d:
            dates_found.append(d)
    
    if len(dates_found) >= 2:
        return min(dates_found), max(dates_found)
    elif len(dates_found) == 1:
        if any(kw in text.lower() for kw in ['from', 'after', 'since', 'starting']):
            return dates_found[0], None
        elif any(kw in text.lower() for kw in ['until', 'to', 'before', 'ending']):
            return None, dates_found[0]
        return dates_found[0], dates_found[0]
    return None, None

def is_unavail_applicable_to_date(dt_obj, u_day, u_win):
    u_day_clean = str(u_day).strip()
    u_win_clean = str(u_win).strip()
    
    # 1. Date range bounds
    start_d, end_d = extract_date_range(f"{u_day_clean} {u_win_clean}")
    if start_d and dt_obj < start_d:
        return False
    if end_d and dt_obj > end_d:
        return False
        
    # 2. Explicit date check
    explicit_date = parse_date_robust(u_day_clean)
    if explicit_date:
        return dt_obj == explicit_date
        
    # 3. Day-of-week matching
    weekday_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    dt_day_name = weekday_names[dt_obj.weekday()]
    dt_day_short = dt_day_name[:3]
    
    u_day_lower = u_day_clean.lower()
    
    if dt_day_name in u_day_lower or dt_day_short in u_day_lower:
        return True
        
    if any(kw in u_day_lower for kw in ["all week", "everyday", "any day"]):
        return True
    if "weekend" in u_day_lower and dt_obj.weekday() in [5, 6]:
        return True
    if "weekday" in u_day_lower and dt_obj.weekday() in [0, 1, 2, 3, 4]:
        return True
        
def calculate_age_from_dob(dob_str):
    if not dob_str or str(dob_str).strip().lower() in ["nan", "none", "nat", "null", ""]:
        return None
    try:
        dt = pd.to_datetime(dob_str, dayfirst=True, errors='coerce')
        if pd.notna(dt):
            today = datetime.now()
            age = today.year - dt.year - ((today.month, today.day) < (dt.month, dt.day))
            return age, dt.strftime('%d/%m/%Y')
    except:
        pass
    return None

def get_employee_team(role_str):
    if not role_str or pd.isna(role_str):
        return "Service Team"
    role_lower = str(role_str).strip().lower()
    if any(kw in role_lower for kw in ["baker", "baking", "pastry"]):
        return "Baking Team"
    return "Service Team"

def get_employee_team_and_age(emp_name, role_str="", age_val=None, dob_str=""):
    emp_name_clean = str(emp_name).strip() if pd.notna(emp_name) else ""
    if not emp_name_clean or emp_name_clean.lower() in ["none", "nan", "total", "summary", "select", ""]:
        return (99, 0, "")
        
    emp_lower = emp_name_clean.lower()

    # Known store bakers
    known_bakers = ["robert", "viet", "aroha"]
    
    # Store master ages for accurate intra-group ordering
    known_ages = {
        "jane": 50,
        "robert": 45,
        "viet": 35,
        "aroha": 32,
        "elizabeth": 28,
        "anastasia": 26,
        "jude": 25,
        "esther amataiti": 24,
        "esther": 24,
        "aimi": 20,
        "ainsley mactier": 19,
        "ainsley": 19,
        "stella": 17,
        "amy": 16,
        "jack": 15,
        "shaelyn": 15,
        "olivia": 15
    }

    # 1. Determine Team (0 = Baking Team, 1 = Service Team)
    team = None
    if any(b in emp_lower for b in known_bakers):
        team = "Baking Team"
    else:
        role_check = str(role_str).strip().lower() if role_str else ""
        if role_check:
            if any(kw in role_check for kw in ["baker", "baking", "pastry"]):
                team = "Baking Team"
            elif any(kw in role_check for kw in ["service", "senior team", "junior team", "sales", "retail"]):
                team = "Service Team"

    if not team:
        profiles = get_active_user_profiles()
        for u_key, u_data in profiles.items():
            ename = str(u_data.get("employee_name", u_key)).strip().lower()
            if ename == emp_lower or emp_lower in ename:
                prof = u_data.get("profile", {})
                prof_role = str(prof.get("employment_level", "")).lower()
                if any(kw in prof_role for kw in ["baker", "baking", "pastry"]):
                    team = "Baking Team"
                break

    if not team and "manual_employees" in st.session_state and isinstance(st.session_state.manual_employees, pd.DataFrame):
        e_df = st.session_state.manual_employees
        n_col = find_column(e_df, ["name", "employee", "staff"], "NAME")
        p_col = find_column(e_df, ["position", "role", "employment level", "team"], "position")
        for _, r in e_df.iterrows():
            r_name = str(r.get(n_col, "")).strip().lower()
            if r_name == emp_lower or emp_lower in r_name:
                r_pos = str(r.get(p_col, "")).lower()
                if any(kw in r_pos for kw in ["baker", "baking", "pastry"]):
                    team = "Baking Team"
                break

    if not team:
        team = "Service Team"

    team_code = 0 if team == "Baking Team" else 1

    # 2. Determine Age
    final_age = known_ages.get(emp_lower)
    
    if final_age is None and any(k in emp_lower for k in known_ages):
        for k, v in known_ages.items():
            if k in emp_lower:
                final_age = v
                break

    if final_age is None and age_val is not None and str(age_val).strip().replace('.', '', 1).isdigit():
        try:
            final_age = int(float(age_val))
        except:
            pass
            
    if final_age is None and dob_str:
        res = calculate_age_from_dob(dob_str)
        if res:
            final_age = res[0]

    if final_age is None:
        final_age = 21

    return (team_code, -final_age, emp_lower)

def sort_dataframe_by_team_and_age(df):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df

    df_copy = df.copy()

    emp_col = find_column(df_copy, ["employee", "name", "staff", "employee name", "staff name"], "")
    if not emp_col:
        if df_copy.columns[0] not in ["Shift", "Date", "Day"]:
            emp_col = df_copy.columns[0]
        else:
            return df_copy

    role_col = find_column(df_copy, ["position", "role", "employment level", "job", "title", "team"], "")
    age_col = find_column(df_copy, ["age"], "")
    dob_col = find_column(df_copy, ["dob", "date of birth", "birth date"], "")

    sort_keys = []
    for idx, row in df_copy.iterrows():
        emp_name = str(row.get(emp_col, "")).strip()
        role_val = str(row.get(role_col, "")) if role_col else ""
        age_val = row.get(age_col) if age_col else None
        dob_val = str(row.get(dob_col, "")) if dob_col else ""

        key = get_employee_team_and_age(emp_name, role_val, age_val, dob_val)
        sort_keys.append(key)

    df_copy["_sort_key"] = sort_keys
    df_copy = df_copy.sort_values(by="_sort_key").drop(columns=["_sort_key"]).reset_index(drop=True)
    return df_copy

def cleanup_duplicate_employee_columns(df):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    
    df = df.copy()
    
    concept_mappings = {
        "NAME": ["name", "employee", "staff", "staff name", "employee name"],
        "Team": ["team", "department", "group"],
        "DOB": ["dob", "date of birth", "birth date"],
        "Commencing Date": ["commencing date", "commence date", "commencement date", "start date", "started", "start"],
        "status": ["status", "employment type", "type", "classification", "employmenttype"],
        "position": ["position", "role", "employment level", "title", "job"]
    }
    
    matched_cols = {}
    for canonical, candidates in concept_mappings.items():
        matched_cols[canonical] = []
        for col in df.columns:
            if str(col).strip().lower() in [cand.lower() for cand in candidates]:
                matched_cols[canonical].append(col)
                
    new_rows = []
    for idx in df.index:
        row_dict = {}
        for canonical, source_cols in matched_cols.items():
            val = ""
            for sc in source_cols:
                v = str(df.at[idx, sc]).strip() if pd.notna(df.at[idx, sc]) else ""
                if v and v.lower() not in ["none", "nan", "nat", "null", ""]:
                    val = v
                    break
            row_dict[canonical] = val
            
        if row_dict["DOB"]:
            try:
                dt = pd.to_datetime(row_dict["DOB"], dayfirst=True, errors='coerce')
                if pd.notna(dt):
                    row_dict["DOB"] = dt.strftime('%d/%m/%Y')
            except:
                pass
                
        if row_dict["Commencing Date"]:
            try:
                dt = pd.to_datetime(row_dict["Commencing Date"], dayfirst=True, errors='coerce')
                if pd.notna(dt):
                    row_dict["Commencing Date"] = dt.strftime('%d/%m/%Y')
            except:
                pass

        if not row_dict.get("Team"):
            row_dict["Team"] = get_employee_team(row_dict.get("position", ""))

        new_rows.append(row_dict)

    res_df = pd.DataFrame(new_rows, columns=["NAME", "Team", "DOB", "Commencing Date", "status", "position"])
    res_df = sort_dataframe_by_team_and_age(res_df)
    return res_df

def sync_user_profiles_to_employees(emp_df):
    if emp_df is None:
        emp_df = pd.DataFrame(columns=["NAME", "Team", "DOB", "Commencing Date", "status", "position"])
    
    emp_df = cleanup_duplicate_employee_columns(emp_df)
    
    # Exclude demo accounts from employee dataframe
    if "NAME" in emp_df.columns:
        emp_df = emp_df[~emp_df["NAME"].astype(str).str.lower().str.contains("demo")].reset_index(drop=True)
    
    profiles = get_active_user_profiles()
    existing_names = [str(n).strip().lower() for n in emp_df["NAME"].tolist() if pd.notna(n)]
    
    new_rows = []
    for u_key, u_data in profiles.items():
        if u_key.startswith("demo.") or "demo" in u_key.lower():
            continue
        if u_data.get("role") == "Employee":
            emp_name = u_data.get("employee_name", u_key).strip()
            if "demo" in emp_name.lower():
                continue
            if emp_name and emp_name.lower() not in existing_names:
                prof = u_data.get("profile", {})
                comm_date = prof.get("commencement_date", datetime.now().strftime("%d/%m/%Y"))
                try:
                    dt = pd.to_datetime(comm_date, dayfirst=True, errors='coerce')
                    if pd.notna(dt):
                        comm_date = dt.strftime('%d/%m/%Y')
                except:
                    pass
                pos_val = str(prof.get("employment_level", "Service Staff"))
                new_rows.append({
                    "NAME": emp_name,
                    "Team": get_employee_team(pos_val),
                    "DOB": prof.get("dob", ""),
                    "Commencing Date": comm_date,
                    "status": str(prof.get("classification", "casual")).lower(),
                    "position": pos_val
                })
                existing_names.append(emp_name.lower())
                
    if new_rows:
        combined = pd.concat([emp_df, pd.DataFrame(new_rows)], ignore_index=True)
        emp_df = cleanup_duplicate_employee_columns(combined)
        
    emp_df = sort_dataframe_by_team_and_age(emp_df)
    save_persisted_df(emp_df, "employees.csv")
    return emp_df

def standardize_unavailability_df(df):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    
    df = df.copy()
    
    emp_col = find_column(df, ["employee", "name", "staff", "user", "person", "employee name", "staff name"], "Employee")
    day_col = find_column(df, ["day", "date", "weekday", "when", "day of week", "unavailable date"], "Day")
    win_col = find_column(df, ["time window", "window", "time", "unavailability", "reason", "constraint", "time constraint", "notes", "details"], "Time Window")

    weekday_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    for idx, row in df.iterrows():
        # 1. Clean Employee Name
        emp_val = str(row.get(emp_col, "")).strip()
        if emp_val and emp_val.lower() not in ["nan", "none", "nat", "null", ""]:
            if emp_val.islower():
                emp_val = emp_val.title()
            df.at[idx, emp_col] = emp_val
            
        # 2. Clean & Standardize Day / Date Column
        day_val = str(row.get(day_col, "")).strip()
        if day_val and day_val.lower() not in ["nan", "none", "nat", "null", ""]:
            dt = parse_date_robust(day_val)
            if dt:
                w_name = weekday_names[dt.weekday()]
                date_formatted = dt.strftime('%d/%m/%Y')
                if w_name.lower() not in day_val.lower():
                    df.at[idx, day_col] = f"{w_name} ({date_formatted})"
                else:
                    df.at[idx, day_col] = day_val
            else:
                matched_w = [w for w in weekday_names if w.lower() == day_val.lower()]
                if matched_w:
                    df.at[idx, day_col] = matched_w[0]
                    
        # 3. Clean Time Window Column
        win_val = str(row.get(win_col, "")).strip()
        if not win_val or win_val.lower() in ["nan", "none", "nat", "null", ""]:
            df.at[idx, win_col] = "All Day"
            
    return df

def clean_win_display(win_str):
    if not win_str or str(win_str).strip().lower() in ["nan", "none", "nat", "null", ""]:
        return "All Day"
    cleaned = re.sub(r'\(From.*?\)', '', str(win_str), flags=re.IGNORECASE).strip()
    if not cleaned or cleaned.lower() in ["nan", "none", "nat", "null", ""]:
        return "All Day"
    return cleaned

if 'manual_employees' not in st.session_state:
    default_emp = pd.DataFrame([
        {"NAME": "Aimi", "Team": "Service Staff", "DOB": "10/11/2006", "Commencing Date": "01/10/2023", "status": "casual", "position": "Service Staff"},
        {"NAME": "Ainsley Mactier", "Team": "Service Staff", "DOB": "14/08/2006", "Commencing Date": "04/10/2021", "status": "casual", "position": "Service Staff"},
        {"NAME": "Aroha", "Team": "Bakery Staff", "DOB": "24/05/2005", "Commencing Date": "27/09/2021", "status": "part time", "position": "baker assitant"},
        {"NAME": "Elizabeth", "Team": "Service Staff", "DOB": "30/07/2004", "Commencing Date": "03/06/2024", "status": "casual", "position": "Service Staff"},
        {"NAME": "Olivia", "Team": "Service Staff", "DOB": "15/01/2007", "Commencing Date": "27/05/2024", "status": "casual", "position": "Service Staff"},
        {"NAME": "Robert", "Team": "Bakery Staff", "DOB": "19/02/2004", "Commencing Date": "22/01/2024", "status": "part time", "position": "baker"},
        {"NAME": "Stella", "Team": "Service Staff", "DOB": "03/07/2007", "Commencing Date": "09/01/2024", "status": "casual", "position": "Service Staff"},
        {"NAME": "Violet", "Team": "Service Staff", "DOB": "27/02/2010", "Commencing Date": "25/05/2026", "status": "casual", "position": "Service Staff"},
        {"NAME": "Esther Amataiti", "Team": "Service Staff", "DOB": "20/09/2001", "Commencing Date": "20/09/2021", "status": "casual", "position": "Service Staff"},
        {"NAME": "Anastasia", "Team": "Service Staff", "DOB": "02/03/2000", "Commencing Date": "04/10/2021", "status": "casual", "position": "Service Staff"},
        {"NAME": "Jude", "Team": "Service Staff", "DOB": "28/04/2011", "Commencing Date": "27/07/2026", "status": "casual", "position": "Service Staff"},
        {"NAME": "Jack", "Team": "Service Staff", "DOB": "28/04/2011", "Commencing Date": "27/07/2026", "status": "casual", "position": "Service Staff"},
        {"NAME": "Jane", "Team": "Store Owners", "DOB": "", "Commencing Date": "", "status": "owner", "position": "Service Staff"},
        {"NAME": "Amy", "Team": "Service Staff", "DOB": "27/02/2010", "Commencing Date": "25/05/2026", "status": "casual", "position": "Service Staff"},
        {"NAME": "Viet", "Team": "Bakery Staff", "DOB": "", "Commencing Date": "", "status": "owner", "position": "baker"},
        {"NAME": "Shaelyn", "Team": "Service Staff", "DOB": "01/08/2011", "Commencing Date": "01/08/2026", "status": "casual", "position": "Service Staff"}
    ])
    st.session_state.manual_employees = sync_user_profiles_to_employees(load_persisted_df("employees.csv", default_emp))

default_unavail = pd.DataFrame([
    {"Employee": "Elizabeth", "Day": "Saturday", "Time Window": "All Day"},
    {"Employee": "Elizabeth", "Day": "Sunday", "Time Window": "All Day"},
    {"Employee": "Stella", "Day": "Monday", "Time Window": "Before 3:30pm"},
    {"Employee": "Stella", "Day": "Tuesday", "Time Window": "Before 3:30pm"},
    {"Employee": "Stella", "Day": "Thursday", "Time Window": "Before 3:30pm"},
    {"Employee": "Stella", "Day": "Friday", "Time Window": "Before 3:30pm"},
    {"Employee": "Ainsley Mactier", "Day": "Monday", "Time Window": "After 5:00pm"},
    {"Employee": "Ainsley Mactier", "Day": "Friday", "Time Window": "After 5:00pm"},
    {"Employee": "Jude", "Day": "Sunday", "Time Window": "Before 12:00pm"},
    {"Employee": "Jack", "Day": "Monday", "Time Window": "4:30pm-7:00pm"},
    {"Employee": "Jack", "Day": "Tuesday", "Time Window": "4:30pm-7:00pm"},
    {"Employee": "Jack", "Day": "Wednesday", "Time Window": "4:30pm-7:00pm"},
    {"Employee": "Jack", "Day": "Thursday", "Time Window": "4:30pm-7:00pm"},
    {"Employee": "Jack", "Day": "Friday", "Time Window": "2:30pm-5:30pm"},
    {"Employee": "Violet", "Day": "Saturday", "Time Window": "All Day"},
    {"Employee": "Violet", "Day": "Sunday", "Time Window": "All Day"},
    {"Employee": "Amy", "Day": "Monday", "Time Window": "All Day"},
    {"Employee": "Amy", "Day": "Tuesday", "Time Window": "All Day"},
    {"Employee": "Amy", "Day": "Wednesday", "Time Window": "All Day"},
    {"Employee": "Amy", "Day": "Thursday", "Time Window": "All Day"},
    {"Employee": "Amy", "Day": "Friday", "Time Window": "All Day"},
    {"Employee": "Olivia", "Day": "Monday", "Time Window": "All Day"},
    {"Employee": "Olivia", "Day": "Tuesday", "Time Window": "All Day"},
    {"Employee": "Olivia", "Day": "Wednesday", "Time Window": "All Day"},
    {"Employee": "Olivia", "Day": "Thursday", "Time Window": "All Day"},
    {"Employee": "Olivia", "Day": "Friday", "Time Window": "All Day"}
])

if 'manual_unavailability' not in st.session_state or st.session_state.manual_unavailability is None or st.session_state.manual_unavailability.empty:
    st.session_state.manual_unavailability = sort_dataframe_by_team_and_age(standardize_unavailability_df(load_persisted_df("unavailability.csv", default_unavail)))
    save_persisted_df(st.session_state.manual_unavailability, "unavailability.csv")

if 'manual_requirements' not in st.session_state or st.session_state.manual_requirements is None or len(st.session_state.manual_requirements) <= 2:
    default_req = pd.DataFrame([
        {"Shift": "12:00pm-5:00pm", "Monday": "1", "Tuesday": "0", "Wednesday": "0", "Thursday": "0", "Friday": "0", "Saturday": "0", "Sunday": "0"},
        {"Shift": "7:00am-10:00am", "Monday": "0", "Tuesday": "0", "Wednesday": "0", "Thursday": "0", "Friday": "0", "Saturday": "1", "Sunday": "0"},
        {"Shift": "7:00am-12:00pm", "Monday": "1", "Tuesday": "1", "Wednesday": "1", "Thursday": "1", "Friday": "1", "Saturday": "0", "Sunday": "0"},
        {"Shift": "7:30am-10:30am", "Monday": "0", "Tuesday": "0", "Wednesday": "0", "Thursday": "0", "Friday": "0", "Saturday": "1", "Sunday": "0"},
        {"Shift": "7:30am-12:30pm", "Monday": "1", "Tuesday": "1", "Wednesday": "1", "Thursday": "0", "Friday": "0", "Saturday": "0", "Sunday": "0"},
        {"Shift": "7:30am-3:30pm", "Monday": "0", "Tuesday": "0", "Wednesday": "0", "Thursday": "0", "Friday": "0", "Saturday": "0", "Sunday": "0"},
        {"Shift": "8:30am-1:30pm", "Monday": "0", "Tuesday": "0", "Wednesday": "0", "Thursday": "0", "Friday": "0", "Saturday": "1", "Sunday": "0"},
        {"Shift": "9:00am-5:00pm", "Monday": "0", "Tuesday": "0", "Wednesday": "0", "Thursday": "0", "Friday": "1", "Saturday": "0", "Sunday": "0"},
        {"Shift": "11:30am-2:30pm", "Monday": "0", "Tuesday": "0", "Wednesday": "0", "Thursday": "0", "Friday": "0", "Saturday": "0", "Sunday": "1"},
        {"Shift": "12:00pm-7:00pm", "Monday": "1", "Tuesday": "0", "Wednesday": "1", "Thursday": "1", "Friday": "1", "Saturday": "0", "Sunday": "0"},
        {"Shift": "12:30pm-5:30pm", "Monday": "0", "Tuesday": "1", "Wednesday": "0", "Thursday": "0", "Friday": "0", "Saturday": "1", "Sunday": "1"},
        {"Shift": "4:00pm-7:00pm", "Monday": "0", "Tuesday": "1", "Wednesday": "0", "Thursday": "1", "Friday": "0", "Saturday": "0", "Sunday": "0"},
        {"Shift": "2:30pm-5:30pm", "Monday": "0", "Tuesday": "0", "Wednesday": "0", "Thursday": "0", "Friday": "0", "Saturday": "0", "Sunday": "1"}
    ])
    st.session_state.manual_requirements = load_persisted_df("requirements.csv", default_req)

if 'manual_fixed' not in st.session_state or st.session_state.manual_fixed is None or len(st.session_state.manual_fixed) <= 2:
    default_fixed = pd.DataFrame([
        {"Employee": "Viet Nguyen", "Monday": "off", "Tuesday": "4:00am-12:00pm", "Wednesday": "off", "Thursday": "5:30am-12:30pm", "Friday": "5:30am-12:30pm", "Saturday": "4:00am-12:00pm", "Sunday": "5:30am-12:30pm"},
        {"Employee": "Anastasia", "Monday": "12:00am-5:00pm", "Tuesday": "off", "Wednesday": "off", "Thursday": "", "Friday": "9:00am-5:00pm", "Saturday": "12:30pm-5:30pm", "Sunday": "12:30pm-5:30pm"},
        {"Employee": "Esther Amataiti", "Monday": "7:00am-12:00pm", "Tuesday": "7:00am-12:00pm", "Wednesday": "off", "Thursday": "7:30am-12:30pm", "Friday": "7:00am-12:00pm", "Saturday": "off", "Sunday": "off"},
        {"Employee": "Jane", "Monday": "off", "Tuesday": "12:00pm-5:00pm", "Wednesday": "12:30pm-5:30pm", "Thursday": "12:30pm-5:30pm", "Friday": "off", "Saturday": "10:00am-3:00pm", "Sunday": "9:00am-2:00am"},
        {"Employee": "Amy", "Monday": "unavailable", "Tuesday": "unavailable", "Wednesday": "unavailable", "Thursday": "unavailable", "Friday": "unavailable", "Saturday": "", "Sunday": ""},
        {"Employee": "Olivia", "Monday": "unavailable", "Tuesday": "unavailable", "Wednesday": "unavailable", "Thursday": "unavailable", "Friday": "unavailable", "Saturday": "", "Sunday": ""},
        {"Employee": "Aroha", "Monday": "6:00am-1:00pm", "Tuesday": "6:00am-1:00pm", "Wednesday": "6:00am-1:00pm", "Thursday": "off", "Friday": "off", "Saturday": "6:00am-2:00pm", "Sunday": "6:00am-11:00am"},
        {"Employee": "Robert", "Monday": "4:00am-12:00pm", "Tuesday": "off", "Wednesday": "4:00am-12:00pm", "Thursday": "4:00pm-12:00pm", "Friday": "4:00am-12:00pm", "Saturday": "off", "Sunday": "4:00am-12:00pm"}
    ])
    st.session_state.manual_fixed = sort_dataframe_by_team_and_age(load_persisted_df("fixed.csv", default_fixed))
if st.session_state.manual_fixed is not None:
    st.session_state.manual_fixed = st.session_state.manual_fixed.replace(["off", "Off", "OFF", "None", "none", "nan", "NaN", None], "")
def render_store_kiosk_timeclock():
    st.markdown("""
    <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 22px 26px; border-radius: 16px; color: #ffffff !important; border: 2.5px solid #e5a93c; margin-bottom: 24px; text-align: center; box-shadow: 0 8px 24px rgba(0,0,0,0.5);">
        <h2 style="margin: 0; color: #e5a93c; font-size: 1.8rem; font-weight: 900;">🥐 Brumby's Bakery Pakenham — Shared Store Timeclock</h2>
        <p style="margin: 6px 0 0 0; color: #d0e6df; font-size: 1.05rem; font-weight: 600;">Store Terminal Mode • Select your name to Clock In or Clock Out</p>
    </div>
    """, unsafe_allow_html=True)

    today_dt = datetime.utcnow() + timedelta(hours=10)
    today_str = today_dt.strftime("%d/%m/%Y")
    day_name = today_dt.strftime("%A")

    st.markdown(f"""
    <div style="background: #0c2b25; padding: 14px 20px; border-radius: 12px; border: 1.5px solid #1f5c50; text-align: center; margin-bottom: 20px;">
        <div style="font-size: 2.2rem; font-weight: 900; color: #e5a93c; letter-spacing: 1px;">{today_dt.strftime('%I:%M:%S %p')}</div>
        <div style="font-size: 0.95rem; color: #a0aec0; font-weight: 700; margin-top: 2px;">🗓️ {day_name}, {today_str} (Melbourne AEST)</div>
    </div>
    """, unsafe_allow_html=True)

    if "kiosk_success_msg" in st.session_state and st.session_state.kiosk_success_msg:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #1c4532 0%, #22543d 100%); border: 2px solid #48bb78; color: #ffffff; padding: 18px 24px; border-radius: 14px; text-align: center; font-size: 1.2rem; font-weight: 800; margin-bottom: 24px; box-shadow: 0 6px 20px rgba(72,187,120,0.4);">
            {st.session_state.kiosk_success_msg}
        </div>
        """, unsafe_allow_html=True)
        st.session_state.kiosk_success_msg = None

    df_emp = load_persisted_df("employees.csv")
    emp_list = []
    if df_emp is not None and not df_emp.empty:
        n_col = find_column(df_emp, ["name", "employee", "staff"], "NAME")
        if n_col in df_emp.columns:
            for val in df_emp[n_col].dropna().tolist():
                name_str = str(val).strip()
                if name_str and name_str.lower() not in ["none", "nan", "total", "summary", ""] and "demo" not in name_str.lower():
                    if name_str not in emp_list:
                        emp_list.append(name_str)

    for u_k, u_v in user_profiles.items():
        if u_k.startswith("demo.") or "demo" in u_k.lower():
            continue
        ename = u_v.get("employee_name", u_k).strip()
        if ename and ename.lower() not in ["viet", "jane"] and ename not in emp_list:
            emp_list.append(ename)

    emp_list = sorted(list(set(emp_list)))
    
    if "reset_kiosk_emp" not in st.session_state:
        st.session_state.reset_kiosk_emp = False

    if st.session_state.reset_kiosk_emp:
        st.session_state["kiosk_selected_emp"] = "-- Select Your Name --"
        st.session_state.reset_kiosk_emp = False

    st.markdown("#### 👤 Select Your Name to Punch In / Out")
    selected_emp = st.selectbox(
        "Choose Employee:",
        options=["-- Select Your Name --"] + emp_list,
        key="kiosk_selected_emp"
    )

    if not selected_emp or selected_emp == "-- Select Your Name --":
        st.info("👇 Please select your name from the dropdown menu above to view status and clock in/out.")
        return

    df_cards = load_persisted_timecards()
    today_punch = None
    if df_cards is not None and not df_cards.empty and "Date" in df_cards.columns:
        for idx, r in df_cards.iterrows():
            if str(r.get("Date", "")).strip() == today_str and str(r.get("Employee", "")).strip().lower() == selected_emp.strip().lower():
                today_punch = r.to_dict()
                break

    scheduled_shift = "7:00am-3:30pm"
    past_rosters = list_finalized_rosters()
    if past_rosters:
        target_dt = parse_date_robust(today_str)
        matching_rosters = []
        if target_dt:
            for r_item in past_rosters:
                s_dt = r_item.get("start_date")
                if s_dt and (s_dt <= target_dt <= s_dt + timedelta(days=6)):
                    matching_rosters.append(r_item)
                    
        rosters_to_scan = matching_rosters if matching_rosters else past_rosters
        for r_item in rosters_to_scan:
            r_df = load_finalized_roster(r_item["csv_filename"])
            if r_df is not None and not r_df.empty and day_name in r_df.columns:
                emp_col = find_column(r_df, ["name", "employee", "staff"])
                if emp_col in r_df.columns:
                    for _, r in r_df.iterrows():
                        if find_matching_employee(selected_emp, {str(r.get(emp_col, "")).strip().lower(): str(r.get(emp_col, "")).strip()}):
                            val = str(r.get(day_name, "")).strip()
                            if val and val.lower() not in ["off", "nan", "unavailable"]:
                                scheduled_shift = val
                                break
            if scheduled_shift != "7:00am-3:30pm":
                break

    c_in = today_punch.get("Clock In", "") if today_punch else ""
    c_out = today_punch.get("Clock Out", "") if today_punch else ""

    st.markdown(f"""
    <div style="background: #081d19; padding: 20px; border-radius: 14px; border: 1.5px solid #1f5c50; margin-bottom: 20px;">
        <h3 style="color: #e5a93c; margin: 0 0 10px 0;">👤 Staff Member: {selected_emp}</h3>
        <p style="color: #d0e6df; margin: 4px 0; font-size: 1rem;"><b>Scheduled Shift Today:</b> <code>{scheduled_shift}</code></p>
    """, unsafe_allow_html=True)

    if c_in and not c_out:
        st.markdown(f"<p style='color: #48bb78; font-size: 1.15rem; font-weight: 800; margin: 6px 0;'>🟢 Status: WORKING NOW (Clocked IN at {c_in})</p></div>", unsafe_allow_html=True)
    elif c_in and c_out:
        st.markdown(f"<p style='color: #4299e1; font-size: 1.15rem; font-weight: 800; margin: 6px 0;'>✅ Status: SHIFT COMPLETED ({c_in} - {c_out})</p></div>", unsafe_allow_html=True)
    else:
        st.markdown(f"<p style='color: #fc8181; font-size: 1.15rem; font-weight: 800; margin: 6px 0;'>🔴 Status: CLOCKED OUT</p></div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col_in, col_out = st.columns(2)

    melbourne_now = datetime.utcnow() + timedelta(hours=10)

    with col_in:
        is_in_disabled = bool(c_in and not c_out)
        btn_in = st.button("🟢 CLOCK IN NOW", key=f"kiosk_btn_in_{selected_emp}", use_container_width=True, disabled=is_in_disabled)
        if btn_in:
            clock_in_time_str = melbourne_now.strftime("%I:%M %p")
            rec_id = f"TC_{today_str.replace('/', '')}_{selected_emp.replace(' ', '')}"
            loc_badge = "✅ Verified via Store Terminal (Brumby's Bakery Pakenham)"
            
            new_rec = {
                "Record ID": rec_id,
                "Date": today_str,
                "Employee": selected_emp,
                "Scheduled Shift": scheduled_shift,
                "Clock In": clock_in_time_str,
                "Clock Out": "",
                "Net Hours": "0",
                "Variance (Mins)": "0",
                "GPS Lat": "",
                "GPS Lon": "",
                "Distance (m)": "0.0",
                "Location Verification": loc_badge,
                "Note": "✅ Store Terminal Punch",
                "Late Correction Status": "Normal",
                "Status": "Working"
            }
            
            if df_cards is None or df_cards.empty:
                df_updated = pd.DataFrame([new_rec])
            else:
                df_cards = df_cards[df_cards["Record ID"] != rec_id]
                df_updated = pd.concat([df_cards, pd.DataFrame([new_rec])], ignore_index=True)
                
            save_timecard_records(df_updated)
            st.session_state.kiosk_success_msg = f"✅ Welcome {selected_emp}! Successfully Clocked IN at {clock_in_time_str} via Store Terminal."
            st.session_state.reset_kiosk_emp = True
            st.rerun()

    with col_out:
        is_out_disabled = bool(not c_in or c_out)
        btn_out = st.button("🔴 CLOCK OUT NOW", key=f"kiosk_btn_out_{selected_emp}", use_container_width=True, disabled=is_out_disabled)
        if btn_out:
            clock_out_time_str = melbourne_now.strftime("%I:%M %p")
            clock_in_str = today_punch.get("Clock In", "") if today_punch else ""
            
            c_in_dec = parse_time_to_decimal(clock_in_str)
            c_out_dec = parse_time_to_decimal(clock_out_time_str)
            net_h = round(c_out_dec - c_in_dec, 2) if c_out_dec > c_in_dec else 0.0
            
            if today_punch:
                today_punch["Clock Out"] = clock_out_time_str
                today_punch["Net Hours"] = str(net_h)
                today_punch["Status"] = "Completed"
                rec_id = today_punch.get("Record ID")
                df_cards = df_cards[df_cards["Record ID"] != rec_id]
                df_updated = pd.concat([df_cards, pd.DataFrame([today_punch])], ignore_index=True)
            else:
                rec_id = f"TC_{today_str.replace('/', '')}_{selected_emp.replace(' ', '')}"
                loc_badge = "✅ Verified via Store Terminal (Brumby's Bakery Pakenham)"
                new_rec = {
                    "Record ID": rec_id,
                    "Date": today_str,
                    "Employee": selected_emp,
                    "Scheduled Shift": scheduled_shift,
                    "Clock In": clock_out_time_str,
                    "Clock Out": clock_out_time_str,
                    "Net Hours": "0",
                    "Variance (Mins)": "0",
                    "GPS Lat": "",
                    "GPS Lon": "",
                    "Distance (m)": "0.0",
                    "Location Verification": loc_badge,
                    "Note": "✅ Store Terminal Punch",
                    "Late Correction Status": "Normal",
                    "Status": "Completed"
                }
                df_updated = pd.concat([df_cards, pd.DataFrame([new_rec])], ignore_index=True)

            save_timecard_records(df_updated)
            st.session_state.kiosk_success_msg = f"✅ Goodbye {selected_emp}! Successfully Clocked OUT at {clock_out_time_str} (Total: {net_h} hrs)."
            st.session_state.reset_kiosk_emp = True
            st.rerun()

# --- ROLE-BASED TAB NAVIGATION ---
is_kiosk = (st.session_state.user_role == "Kiosk" or st.session_state.get("is_kiosk_mode", False))
is_manager = (st.session_state.user_role == "Manager")

if is_kiosk:
    render_store_kiosk_timeclock()
    st.stop()
elif is_manager:
    tab_dash, tab_roster, tab_gen, tab_emp, tab_unavail, tab_req, tab_fixed, tab_timesheets = st.tabs([
        "🏠 Home / Dashboard",
        "📅 Roster Inside",
        "⚡ Weekly Roster Generator",
        "👥 Staff Members", 
        "🚫 Unavailability", 
        "📋 Daily Requirements", 
        "📌 Fixed Shifts",
        "⏱️ Shift Timesheet Audit & Live Attendance"
    ])
else:
    # Employee sees 3 tabs: Current Roster (1st), Personal Information (2nd), Availability (3rd)
    tab_my_current_roster, tab_my_info, tab_my_avail = st.tabs([
        "📅 Current Roster",
        "📋 Personal Information Form",
        "📅 My Availability & Constraints"
    ])

def render_employee_current_roster_tab(user_key):
    user_info = user_profiles.get(user_key, {})
    emp_name = user_info.get("employee_name", user_key)
    
    today = datetime.now().date()
    
    past_rosters = list_finalized_rosters()
    matched_roster = None
    matched_start_date = None
    
    # Compare current date against roster week periods (Monday to Sunday)
    for r in past_rosters:
        try:
            s_dt = datetime.strptime(r["date_str"], "%Y-%m-%d").date()
            e_dt = s_dt + timedelta(days=6)
            if s_dt <= today <= e_dt:
                matched_roster = r
                matched_start_date = s_dt
                break
        except:
            pass
            
    # Fallback to most recent finalized roster if no exact match for current week
    if not matched_roster and past_rosters:
        matched_roster = past_rosters[0]
        try:
            matched_start_date = datetime.strptime(matched_roster["date_str"], "%Y-%m-%d").date()
        except:
            matched_start_date = today
            
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0e2b26 0%, #1a4d43 100%); padding: 20px; border-radius: 16px; border: 2px solid #e5a93c; box-shadow: 0 8px 24px rgba(0,0,0,0.4); margin-bottom: 20px;">
        <h2 style="color: #f7d594 !important; margin-top: 0; font-size: 1.6rem; font-weight: 800;">📅 Bakery Weekly Staff Roster</h2>
        <p style="color: #ffffff !important; font-size: 0.95rem; margin-bottom: 0;">Welcome! View your scheduled shifts and General Retail Industry Award 2020 break entitlements below.</p>
    </div>
    """, unsafe_allow_html=True)
    
    if matched_roster:
        roster_df = load_finalized_roster(matched_roster["csv_filename"])
        if roster_df is not None and not roster_df.empty:
            end_date = matched_start_date + timedelta(days=6)
            
            # Header week badge
            st.markdown(f"""
            <div style="background: rgba(9, 32, 28, 0.6); padding: 10px 16px; border-radius: 10px; border-left: 4px solid #e5a93c; margin-bottom: 15px;">
                <span style="color: #e5a93c; font-weight: 800; font-size: 1.05rem;">🗓️ Week Period:</span>
                <span style="color: #ffffff; font-weight: 700; font-size: 1.05rem; margin-left: 8px;">Monday {matched_start_date.strftime('%d/%m/%Y')} — Sunday {end_date.strftime('%d/%m/%Y')}</span>
            </div>
            """, unsafe_allow_html=True)
            
            # Personalized Employee Shift Callout Summary
            emp_c = find_column(roster_df, ["employee", "name", "staff"], "Employee")
            if emp_c in roster_df.columns:
                matched_rows = roster_df[roster_df[emp_c].astype(str).str.strip().str.lower() == emp_name.strip().lower()]
                if not matched_rows.empty:
                    emp_row = matched_rows.iloc[0]
                    shifts_list = []
                    days_arr = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                    for d in days_arr:
                        if d in emp_row:
                            val = str(emp_row[d]).strip()
                            if val and val.lower() not in ["off", "unavailable", " unavailable", "none", "nan"]:
                                shifts_list.append(f"<b>{d}:</b> {val}")
                    
                    if shifts_list:
                        shifts_str = " &nbsp;|&nbsp; ".join(shifts_list)
                        st.markdown(f"""
                        <div style="background: rgba(229, 169, 60, 0.15); border: 1px solid #e5a93c; padding: 12px 18px; border-radius: 12px; margin-bottom: 20px;">
                            <div style="color: #f7d594; font-weight: 800; font-size: 1.05rem;">👋 Hello {emp_name}, your shifts for this week:</div>
                            <div style="color: #ffffff; font-size: 0.95rem; margin-top: 6px;">{shifts_str}</div>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.markdown(f"""
                        <div style="background: rgba(229, 169, 60, 0.1); border: 1px solid rgba(229, 169, 60, 0.4); padding: 12px 18px; border-radius: 12px; margin-bottom: 20px;">
                            <div style="color: #f7d594; font-weight: 700; font-size: 0.95rem;">👋 Hello {emp_name}, you have no shifts scheduled for this week.</div>
                        </div>
                        """, unsafe_allow_html=True)

            # Store Announcements Display for Employee Portal
            announcements = load_announcements()
            if announcements:
                ann = announcements[0]
                prio = ann.get("priority", "Normal")
                badge_bg = "#e53e3e" if prio == "Urgent" else ("#dd6b20" if prio == "Important" else "#319795")
                st.markdown(f"""
                <div style="background: rgba(8, 29, 25, 0.95); border: 1.5px solid #e5a93c; border-radius: 12px; padding: 16px; margin-bottom: 20px;">
                    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 6px;">
                        <span style="font-weight: 800; font-size: 1.1rem; color: #f7d594;">📢 Store Announcement: {ann.get('title', '')}</span>
                        <span style="background: {badge_bg}; color: white; padding: 2px 10px; border-radius: 12px; font-size: 0.78rem; font-weight: 800;">{prio}</span>
                    </div>
                    <div style="color: #ffffff; font-size: 0.95rem; margin-bottom: 8px;">{ann.get('content', '')}</div>
                    <div style="color: #a0aec0; font-size: 0.8rem;">✍️ Posted by <b>{ann.get('author', 'Store Owner')}</b> on {ann.get('date', '')}</div>
                </div>
                """, unsafe_allow_html=True)

            # Full Schedule Display
            st.markdown("### 📋 Full Team Schedule")
            st.dataframe(roster_df, use_container_width=True)

            # Download XLSX Button
            excel_bytes = build_roster_excel_bytes(roster_df, matched_start_date)
            dl_filename = f"Team_Roster_{matched_start_date.strftime('%d.%m.%Y')}.xlsx"
            st.download_button(
                label=f"📥 Download Roster for {matched_start_date.strftime('%d/%m/%Y')} (.XLSX)",
                data=excel_bytes,
                file_name=dl_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_emp_dl_current_roster",
                use_container_width=True
            )

            # Option to select other past weeks
            if len(past_rosters) > 1:
                st.markdown("<br><hr>", unsafe_allow_html=True)
                st.markdown("**🔍 Select Other Roster Weeks to View:**")
                roster_opts = {r["label"]: r for r in past_rosters}
                sel_label = st.selectbox("View another week's schedule:", list(roster_opts.keys()), key="emp_select_other_roster")
                if sel_label != matched_roster["label"]:
                    sel_r = roster_opts[sel_label]
                    other_df = load_finalized_roster(sel_r["csv_filename"])
                    if other_df is not None:
                        st.markdown(f"**Viewing Schedule for: `{sel_label}`**")
                        st.dataframe(other_df, use_container_width=True)
    else:
        st.info("ℹ️ No finalized rosters have been published yet by management. Please check back soon!")

    # --- GENERAL RETAIL INDUSTRY AWARD 2020 BREAK RULES SECTION ---
    st.markdown("<br><hr>", unsafe_allow_html=True)
    st.markdown("""
    <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 14px 20px; border-radius: 12px; border: 1px solid rgba(229, 169, 60, 0.4); margin-top: 15px; margin-bottom: 15px;">
        <h3 style="color: #e5a93c !important; margin: 0; font-size: 1.25rem; font-weight: 800;">☕ General Retail Industry Award 2020 — Shift Break Entitlements</h3>
    </div>
    """, unsafe_allow_html=True)
    
    col_b1, col_b2 = st.columns([1.5, 1])
    with col_b1:
        break_table_html = """
        <table style="width:100%; border-collapse: collapse; margin-top: 5px; font-size: 0.9rem; color: #ffffff;">
            <thead>
                <tr style="background-color: #1F4E78; text-align: center; font-weight: bold; color: #ffffff;">
                    <th style="padding: 8px; border: 1px solid #336699;">Shift Duration</th>
                    <th style="padding: 8px; border: 1px solid #336699;">Paid Rest Break (10m)</th>
                    <th style="padding: 8px; border: 1px solid #336699;">Unpaid Meal Break (30m)</th>
                    <th style="padding: 8px; border: 1px solid #336699;">Total Entitlement</th>
                </tr>
            </thead>
            <tbody>
                <tr style="background-color: rgba(255,255,255,0.05); text-align: center;">
                    <td style="padding: 8px; border: 1px solid #2a5a50;">Less than 4 hours</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">No break</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">No break</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">No breaks required</td>
                </tr>
                <tr style="background-color: rgba(255,255,255,0.02); text-align: center;">
                    <td style="padding: 8px; border: 1px solid #2a5a50;">4 hrs up to 5 hrs</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">1 x 10 min</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">No break</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">10 min Paid</td>
                </tr>
                <tr style="background-color: rgba(255,255,255,0.05); text-align: center;">
                    <td style="padding: 8px; border: 1px solid #2a5a50;">5 hrs up to 7 hrs</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">1 x 10 min</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">1 x 30 min</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">10m Paid + 30m Unpaid</td>
                </tr>
                <tr style="background-color: rgba(255,255,255,0.02); text-align: center;">
                    <td style="padding: 8px; border: 1px solid #2a5a50;">7 hrs up to 10 hrs</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">2 x 10 min</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">1 x 30 min</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">20m Paid + 30m Unpaid</td>
                </tr>
                <tr style="background-color: rgba(255,255,255,0.05); text-align: center;">
                    <td style="padding: 8px; border: 1px solid #2a5a50;">10 hours or more</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">2 x 10 min</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">2 x 30 min</td>
                    <td style="padding: 8px; border: 1px solid #2a5a50;">20m Paid + 60m Unpaid</td>
                </tr>
            </tbody>
        </table>
        """
        st.markdown(break_table_html, unsafe_allow_html=True)
        
    with col_b2:
        st.markdown("""
        <div style="background: rgba(9, 32, 28, 0.6); border: 1px solid rgba(229, 169, 60, 0.4); border-radius: 12px; padding: 15px; height: 100%;">
            <h4 style="color: #e5a93c !important; margin-top: 0;">📌 Key Shift Break Rules</h4>
            <ul style="margin-bottom: 0; padding-left: 18px; font-size: 0.88rem; color: #ffffff !important;">
                <li><b>Paid vs Unpaid:</b> Rest breaks (10 min) are paid by employer. Meal breaks (30 min) are unpaid.</li>
                <li><b>Start & End Buffer:</b> No break may be taken during the first or last hour of work.</li>
                <li><b>5-Hour Limit:</b> An unpaid meal break must be taken no later than after 5 hours of continuous work.</li>
                <li><b>No Combining:</b> Rest breaks and meal breaks cannot be combined together.</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

# Helper function to render Confidential Profile Form
def render_confidential_profile_form(user_key, is_admin=False):
    user_data = user_profiles.get(user_key, {})
    prof = user_data.get("profile", {})
    emp_name = user_data.get("employee_name", user_key)
    
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 16px 24px; border-radius: 14px; border: 2px solid #e5a93c; margin-bottom: 20px;">
        <h2 style="color: #e5a93c !important; margin: 0;">📜 Employee Confidential Information Form</h2>
        <p style="color: #c8e6e0 !important; margin-top: 4px; margin-bottom: 0;">Employee: <b>{emp_name}</b> &nbsp;|&nbsp; Store: <b>{prof.get('store', "Brumby's Pakenham")}</b></p>
    </div>
    """, unsafe_allow_html=True)
    
    with st.form(key=f"form_profile_{user_key}"):
        st.markdown("### 1. 👤 Employee Personal Details")
        c1, c2 = st.columns(2)
        with c1:
            full_name = st.text_input("Full Name", value=str(prof.get("full_name", emp_name)))
            address = st.text_area("Home Address", value=str(prof.get("address", "")), height=100)
            home_phone = st.text_input("Home Phone Number", value=str(prof.get("home_phone", "")))
            mobile = st.text_input("Mobile Phone Number", value=str(prof.get("mobile", "")))
            email = st.text_input("Email Address", value=str(prof.get("email", "")))
            dob = st.text_input("Date of Birth", value=str(prof.get("dob", "")))
        with c2:
            gender_options = ["Female", "Male", "Other", "Prefer not to say"]
            curr_gender = str(prof.get("gender", "Female"))
            g_idx = gender_options.index(curr_gender) if curr_gender in gender_options else 0
            gender = st.selectbox("Gender", gender_options, index=g_idx)
            
            tfn = st.text_input("Tax File Number (TFN)", value=str(prof.get("tfn", "")))
            store = st.text_input("Store Location", value=str(prof.get("store", "Brumby's Pakenham")))
            
            class_options = ["Casual", "Part-Time", "Full-Time"]
            curr_class = str(prof.get("classification", "Casual"))
            c_idx = class_options.index(curr_class) if curr_class in class_options else 0
            classification = st.selectbox("Employment Classification", class_options, index=c_idx)
            
            commencement_date = st.text_input("Commencement Date", value=str(prof.get("commencement_date", "")))
            employment_level = st.text_input("Employment Level / Role", value=str(prof.get("employment_level", "")))

        st.markdown("---")
        st.markdown("### 2. 🏦 Bank Account Details")
        b1, b2 = st.columns(2)
        with b1:
            bank_name = st.text_input("Bank Name", value=str(prof.get("bank_name", "")))
            bank_branch = st.text_input("Branch Location", value=str(prof.get("bank_branch", "")))
            account_name = st.text_input("Name on Account", value=str(prof.get("account_name", "")))
        with b2:
            bank_bsb = st.text_input("BSB Number (e.g. 062 948)", value=str(prof.get("bank_bsb", "")))
            bank_account = st.text_input("Account Number", value=str(prof.get("bank_account", "")))

        st.markdown("---")
        st.markdown("### 3. 📈 Superannuation Details")
        s1, s2 = st.columns(2)
        with s1:
            super_fund = st.text_input("Name of Super Fund", value=str(prof.get("super_fund", "")))
            super_policy = st.text_input("Policy / Membership Number", value=str(prof.get("super_policy", "")))
            super_address = st.text_input("Super Fund Address", value=str(prof.get("super_address", "")))
        with s2:
            super_contact = st.text_input("Fund Contact Number", value=str(prof.get("super_contact", "")))
            super_abn = st.text_input("Fund ABN / SPIN / USI", value=str(prof.get("super_abn", "")))

        submit_btn = st.form_submit_button("💾 Save Confidential Profile Information")
        
        if submit_btn:
            user_profiles[user_key]["profile"] = {
                "full_name": full_name,
                "address": address,
                "home_phone": home_phone,
                "mobile": mobile,
                "email": email,
                "dob": dob,
                "gender": gender,
                "tfn": tfn,
                "store": store,
                "classification": classification,
                "commencement_date": commencement_date,
                "employment_level": employment_level,
                "bank_name": bank_name,
                "bank_branch": bank_branch,
                "account_name": account_name,
                "bank_bsb": bank_bsb,
                "bank_account": bank_account,
                "super_fund": super_fund,
                "super_policy": super_policy,
                "super_address": super_address,
                "super_contact": super_contact,
                "super_abn": super_abn
            }
            save_user_profiles(user_profiles)
            st.success("✅ Profile information updated and saved successfully!")

# Helper function for Employee Availability Management & Auto-Sync
def render_employee_availability_manager(user_key):
    user_data = user_profiles.get(user_key, {})
    emp_name = user_data.get("employee_name", user_key)
    
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 16px 24px; border-radius: 14px; border: 2px solid #e5a93c; margin-bottom: 20px;">
        <h2 style="color: #e5a93c !important; margin: 0;">📅 My Unavailability & Constraint Manager</h2>
        <p style="color: #c8e6e0 !important; margin-top: 4px; margin-bottom: 0;">Logged in Employee: <b>{emp_name}</b>. Follow the 2-Step form below to log your unavailable days & date range.</p>
    </div>
    """, unsafe_allow_html=True)
    
    from datetime import date
    days_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    with st.form(key=f"form_2step_unavail_{user_key}"):
        st.markdown("### Step 1: Select Unavailability per Day (7 Day Rows)")
        
        day_inputs = {}
        for d in days_list:
            col_check, col_type, col_detail = st.columns([1.2, 1.2, 2.5])
            with col_check:
                is_checked = st.checkbox(f"🔴 {d}", key=f"chk_{d}_{user_key}")
            with col_type:
                time_type = st.radio(f"Time for {d}", ["All Day", "Specific Time"], key=f"rad_{d}_{user_key}", label_visibility="collapsed")
            with col_detail:
                # Input text box ALWAYS visible for every row
                spec_time_input = st.text_input(
                    f"Window for {d}", 
                    value="", 
                    placeholder="Type window e.g. Before 3:30pm, After 5pm, 7:30am-12:30pm", 
                    key=f"input_spec_{d}_{user_key}", 
                    label_visibility="collapsed"
                ).strip()
                
                if time_type == "Specific Time":
                    spec_time = spec_time_input if spec_time_input else "Specific Time"
                else:
                    spec_time = "All Day"
                    
            day_inputs[d] = {"checked": is_checked, "time_type": time_type, "spec_time": spec_time}
            st.markdown("<hr style='margin: 2px 0; border-color: rgba(255,255,255,0.08);'>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### Step 2: Input Date Range")
        col_start, col_end = st.columns(2)
        with col_start:
            start_date = st.date_input("Start Date", date.today(), key=f"step2_start_{user_key}")
        with col_end:
            end_date = st.date_input("End Date", date.today() + timedelta(days=90), key=f"step2_end_{user_key}")

        submit_2step = st.form_submit_button("📌 Save Unavailability Constraints")
        
        if submit_2step:
            checked_days = [d for d in days_list if day_inputs[d]["checked"]]
            if not checked_days:
                st.warning("⚠️ Please check at least one day row in Step 1 before saving.")
            else:
                date_note = f"From {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
                df_curr = st.session_state.manual_unavailability.copy()
                emp_col = find_column(df_curr, ["employee", "name", "staff"], "Employee")
                day_col = find_column(df_curr, ["day", "date", "weekday"], "Day")
                win_col = find_column(df_curr, ["time window", "window", "time", "unavailability"], "Time Window")
                
                new_entries = []
                details_lines = [f"Effective Date Range: {date_note}"]
                for d in checked_days:
                    t_val = day_inputs[d]["spec_time"]
                    final_win_str = f"{t_val} ({date_note})"
                    new_entries.append({emp_col: emp_name, day_col: d, win_col: final_win_str})
                    details_lines.append(f"  • {d}: {t_val}")
                    
                df_updated = pd.concat([df_curr, pd.DataFrame(new_entries)], ignore_index=True)
                st.session_state.manual_unavailability = sort_dataframe_by_team_and_age(df_updated)
                save_persisted_df(st.session_state.manual_unavailability, "unavailability.csv")
                
                # Send email notification to managers
                details_str = "\n".join(details_lines)
                sent_ok, email_msg = send_availability_notification_email_smtp(emp_name, "Logged New Availability Constraints", details_str)

                st.success(f"✅ Saved unavailability for {len(checked_days)} day(s) ({', '.join(checked_days)}) from {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}!")
                if sent_ok:
                    st.toast("📧 Managers notified via email!", icon="📧")
                else:
                    st.toast(f"⚠️ Unavailability saved locally, but email failed: {email_msg}", icon="⚠️")
                st.rerun()

    st.markdown("---")
    st.markdown("### 📋 My Active Logged Constraints")
    df_curr = st.session_state.manual_unavailability.copy()
    emp_col = find_column(df_curr, ["employee", "name", "staff"], "Employee")
    day_col = find_column(df_curr, ["day", "date", "weekday"], "Day")
    win_col = find_column(df_curr, ["time window", "window", "time", "unavailability"], "Time Window")
    
    mask = df_curr[emp_col].astype(str).str.strip().str.lower() == emp_name.lower()
    user_rows = df_curr[mask].copy()
    
    if user_rows.empty:
        st.info("🎉 You have no logged unavailability constraints. You are listed as Available all week!")
    else:
        for idx, row in user_rows.iterrows():
            day_val = row.get(day_col, "")
            win_val = row.get(win_col, "")
            c1, c2, c3 = st.columns([2, 3.5, 1])
            with c1:
                st.markdown(f"🗓️ **{day_val}**")
            with c2:
                badge_bg = "#9b2c2c" if "all day" in str(win_val).lower() else "#b7791f"
                st.markdown(f'<span style="background-color: {badge_bg}; color: white; padding: 4px 14px; border-radius: 12px; font-weight: 700;">{win_val}</span>', unsafe_allow_html=True)
            with c3:
                if st.button("🗑️ Delete", key=f"btn_delete_unavail_{idx}"):
                    day_del = row.get(day_col, "")
                    win_del = row.get(win_col, "")
                    st.session_state.manual_unavailability = df_curr.drop(idx).reset_index(drop=True)
                    save_persisted_df(st.session_state.manual_unavailability, "unavailability.csv")
                    
                    del_details = f"Removed Constraint:\n  • Day: {day_del}\n  • Details: {win_del}"
                    sent_ok_del, email_msg_del = send_availability_notification_email_smtp(emp_name, "Deleted Availability Constraint", del_details)
                    if sent_ok_del:
                        st.toast("📧 Managers notified of deletion!", icon="📧")
                    else:
                        st.toast(f"⚠️ Constraint deleted locally, but email failed: {email_msg_del}", icon="⚠️")

                    st.success("Constraint deleted.")
                    st.rerun()

# Helper function to render Visual Monthly Calendar Grid with Color-Coded Event Badges for Manager
def render_team_monthly_calendar_grid():
    import calendar
    from datetime import date
    
    st.markdown("""
    <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 14px 22px; border-radius: 14px; color: #ffffff !important; font-weight: 800; font-size: 1.25rem; letter-spacing: 0.3px; border: 2px solid #e5a93c; margin-top: 15px;">
        📅 Bakery Team Monthly Calendar & Unavailability Grid
    </div>
    """, unsafe_allow_html=True)
        
    # Month / Year Selection Controls
    col_nav1, col_nav2, col_nav3 = st.columns([1, 2, 1])
    with col_nav2:
        selected_month_str = st.selectbox(
            "Select Month & Year View", 
            ["August 2026", "September 2026", "October 2026", "November 2026", "December 2026", "January 2027"],
            index=0,
            key="cal_grid_month_picker"
        )
        
    month_names = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    parts = selected_month_str.split()
    sel_month_name = parts[0]
    sel_year = int(parts[1])
    sel_month = month_names.index(sel_month_name) + 1
    
    # Render 7 day headers (Mon to Sun - matching Monday-Sunday bakery roster week format)
    days_hdr = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    cols_hdr = st.columns(7)
    for i, h in enumerate(days_hdr):
        with cols_hdr[i]:
            st.markdown(f'<div style="text-align: center; font-weight: 800; color: #e5a93c; background: #0c2b25; padding: 8px; border-radius: 8px; font-size: 0.95rem;">{h}</div>', unsafe_allow_html=True)
            
    # Set calendar to Monday-first (firstweekday=0)
    cal = calendar.Calendar(firstweekday=0)
    month_days = cal.monthdayscalendar(sel_year, sel_month)
    
    # Always load fresh data from disk so calendar grid & breakdown table are 100% in sync with file
    unavail_df = standardize_unavailability_df(load_persisted_df("unavailability.csv", default_unavail))
    st.session_state.manual_unavailability = unavail_df

    emp_col = find_column(unavail_df, ["employee", "name", "staff", "user", "person", "employee name", "staff name"], "Employee")
    day_col = find_column(unavail_df, ["day", "date", "weekday", "when", "day of week", "unavailable date"], "Day")
    win_col = find_column(unavail_df, ["time window", "window", "time", "unavailability", "reason", "constraint", "time constraint", "notes", "details"], "Time Window")

    # Positional fallback if named column matching is empty
    if unavail_df is not None and hasattr(unavail_df, "columns") and len(unavail_df.columns) > 0:
        if not emp_col or emp_col not in unavail_df.columns:
            emp_col = unavail_df.columns[0]
        if (not day_col or day_col not in unavail_df.columns) and len(unavail_df.columns) >= 2:
            day_col = unavail_df.columns[1]
        if (not win_col or win_col not in unavail_df.columns) and len(unavail_df.columns) >= 3:
            win_col = unavail_df.columns[2]

    # Filter out empty or 'nan' rows
    clean_rows = []
    if emp_col and day_col and win_col and not unavail_df.empty:
        for _, r in unavail_df.iterrows():
            e = str(r.get(emp_col, "")).strip()
            d = str(r.get(day_col, "")).strip()
            w = str(r.get(win_col, "")).strip()
            if e and e.lower() not in ["nan", "none", ""] and d and d.lower() not in ["nan", "none", ""]:
                clean_rows.append({emp_col: e, day_col: d, win_col: w})
                
    clean_unavail_df = sort_dataframe_by_team_and_age(standardize_unavailability_df(pd.DataFrame(clean_rows)))

    # Build name_map for matching employee names
    emp_df = st.session_state.get("manual_employees", None)
    name_map = {}
    if emp_df is not None and hasattr(emp_df, "columns") and len(emp_df.columns) > 0:
        name_c = find_column(emp_df, ["name", "employee", "staff"])
        if name_c and name_c in emp_df.columns:
            for n in emp_df[name_c].dropna():
                n_str = str(n).strip()
                if n_str and n_str.lower() not in ["nan", "none", ""]:
                    name_map[n_str.lower()] = n_str

    for week in month_days:
        week_cols = st.columns(7)
        for i, day_num in enumerate(week):
            with week_cols[i]:
                if day_num == 0:
                    st.markdown('<div style="min-height: 105px; background: rgba(255,255,255,0.02); border-radius: 8px; margin-top: 4px;"></div>', unsafe_allow_html=True)
                else:
                    dt_obj = date(sel_year, sel_month, day_num)
                    
                    chips_html = []
                    
                    # Unavailability Chips (Employee Names who are not available)
                    for _, urow in clean_unavail_df.iterrows():
                        u_emp = str(urow.get(emp_col, "")).strip()
                        u_day = str(urow.get(day_col, "")).strip()
                        u_win = str(urow.get(win_col, "")).strip()
                        
                        if u_emp and u_day and is_unavail_applicable_to_date(dt_obj, u_day, u_win):
                            matched = find_matching_employee(u_emp, name_map) if name_map else u_emp
                            display_name = matched if matched else u_emp
                            
                            badge_win = clean_win_display(u_win)
                            tooltip_win = u_win if (u_win and str(u_win).strip().lower() not in ["nan", "none", "nat", "null", ""]) else "All Day"
                            
                            if "all day" in tooltip_win.lower():
                                chips_html.append(f'<div style="background-color: #e53e3e; color: white; padding: 3px 8px; border-radius: 6px; font-size: 0.75rem; font-weight: 700; margin-bottom: 3px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;" title="{display_name}: {tooltip_win}">🔴 {display_name} ({badge_win})</div>')
                            else:
                                chips_html.append(f'<div style="background-color: #dd6b20; color: white; padding: 3px 8px; border-radius: 6px; font-size: 0.75rem; font-weight: 700; margin-bottom: 3px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;" title="{display_name}: {tooltip_win}">🟨 {display_name} ({badge_win})</div>')
                                    
                    chips_block = "".join(chips_html) if chips_html else '<div style="color: #718096; font-size: 0.75rem; font-weight: 600; padding: 4px 0;">Clear (All Available)</div>'
                    
                    st.markdown(f"""
                    <div style="min-height: 120px; max-height: 150px; background: #11362f; border: 1px solid #1f5c50; border-radius: 8px; padding: 6px; margin-top: 4px; display: flex; flex-direction: column;">
                        <div style="font-weight: 800; font-size: 0.85rem; color: #e5a93c; border-bottom: 1px solid #1f5c50; margin-bottom: 4px; padding-bottom: 2px; flex-shrink: 0;">{day_num}</div>
                        <div style="overflow-y: auto; flex-grow: 1; padding-right: 2px;">
                            {chips_block}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

    # --- Unavailability Breakdown by Employee Table according to chosen month ---
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 12px 20px; border-radius: 12px 12px 0 0; color: #ffffff !important; font-weight: 800; font-size: 1.15rem; letter-spacing: 0.3px; border: 2px solid #e5a93c; border-bottom: none; margin-top: 15px;">
        🚫 Unavailability Breakdown by Employee ({selected_month_str})
    </div>
    """, unsafe_allow_html=True)

    month_days_list = [d for w in month_days for d in w if d > 0]
    active_rows = []
    if not clean_unavail_df.empty:
        for _, urow in clean_unavail_df.iterrows():
            u_emp = str(urow.get(emp_col, "")).strip()
            u_day = str(urow.get(day_col, "")).strip()
            u_win = str(urow.get(win_col, "")).strip()
            
            applies_to_month = False
            for day_num in month_days_list:
                dt_obj = date(sel_year, sel_month, day_num)
                if is_unavail_applicable_to_date(dt_obj, u_day, u_win):
                    applies_to_month = True
                    break
                    
            if applies_to_month:
                matched_name = find_matching_employee(u_emp, name_map) if name_map else u_emp
                active_rows.append({
                    "Employee": matched_name if matched_name else u_emp,
                    "Day": u_day,
                    "Time Window": u_win
                })

    if active_rows:
        month_unavail_df = pd.DataFrame(active_rows).drop_duplicates()
        month_unavail_df = sort_dataframe_by_team_and_age(standardize_unavailability_df(month_unavail_df))
    else:
        month_unavail_df = pd.DataFrame(columns=["Employee", "Day", "Time Window"])

    current_unavail_sig = str(len(month_unavail_df)) + "_" + str(hash(tuple(month_unavail_df.astype(str).values.flatten())))
    if st.session_state.get(f"unavail_sig_{sel_month}_{sel_year}") != current_unavail_sig:
        if f"edit_unavail_month_{sel_month}_{sel_year}" in st.session_state:
            del st.session_state[f"edit_unavail_month_{sel_month}_{sel_year}"]
        st.session_state[f"unavail_sig_{sel_month}_{sel_year}"] = current_unavail_sig

    edited_month_df = st.data_editor(month_unavail_df, num_rows="dynamic", key=f"edit_unavail_month_{sel_month}_{sel_year}")
    if edited_month_df is not None:
        std_edited = standardize_unavailability_df(edited_month_df)
        if not std_edited.equals(month_unavail_df):
            # Merge edits safely back into master dataset without dropping rows for other months
            full_df = load_persisted_df("unavailability.csv", default_unavail)
            full_df = standardize_unavailability_df(full_df)
            
            keep_rows = []
            if full_df is not None and not full_df.empty:
                for _, r in full_df.iterrows():
                    e = str(r.get("Employee", "")).strip()
                    d = str(r.get("Day", "")).strip()
                    w = str(r.get("Time Window", "")).strip()
                    applies = False
                    for day_num in month_days_list:
                        dt_obj = date(sel_year, sel_month, day_num)
                        if is_unavail_applicable_to_date(dt_obj, d, w):
                            applies = True
                            break
                    if not applies and e:
                        keep_rows.append({"Employee": e, "Day": d, "Time Window": w})
            
            combined_df = pd.concat([pd.DataFrame(keep_rows), std_edited], ignore_index=True).drop_duplicates()
            combined_df = sort_dataframe_by_team_and_age(standardize_unavailability_df(combined_df))
            st.session_state.manual_unavailability = combined_df
            save_persisted_df(combined_df, "unavailability.csv")
            st.session_state[f"unavail_sig_{sel_month}_{sel_year}"] = str(len(std_edited)) + "_" + str(hash(tuple(std_edited.astype(str).values.flatten())))
            st.rerun()

def build_weekly_timesheet_excel_bytes(selected_week_str):
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Weekly Audited Summary"
    ws_details = wb.create_sheet(title="Daily Punch Details")
    
    # Parse week start and end dates
    mon_date_obj = parse_date_robust(selected_week_str)
    if not mon_date_obj:
        mon_date_obj = datetime.now().date()
    if isinstance(mon_date_obj, datetime):
        mon_date_obj = mon_date_obj.date()
    sun_date_obj = mon_date_obj + timedelta(days=6)
    
    week_period_str = f"Roster Week Period: Monday {mon_date_obj.strftime('%d/%m/%Y')} – Sunday {sun_date_obj.strftime('%d/%m/%Y')}"
    
    # Fills & Fonts
    title_fill = PatternFill(start_color="081D19", end_color="081D19", fill_type="solid")
    subtitle_fill = PatternFill(start_color="0C2B25", end_color="0C2B25", fill_type="solid")
    header_fill = PatternFill(start_color="11362F", end_color="11362F", fill_type="solid")
    day_header_fill = PatternFill(start_color="16443C", end_color="16443C", fill_type="solid")
    
    title_font = Font(name="Arial", size=14, bold=True, color="E5A93C")
    subtitle_font = Font(name="Arial", size=11, bold=True, italic=True, color="FFFFFF")
    header_font = Font(name="Arial", size=11, bold=True, color="E5A93C")
    day_header_font = Font(name="Arial", size=11, bold=True, color="E5A93C")
    data_font = Font(name="Arial", size=10)
    
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    border_side = Side(style='thin', color='1F5C50')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # ----------------------------------------------------
    # 1. SETUP SHEET TITLES & PERIOD HEADERS
    # ----------------------------------------------------
    # Summary Sheet Title
    ws_summary.merge_cells("A1:F1")
    cell_s1 = ws_summary["A1"]
    cell_s1.value = "BRUMBY'S BAKERY PAKENHAM — AUDITED WEEKLY TIMESHEET SUMMARY"
    cell_s1.fill = title_fill
    cell_s1.font = title_font
    cell_s1.alignment = center_align
    
    ws_summary.merge_cells("A2:F2")
    cell_s2 = ws_summary["A2"]
    cell_s2.value = week_period_str
    cell_s2.fill = subtitle_fill
    cell_s2.font = subtitle_font
    cell_s2.alignment = center_align
    
    # Details Sheet Title
    ws_details.merge_cells("A1:M1")
    cell_d1 = ws_details["A1"]
    cell_d1.value = "BRUMBY'S BAKERY PAKENHAM — DAILY SHIFT PUNCH DETAILS"
    cell_d1.fill = title_fill
    cell_d1.font = title_font
    cell_d1.alignment = center_align
    
    ws_details.merge_cells("A2:M2")
    cell_d2 = ws_details["A2"]
    cell_d2.value = week_period_str
    cell_d2.fill = subtitle_fill
    cell_d2.font = subtitle_font
    cell_d2.alignment = center_align
    
    # Row 3 blank separator
    ws_summary.row_dimensions[3].height = 10
    ws_details.row_dimensions[3].height = 10
    
    # ----------------------------------------------------
    # 2. COLUMN HEADERS (ROW 4)
    # ----------------------------------------------------
    headers_summary = ["Employee", "Scheduled Shift Hours", "Actual Audited Hours", "Variance (Mins)", "Late Shift Status", "Manager Approval Status"]
    ws_summary.append([]) # Row 3 blank
    ws_summary.append(headers_summary) # Row 4
    for col_num, h_text in enumerate(headers_summary, 1):
        cell = ws_summary.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = cell_border
        
    headers_details = ["Record ID", "Date", "Day", "Employee", "Scheduled Shift", "Clock In", "Clock Out", "Net Hours", "Variance (Mins)", "GPS Distance (m)", "Location Verification", "Note / Alert", "Status"]
    ws_details.append([]) # Row 3 blank
    ws_details.append(headers_details) # Row 4
    for col_num, h_text in enumerate(headers_details, 1):
        cell = ws_details.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = cell_border

    # ----------------------------------------------------
    # 3. LOAD & GROUP TIMECARDS BY DAY OF WEEK
    # ----------------------------------------------------
    df_cards = load_persisted_timecards()
    week_cards = []
    if df_cards is not None and not df_cards.empty and "Date" in df_cards.columns:
        for idx, row in df_cards.iterrows():
            d_obj = parse_date_robust(row.get("Date", ""))
            if d_obj and get_week_start_date_str(d_obj) == selected_week_str:
                row_dict = row.to_dict()
                row_dict["_d_obj"] = d_obj
                week_cards.append(row_dict)

    # Sort chronologically by date
    week_cards.sort(key=lambda x: x["_d_obj"])

    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    cards_by_day = {day: [] for day in days_order}
    emp_summary_map = {}

    for r_dict in week_cards:
        d_obj = r_dict["_d_obj"]
        day_w_name = days_order[d_obj.weekday()]
        cards_by_day[day_w_name].append(r_dict)
        
        emp = r_dict.get("Employee", "")
        if emp not in emp_summary_map:
            emp_summary_map[emp] = {"sched_hrs": 0.0, "actual_hrs": 0.0, "var_mins": 0, "late_status": "Normal", "approval": "Approved"}
            
        try:
            actual_h = float(r_dict.get("Net Hours", "0"))
        except:
            actual_h = 0.0
        try:
            var_m = int(float(r_dict.get("Variance (Mins)", "0")))
        except:
            var_m = 0
            
        emp_summary_map[emp]["actual_hrs"] += actual_h
        emp_summary_map[emp]["var_mins"] += var_m
        if "Late" in r_dict.get("Late Correction Status", "") or "Late" in r_dict.get("Note", ""):
            emp_summary_map[emp]["late_status"] = r_dict.get("Note", "Late Shift")
        if "Missing" in r_dict.get("Note", ""):
            emp_summary_map[emp]["approval"] = "Action Required"

    # ----------------------------------------------------
    # 4. WRITE DAY OF WEEK SEPARATED ROWS (DETAILS SHEET)
    # ----------------------------------------------------
    current_row = 5
    for day_idx, day_name in enumerate(days_order):
        day_date_obj = mon_date_obj + timedelta(days=day_idx)
        day_date_str = day_date_obj.strftime("%d/%m/%Y")
        day_punches = cards_by_day[day_name]
        
        if day_punches:
            # Day Section Header Divider Row
            ws_details.append([f"📅 {day_name.upper()} — {day_date_str}"])
            ws_details.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
            
            for c_i in range(1, 14):
                cell = ws_details.cell(row=current_row, column=c_i)
                cell.fill = day_header_fill
                cell.font = day_header_font
                cell.alignment = center_align
                cell.border = cell_border
            current_row += 1
            
            # Punch rows for this day
            for r_dict in day_punches:
                row_vals = [
                    r_dict.get("Record ID", ""),
                    r_dict.get("Date", ""),
                    day_name,
                    r_dict.get("Employee", ""),
                    r_dict.get("Scheduled Shift", ""),
                    r_dict.get("Clock In", ""),
                    r_dict.get("Clock Out", ""),
                    r_dict.get("Net Hours", "0"),
                    r_dict.get("Variance (Mins)", "0"),
                    r_dict.get("Distance (m)", "0"),
                    r_dict.get("Location Verification", ""),
                    r_dict.get("Note", r_dict.get("Missing Punch Alert", "")),
                    r_dict.get("Status", "")
                ]
                ws_details.append(row_vals)
                for c_i in range(1, 14):
                    cell = ws_details.cell(row=current_row, column=c_i)
                    cell.font = data_font
                    cell.border = cell_border
                    if c_i in [8, 9, 10]:
                        cell.alignment = right_align
                    elif c_i in [2, 3, 6, 7]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                current_row += 1

    # ----------------------------------------------------
    # 5. WRITE SUMMARY SHEET ROWS
    # ----------------------------------------------------
    tot_hrs = 0.0
    for emp, s_data in emp_summary_map.items():
        act_h = round(s_data["actual_hrs"], 2)
        tot_hrs += act_h
        ws_summary.append([
            emp,
            act_h,
            act_h,
            s_data["var_mins"],
            s_data["late_status"],
            s_data["approval"]
        ])
        for c_i in range(1, 7):
            cell = ws_summary.cell(row=ws_summary.max_row, column=c_i)
            cell.font = data_font
            cell.border = cell_border
            if c_i in [2, 3, 4]:
                cell.alignment = right_align
            else:
                cell.alignment = left_align
        
    # Total Summary Row
    ws_summary.append([
        "TOTAL WEEKLY AUDITED HOURS",
        round(tot_hrs, 2),
        round(tot_hrs, 2),
        "",
        "",
        ""
    ])
    tot_row_idx = ws_summary.max_row
    for c_i in range(1, 7):
        cell = ws_summary.cell(row=tot_row_idx, column=c_i)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
        if c_i in [2, 3]:
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    # ----------------------------------------------------
    # 6. AUTO-FIT COLUMN WIDTHS & STYLING
    # ----------------------------------------------------
    for ws in [ws_summary, ws_details]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                # Skip merged title cells for width calculation
                if cell.coordinate in ws.merged_cells:
                    continue
                max_len = max(max_len, len(val_str))
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def render_manager_timesheet_audit_dashboard():
    st.markdown("""
    <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 16px 22px; border-radius: 12px; color: #ffffff !important; border: 2px solid #e5a93c; margin-bottom: 20px;">
        <h3 style="margin: 0; color: #e5a93c;">⏱️ Shift Timesheet Audit & Live Attendance</h3>
        <p style="margin: 4px 0 0 0; color: #d0e6df; font-size: 0.95rem;">Review live working staff, resolve late/missing clockings, approve or reject roster adjustments, and download weekly timesheets.</p>
    </div>
    """, unsafe_allow_html=True)
    
    df_cards = load_persisted_timecards()
    
    # Clean auto-generated "Missing" records for past historical dates (older than today)
    if df_cards is not None and not df_cards.empty:
        today_dt = datetime.now().date()
        valid_rows = []
        for idx, r in df_cards.iterrows():
            status = str(r.get("Status", "")).strip()
            c_in = str(r.get("Clock In", "")).strip()
            d_str = str(r.get("Date", "")).strip()
            d_obj = parse_date_robust(d_str)
            # Retain actual punches (with Clock In), or today's missing punches. Drop historical unpunched missing records.
            if c_in or status != "Missing" or (d_obj and d_obj == today_dt):
                valid_rows.append(r)
        
        df_cards = pd.DataFrame(valid_rows).reset_index(drop=True) if valid_rows else pd.DataFrame()
        save_timecard_records(df_cards)

    # 1. Scan current active week roster ONLY for today's missing clockings (not past historical dates)
    today_dt = datetime.now().date()
    past_rosters = list_finalized_rosters()
    if past_rosters:
        for r_item in past_rosters:
            r_df = load_finalized_roster(r_item["csv_filename"])
            start_dt = r_item.get("start_date")
            if r_df is not None and not r_df.empty and start_dt:
                if start_dt <= today_dt <= start_dt + timedelta(days=6):
                    emp_col = find_column(r_df, ["name", "employee", "staff"])
                    if emp_col in r_df.columns:
                        days_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                        d_idx = today_dt.weekday()
                        day_name = days_list[d_idx]
                        shift_date_str = today_dt.strftime("%d/%m/%Y")
                        
                        if day_name in r_df.columns:
                            for _, r_row in r_df.iterrows():
                                emp_name = str(r_row.get(emp_col, "")).strip()
                                shift_val = str(r_row.get(day_name, "")).strip()
                                
                                is_owner = emp_name.lower() in ["viet", "jane"]
                                for u_k, u_v in user_profiles.items():
                                    if u_v.get("employee_name", "").strip().lower() == emp_name.lower():
                                        if u_v.get("profile", {}).get("classification", "").lower() == "owner":
                                            is_owner = True

                                if not is_owner and emp_name and shift_val and shift_val.lower() not in ["off", "nan", "unavailable"]:
                                    rec_id = f"TC_{shift_date_str.replace('/', '')}_{emp_name.replace(' ', '')}"
                                    
                                    card_exists = False
                                    if df_cards is not None and not df_cards.empty and "Record ID" in df_cards.columns:
                                        if rec_id in df_cards["Record ID"].values:
                                            card_exists = True
                                            
                                    if not card_exists:
                                        missing_rec = {
                                            "Record ID": rec_id,
                                            "Date": shift_date_str,
                                            "Employee": emp_name,
                                            "Scheduled Shift": shift_val,
                                            "Clock In": "",
                                            "Clock Out": "",
                                            "Net Hours": "0",
                                            "Variance (Mins)": "0",
                                            "GPS Lat": str(BAKERY_LAT),
                                            "GPS Lon": str(BAKERY_LON),
                                            "Distance (m)": "0.0",
                                            "Location Verification": "⚠️ Missing Clocking",
                                            "Note": "⚠️ Missing Clock-In",
                                            "Late Correction Status": "Missing Punch",
                                            "Status": "Missing"
                                        }
                                        if df_cards is None or df_cards.empty:
                                            df_cards = pd.DataFrame([missing_rec])
                                        else:
                                            df_cards = pd.concat([df_cards, pd.DataFrame([missing_rec])], ignore_index=True)
                                        save_timecard_records(df_cards)

    # Re-calculate Note column for all rows
    if df_cards is not None and not df_cards.empty:
        notes = []
        for idx, r in df_cards.iterrows():
            existing_note = str(r.get("Note", "")).strip()
            c_in = str(r.get("Clock In", "")).strip()
            c_out = str(r.get("Clock Out", "")).strip()
            sched = str(r.get("Scheduled Shift", "")).strip()
            status = str(r.get("Status", "")).strip()
            
            if "Approved" in existing_note or "Rejected" in existing_note:
                notes.append(existing_note)
            elif not c_in and status == "Missing":
                notes.append("⚠️ Missing Clock-In")
            elif c_in and not c_out and status == "Completed":
                notes.append("⚠️ Missing Clock-Out")
            elif c_in and sched and "-" in sched:
                sched_start_str = sched.split("-")[0].strip()
                c_in_dec = parse_time_to_decimal(c_in)
                s_in_dec = parse_time_to_decimal(sched_start_str)
                var_mins = round((c_in_dec - s_in_dec) * 60)
                if var_mins >= 10:
                    notes.append(f"⚠️ Late Clocking (+{var_mins} mins)")
                else:
                    notes.append("✅ Verified / Normal")
            else:
                notes.append(existing_note if existing_note else "✅ Verified / Normal")
        df_cards["Note"] = notes

    # Metrics Bar
    working_count = 0
    late_count = 0
    missing_count = 0
    if df_cards is not None and not df_cards.empty:
        if "Status" in df_cards.columns:
            working_count = len(df_cards[df_cards["Status"] == "Working"])
        if "Note" in df_cards.columns:
            late_count = len(df_cards[df_cards["Note"].str.contains("Late Clocking", na=False)])
            missing_count = len(df_cards[df_cards["Note"].str.contains("Missing", na=False)])

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.markdown(f"""
        <div style="background: #0c2b25; padding: 12px; border-radius: 10px; border: 1px solid #1f5c50; text-align: center;">
            <div style="font-size: 0.85rem; color: #a0aec0; font-weight: 700;">ON-DUTY NOW</div>
            <div style="font-size: 1.6rem; font-weight: 900; color: #48bb78;">🟢 {working_count} Staff</div>
        </div>
        """, unsafe_allow_html=True)
    with m2:
        st.markdown(f"""
        <div style="background: #0c2b25; padding: 12px; border-radius: 10px; border: 1px solid #1f5c50; text-align: center;">
            <div style="font-size: 0.85rem; color: #a0aec0; font-weight: 700;">LATE CLOCKINGS</div>
            <div style="font-size: 1.6rem; font-weight: 900; color: #ecc94b;">⚠️ {late_count} Pending</div>
        </div>
        """, unsafe_allow_html=True)
    with m3:
        st.markdown(f"""
        <div style="background: #0c2b25; padding: 12px; border-radius: 10px; border: 1px solid #1f5c50; text-align: center;">
            <div style="font-size: 0.85rem; color: #a0aec0; font-weight: 700;">MISSING CLOCKINGS</div>
            <div style="font-size: 1.6rem; font-weight: 900; color: #f6ad55;">⚠️ {missing_count} Pending</div>
        </div>
        """, unsafe_allow_html=True)
    with m4:
        st.markdown(f"""
        <div style="background: #0c2b25; padding: 12px; border-radius: 10px; border: 1px solid #1f5c50; text-align: center;">
            <div style="font-size: 0.85rem; color: #a0aec0; font-weight: 700;">TOTAL TIMECARDS</div>
            <div style="font-size: 1.6rem; font-weight: 900; color: #4299e1;">📋 {len(df_cards) if df_cards is not None else 0} Records</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Master Table Header
    st.markdown("""
    <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 10px 18px; border-radius: 12px 12px 0 0; color: #ffffff !important; font-weight: 800; font-size: 1.1rem; letter-spacing: 0.3px; border: 2px solid #e5a93c; border-bottom: none;">
        📋 Master Timesheet Audit & Notification Table (Check First Cell `Select` Row to Approve/Reject)
    </div>
    """, unsafe_allow_html=True)

    if df_cards is not None and not df_cards.empty:
        display_df = df_cards.copy()
        if "Select" not in display_df.columns:
            display_df.insert(0, "Select", False)
        else:
            display_df["Select"] = False
            
        cols_order = ["Select", "Note", "Date", "Employee", "Scheduled Shift", "Clock In", "Clock Out", "Net Hours", "Distance (m)", "Status", "Record ID"]
        existing_cols = [c for c in cols_order if c in display_df.columns]
        display_df = display_df[existing_cols]

        edited_df = st.data_editor(
            display_df,
            num_rows="dynamic",
            key="edit_timecards_master_table",
            column_config={
                "Select": st.column_config.CheckboxColumn("Select", help="Check row to trigger Approve / Reject actions", default=False),
                "Note": st.column_config.TextColumn("Note", help="Notification status and system alerts", disabled=True),
            },
            use_container_width=True
        )

        selected_rows = edited_df[edited_df["Select"] == True]
        
        if not selected_rows.empty:
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("""
            <div style="background: rgba(8, 29, 25, 0.95); border: 2px solid #e5a93c; border-radius: 12px; padding: 14px 20px;">
                <h4 style="margin: 0 0 8px 0; color: #e5a93c;">⚡ Manager Action Controls for Selected Row(s)</h4>
            </div>
            """, unsafe_allow_html=True)
            
            for idx, sel_row in selected_rows.iterrows():
                rec_id = sel_row.get("Record ID", "")
                sel_note = sel_row.get("Note", "")
                sel_emp = sel_row.get("Employee", "")
                sel_date = sel_row.get("Date", "")
                sel_clock_in = sel_row.get("Clock In", "")
                sel_sched = sel_row.get("Scheduled Shift", "")
                
                st.markdown(f"**Selected Record:** `{sel_emp}` on `{sel_date}` | Notification: `{sel_note}` | Scheduled: `{sel_sched}` | Clock In: `{sel_clock_in}`")
                
                act_col1, act_col2 = st.columns(2)
                
                if "Late Clocking" in sel_note:
                    with act_col1:
                        if st.button(f"✅ Approve Late Shift (Adjust Roster to {sel_clock_in})", key=f"btn_app_late_{rec_id}_{idx}", use_container_width=True):
                            l_actual_in = sel_clock_in
                            past_rosters = list_finalized_rosters()
                            if past_rosters:
                                for r_item in past_rosters:
                                    r_df = load_finalized_roster(r_item["csv_filename"])
                                    if r_df is not None and not r_df.empty:
                                        l_date_obj = parse_date_robust(sel_date)
                                        if l_date_obj:
                                            day_w_name = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"][l_date_obj.weekday()]
                                            emp_col = find_column(r_df, ["name", "employee", "staff"])
                                            if emp_col in r_df.columns and day_w_name in r_df.columns:
                                                for r_idx, r_row in r_df.iterrows():
                                                    if find_matching_employee(sel_emp, {str(r_row.get(emp_col, "")).strip().lower(): str(r_row.get(emp_col, "")).strip()}):
                                                        old_shift = str(r_row.get(day_w_name, "")).strip()
                                                        if "-" in old_shift:
                                                            old_end = old_shift.split("-")[1].strip()
                                                            new_shift = f"{l_actual_in}-{old_end}"
                                                            r_df.at[r_idx, day_w_name] = new_shift
                                                            save_persisted_df(r_df, os.path.join("finalized_rosters", r_item["csv_filename"]))

                            df_cards.loc[df_cards["Record ID"] == rec_id, "Note"] = "✅ Approved (Roster Adjusted)"
                            df_cards.loc[df_cards["Record ID"] == rec_id, "Late Correction Status"] = "Late Shift Corrected"
                            save_timecard_records(df_cards)
                            st.success(f"✅ Approved late clocking for **{sel_emp}**. Roster adjusted to `{l_actual_in}`!")
                            st.rerun()

                    with act_col2:
                        if st.button(f"❌ Reject Late Shift (Keep Original Roster {sel_sched})", key=f"btn_rej_late_{rec_id}_{idx}", use_container_width=True):
                            df_cards.loc[df_cards["Record ID"] == rec_id, "Note"] = "❌ Rejected (Roster Maintained)"
                            df_cards.loc[df_cards["Record ID"] == rec_id, "Late Correction Status"] = "Rejected (Unexcused Late)"
                            save_timecard_records(df_cards)
                            st.info(f"ℹ️ Rejected late shift adjustment for **{sel_emp}**. Original roster `{sel_sched}` maintained.")
                            st.rerun()

                elif "Missing" in sel_note:
                    with act_col1:
                        if st.button(f"✅ Approve Missing Shift (Keep Roster {sel_sched})", key=f"btn_app_miss_{rec_id}_{idx}", use_container_width=True):
                            df_cards.loc[df_cards["Record ID"] == rec_id, "Note"] = "✅ Approved (Roster Maintained)"
                            df_cards.loc[df_cards["Record ID"] == rec_id, "Status"] = "Completed"
                            save_timecard_records(df_cards)
                            st.success(f"✅ Approved missing shift for **{sel_emp}**. Original roster shift `{sel_sched}` approved and maintained.")
                            st.rerun()

                    with act_col2:
                        if st.button(f"❌ Reject Missing Shift (Delete Shift from Roster)", key=f"btn_rej_miss_{rec_id}_{idx}", use_container_width=True):
                            past_rosters = list_finalized_rosters()
                            if past_rosters:
                                for r_item in past_rosters:
                                    r_df = load_finalized_roster(r_item["csv_filename"])
                                    if r_df is not None and not r_df.empty:
                                        l_date_obj = parse_date_robust(sel_date)
                                        if l_date_obj:
                                            day_w_name = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"][l_date_obj.weekday()]
                                            emp_col = find_column(r_df, ["name", "employee", "staff"])
                                            if emp_col in r_df.columns and day_w_name in r_df.columns:
                                                for r_idx, r_row in r_df.iterrows():
                                                    if find_matching_employee(sel_emp, {str(r_row.get(emp_col, "")).strip().lower(): str(r_row.get(emp_col, "")).strip()}):
                                                        r_df.at[r_idx, day_w_name] = "OFF"
                                                        save_persisted_df(r_df, os.path.join("finalized_rosters", r_item["csv_filename"]))

                            df_cards.loc[df_cards["Record ID"] == rec_id, "Note"] = "❌ Rejected (Shift Deleted)"
                            df_cards.loc[df_cards["Record ID"] == rec_id, "Status"] = "Rejected"
                            save_timecard_records(df_cards)
                            st.warning(f"⚠️ Rejected missing shift for **{sel_emp}**. Shift has been deleted (`OFF`) from the roster.")
                            st.rerun()
                else:
                    st.info("ℹ️ Selected row is verified and normal. No pending actions required.")

    else:
        st.info("ℹ️ No timecard records or scheduled roster shifts found for audit.")

    st.markdown("<br>", unsafe_allow_html=True)
    st.subheader("📥 Download Weekly Timesheet Workbook")
    
    available_weeks = []
    if os.path.exists(TIMESHEETS_DIR):
        for f in os.listdir(TIMESHEETS_DIR):
            if f.startswith("timesheet_week_") and f.endswith(".csv"):
                w_str = f.replace("timesheet_week_", "").replace(".csv", "")
                available_weeks.append(w_str)
                
    if not available_weeks:
        curr_w = get_week_start_date_str()
        available_weeks = [curr_w]

    sel_download_week = st.selectbox("Select Roster Week to Download:", available_weeks, key="sel_dl_week")
    
    excel_bytes = build_weekly_timesheet_excel_bytes(sel_download_week)
    st.download_button(
        label=f"📥 Download Timesheet_Audit_Week_{sel_download_week}.xlsx",
        data=excel_bytes,
        file_name=f"Timesheet_Audit_Week_{sel_download_week}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="btn_download_timesheet_excel"
    )

# IF EMPLOYEE, RENDER 3 TABS (CURRENT ROSTER 1ST, PERSONAL INFO 2ND, AVAILABILITY CALENDAR 3RD)
if not is_manager:
    with tab_my_current_roster:
        render_employee_current_roster_tab(st.session_state.logged_in_user)
    with tab_my_info:
        render_confidential_profile_form(st.session_state.logged_in_user)
    with tab_my_avail:
        render_employee_availability_manager(st.session_state.logged_in_user)


# Deterministic solver
def solve_roster(employees_raw, unavailability_raw, requirements_raw, fixed_raw, start_dt, debug_logs=None):
    # Standardize column headers to lowercase and stripped
    employees = employees_raw.copy()
    employees.columns = [str(c).strip().lower() for c in employees.columns]
    
    unavailability = unavailability_raw.copy()
    unavailability.columns = [str(c).strip().lower() for c in unavailability.columns]
    
    requirements = requirements_raw.copy()
    requirements.columns = [str(c).strip().lower() for c in requirements.columns]
    
    fixed = fixed_raw.copy()
    fixed.columns = [str(c).strip().lower() for c in fixed.columns]

    # Map employees columns
    name_col = find_column(employees, ["name", "employee", "employee name", "staff name", "staff"])
    start_col = find_column(employees, ["start date", "commencement date", "started", "startdate", "commence date", "commence"])
    age_col = find_column(employees, ["age", "years"])
    dob_col = find_column(employees, ["dob", "date of birth", "birth date"])
    type_col = find_column(employees, ["employment type", "type", "status", "ft/pt/casual", "employmenttype"])

    # Build active employees list
    active_employees = []
    for _, row in employees.iterrows():
        raw_name = str(row.get(name_col, "")).strip() if name_col else ""
        if not raw_name or raw_name.lower() in ["nan", ""] or "demo" in raw_name.lower():
            continue
            
        is_active = True
        if start_col:
            try:
                start_val = row.get(start_col)
                if pd.notna(start_val):
                    start_date_parsed = pd.to_datetime(start_val)
                    if start_date_parsed.date() > start_dt:
                        is_active = False
            except:
                pass
                
        if not is_active:
            continue
            
        age = 25
        if age_col and pd.notna(row.get(age_col)):
            try:
                age = int(row.get(age_col))
            except:
                pass
        elif dob_col and pd.notna(row.get(dob_col)):
            try:
                dob_parsed = pd.to_datetime(row.get(dob_col))
                today = datetime.now()
                age = today.year - dob_parsed.year - ((today.month, today.day) < (dob_parsed.month, dob_parsed.day))
            except:
                pass
                
        emp_type = str(row.get(type_col, "Casual")).strip() if type_col else "Casual"
        
        active_employees.append({
            "Name": raw_name,
            "NormalizedName": raw_name.lower(),
            "Age": age,
            "Type": emp_type
        })

    if not active_employees:
        return pd.DataFrame()

    # Create mapping from normalized name to original name
    name_map = {emp["NormalizedName"]: emp["Name"] for emp in active_employees}
    
    # Initialize Roster
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    roster_output = {emp["Name"]: {day: "" for day in days_of_week} for emp in active_employees}
    weekly_shifts_count = {emp["Name"]: 0 for emp in active_employees}
    elizabeth_weekday_shifts = 0

    # 1. Apply fixed shifts first
    fixed_name_col = find_column(fixed, ["employee", "name", "employee name", "staff name", "staff"])
    if not fixed_name_col and not fixed.empty:
        fixed_name_col = fixed.columns[0]
        
    if fixed_name_col:
        for _, fix_row in fixed.iterrows():
            raw_fixed_name = str(fix_row.get(fixed_name_col, "")).strip().lower()
            name = find_matching_employee(raw_fixed_name, name_map)
            if name:
                for day in days_of_week:
                    day_col = find_column(fixed, [day.lower(), day.lower()[:3]])
                    if day_col:
                        val = str(fix_row.get(day_col, "")).strip()
                        if val.lower() not in ["off", "nan", "none", ""]:
                            roster_output[name][day] = val
                            weekly_shifts_count[name] += 1
                            if name == "Elizabeth" and day not in ["Saturday", "Sunday"]:
                                elizabeth_weekday_shifts += 1

    # 2. Check unavailability
    unavail_name_col = find_column(unavailability, ["employee", "name", "employee name", "staff name", "staff", "user", "person"])
    unavail_day_col = find_column(unavailability, ["day", "date", "weekday", "when", "day of week", "unavailable date"])
    unavail_window_col = find_column(unavailability, ["time window", "window", "time", "unavailability", "reason", "constraint", "time constraint", "notes", "details"])

    unavail_map = {}
    if unavail_name_col and unavail_day_col and unavail_window_col:
        for _, un_row in unavailability.iterrows():
            raw_unavail_name = str(un_row.get(unavail_name_col, "")).strip()
            unavail_day = str(un_row.get(unavail_day_col, "")).strip()
            window = str(un_row.get(unavail_window_col, "All Day")).strip()
            matched_name = find_matching_employee(raw_unavail_name, name_map)
            if matched_name and unavail_day:
                for idx, day in enumerate(days_of_week):
                    day_date = start_dt + timedelta(days=idx)
                    if is_unavail_applicable_to_date(day_date, unavail_day, window):
                        key = (matched_name.lower(), day.lower())
                        if key not in unavail_map:
                            unavail_map[key] = []
                        unavail_map[key].append(window)

    # Mark day-off unavailability
    for (norm_name, day_lower), windows in unavail_map.items():
        name = name_map.get(norm_name)
        if name:
            for window in windows:
                if "all day" in window.lower() or "anytime" in window.lower():
                    for day in days_of_week:
                        if day.lower() == day_lower:
                            if roster_output[name][day] in ["", "off"]:
                                roster_output[name][day] = " unavailable"

    # 3. Schedule required shifts day by day
    req_day_col = find_column(requirements, ["day", "date", "weekday"])
    req_shift_col = find_column(requirements, ["shift", "time", "hours", "shift time"])
    has_day_cols = any(find_column(requirements, [day.lower(), day.lower()[:3]]) for day in days_of_week)

    for day in days_of_week:
        shifts_to_fill = []
        
        if has_day_cols and req_shift_col:
            day_col = find_column(requirements, [day.lower(), day.lower()[:3]])
            if day_col:
                for _, req in requirements.iterrows():
                    shift = str(req.get(req_shift_col, "")).strip()
                    count_val = req.get(day_col)
                    if pd.notna(count_val):
                        try:
                            count = int(float(count_val))
                        except:
                            count = 0
                        if count > 0:
                            parsed = parse_shift_range(shift)
                            if parsed:
                                start, end, duration = parsed
                                for _ in range(count):
                                    shifts_to_fill.append({"shift": shift, "start": start, "end": end, "duration": duration})
        elif req_day_col and req_shift_col:
            req_count_col = find_column(requirements, ["count required", "count", "required", "staff needed", "personnel", "quantity", "countrequired"])
            clean_day_series = requirements[req_day_col].astype(str).str.strip().str.lower()
            day_reqs = requirements[clean_day_series.str.startswith(day.lower()[:3])]
            
            for _, req in day_reqs.iterrows():
                shift = str(req.get(req_shift_col, "")).strip()
                count_val = req.get(req_count_col, 1)
                try:
                    count = int(count_val) if pd.notna(count_val) else 1
                except:
                    count = 1
                parsed = parse_shift_range(shift)
                if parsed:
                    start, end, duration = parsed
                    for _ in range(count):
                        shifts_to_fill.append({"shift": shift, "start": start, "end": end, "duration": duration})
                        
        fixed_shifts_today = []
        for name in roster_output:
            val = roster_output[name][day]
            if val not in ["", "off", " unavailable", "none", "nan"]:
                fixed_shifts_today.append(val)
                
        remaining_shifts_to_fill = []
        for shift_req in shifts_to_fill:
            filled_idx = -1
            req_clean = shift_req["shift"].strip().lower().replace(" ", "")
            for idx, fixed_shift in enumerate(fixed_shifts_today):
                fixed_clean = str(fixed_shift).strip().lower().replace(" ", "")
                if req_clean == fixed_clean:
                    filled_idx = idx
                    break
            if filled_idx >= 0:
                fixed_shifts_today.pop(filled_idx)
            else:
                remaining_shifts_to_fill.append(shift_req)
                
        shifts_to_fill = remaining_shifts_to_fill
        shifts_to_fill = sorted(shifts_to_fill, key=lambda x: x["duration"], reverse=True)

        for shift_info in shifts_to_fill:
            best_candidate = None
            best_score = -999999

            for emp in active_employees:
                name = emp["Name"]
                norm_name = emp["NormalizedName"]
                age = emp["Age"]
                
                if roster_output[name][day] not in ["", "off"]:
                    continue
                
                if weekly_shifts_count[name] >= 5:
                    continue

                is_emp_unavailable = False
                key = (norm_name, day.lower())
                if key in unavail_map:
                    for window in unavail_map[key]:
                        if is_overlapping_unavailability(window, shift_info["start"], shift_info["end"]):
                            is_emp_unavailable = True
                            break
                if is_emp_unavailable:
                    continue

                if name == "Elizabeth":
                    if day in ["Saturday", "Sunday"]:
                        continue
                    if elizabeth_weekday_shifts >= 2:
                        continue

                if age < 18 and day not in ["Saturday", "Sunday"]:
                    if max(shift_info["start"], 9.0) < min(shift_info["end"], 15.5):
                        continue

                score = 0
                score -= weekly_shifts_count[name] * 20
                
                if age < 21:
                    score += (21 - age) * shift_info["duration"] * 2
                
                if name == "Elizabeth" and shift_info["start"] in [7.0, 7.5]:
                    score += 100

                if score > best_score:
                    best_score = score
                    best_candidate = name

            if best_candidate:
                roster_output[best_candidate][day] = shift_info["shift"]
                weekly_shifts_count[best_candidate] += 1
                if best_candidate == "Elizabeth" and day not in ["Saturday", "Sunday"]:
                    elizabeth_weekday_shifts += 1

    roster_rows = []
    for name, sched in roster_output.items():
        norm_name = name.lower()
        row = {"Employee": name}
        for day in days_of_week:
            val = str(sched.get(day, "")).strip()
            key = (norm_name, day.lower())
            if not val or val.lower() in ["off", "none", "nan", "null", "unavailable", ""] or val.lower().startswith("unavail") or val.startswith("🚫"):
                if key in unavail_map and unavail_map[key]:
                    clean_win = clean_win_display(unavail_map[key][0])
                    row[day] = f"🚫 Unavailable ({clean_win})"
                elif "unavail" in val.lower() or val.startswith("🚫"):
                    row[day] = "🚫 Unavailable (Full Day)"
                else:
                    row[day] = ""
            else:
                row[day] = val
        roster_rows.append(row)
    res_df = pd.DataFrame(roster_rows)
    res_df = sort_dataframe_by_team_and_age(res_df)
    return res_df

def render_home_dashboard():
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0e2b26 0%, #1a4d43 100%); padding: 22px; border-radius: 16px; border: 2px solid #e5a93c; box-shadow: 0 8px 30px rgba(0,0,0,0.4); margin-bottom: 25px;">
        <h2 style="color: #f7d594 !important; margin-top: 0; font-size: 1.85rem; font-weight: 800;">🏠 Executive Home & Store Dashboard</h2>
        <p style="color: #ffffff !important; font-size: 1.02rem; margin-bottom: 0;">Brumby's Bakery Pakenham • Live Store Notifications, Team Announcements & Operational KPI Hub</p>
    </div>
    """, unsafe_allow_html=True)

    melbourne_now = datetime.utcnow() + timedelta(hours=10)
    today_str = melbourne_now.strftime("%d/%m/%Y")
    
    # 1. Executive Operations & KPI Cards Bar
    df_cards = load_persisted_timecards()
    working_count = 0
    late_count = 0
    if df_cards is not None and not df_cards.empty and "Status" in df_cards.columns:
        if "Date" in df_cards.columns:
            working_count = len(df_cards[(df_cards["Status"] == "Working") & (df_cards["Date"] == today_str)])
            if "Note" in df_cards.columns:
                late_count = len(df_cards[df_cards["Note"].str.contains("Late Clocking", na=False) & (df_cards["Date"] == today_str)])

    past_rosters = list_finalized_rosters()
    active_rosters_count = len(past_rosters)
    active_emp_count = len(user_profiles)

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.markdown(f"""
        <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 12px; padding: 14px; text-align: center;">
            <div style="color: #e5a93c; font-size: 0.8rem; font-weight: 800;">🗓️ TODAY'S DATE</div>
            <div style="color: #ffffff; font-size: 1.4rem; font-weight: 900; margin-top: 4px;">{today_str}</div>
            <div style="color: #a0aec0; font-size: 0.78rem;">Melbourne AEST</div>
        </div>
        """, unsafe_allow_html=True)

    with k2:
        st.markdown(f"""
        <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 12px; padding: 14px; text-align: center;">
            <div style="color: #48bb78; font-size: 0.8rem; font-weight: 800;">🟢 ON DUTY TODAY</div>
            <div style="color: #ffffff; font-size: 1.4rem; font-weight: 900; margin-top: 4px;">{working_count} Staff</div>
            <div style="color: #a0aec0; font-size: 0.78rem;">Clocked In</div>
        </div>
        """, unsafe_allow_html=True)

    with k3:
        st.markdown(f"""
        <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 12px; padding: 14px; text-align: center;">
            <div style="color: #f7d594; font-size: 0.8rem; font-weight: 800;">📋 ACTIVE ROSTERS</div>
            <div style="color: #ffffff; font-size: 1.4rem; font-weight: 900; margin-top: 4px;">{active_rosters_count} Weeks</div>
            <div style="color: #a0aec0; font-size: 0.78rem;">Published</div>
        </div>
        """, unsafe_allow_html=True)

    with k4:
        st.markdown(f"""
        <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 12px; padding: 14px; text-align: center;">
            <div style="color: #76eec6; font-size: 0.8rem; font-weight: 800;">👥 TOTAL TEAM</div>
            <div style="color: #ffffff; font-size: 1.4rem; font-weight: 900; margin-top: 4px;">{active_emp_count} Members</div>
            <div style="color: #a0aec0; font-size: 0.78rem;">Active Profiles</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # 2. Main Dashboard Layout (2 Columns: Left = Store Announcements & Communication, Right = Notifications & Info)
    c_dash1, c_dash2 = st.columns([1.2, 1])

    with c_dash1:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #0e2b26 0%, #1a4d43 100%); padding: 12px 18px; border-radius: 10px 10px 0 0; color: #e5a93c !important; font-weight: 800; font-size: 1.15rem; border: 1.5px solid #e5a93c; border-bottom: none;">
            📢 Store Announcements & Team Bulletin Board
        </div>
        """, unsafe_allow_html=True)
        
        announcements = load_announcements()
        
        if announcements:
            for idx, ann in enumerate(announcements):
                prio = ann.get("priority", "Normal")
                badge_bg = "#e53e3e" if prio == "Urgent" else ("#dd6b20" if prio == "Important" else "#319795")
                st.markdown(f"""
                <div style="background: rgba(8, 29, 25, 0.95); border: 1.5px solid #1f5c50; border-radius: 8px; padding: 14px; margin-bottom: 10px;">
                    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 6px;">
                        <span style="font-weight: 800; font-size: 1.05rem; color: #f7d594;">{ann.get('title', '')}</span>
                        <span style="background: {badge_bg}; color: white; padding: 2px 8px; border-radius: 12px; font-size: 0.75rem; font-weight: 800;">{prio}</span>
                    </div>
                    <div style="color: #e2e8f0; font-size: 0.93rem; margin-bottom: 8px;">{ann.get('content', '')}</div>
                    <div style="color: #a0aec0; font-size: 0.78rem;">✍️ Posted by <b>{ann.get('author', 'Store Manager')}</b> on {ann.get('date', '')}</div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("ℹ️ No active store announcements.")

        with st.expander("➕ Post & Broadcast New Store Announcement", expanded=False):
            with st.form("form_new_announcement"):
                ann_title = st.text_input("Announcement Title:", placeholder="e.g. Easter Trading Hours & Shift Updates")
                ann_content = st.text_area("Announcement Message:", placeholder="Type store message for staff...")
                c_ann1, c_ann2 = st.columns(2)
                with c_ann1:
                    ann_prio = st.selectbox("Priority:", ["Normal", "Important", "Urgent"])
                with c_ann2:
                    ann_author = st.text_input("Author Name:", value="Viet (Store Owner)")
                
                broadcast_email = st.checkbox("✉️ Automatically Broadcast via Email to All Staff", value=True)
                
                btn_post_ann = st.form_submit_button("🚀 Publish Announcement", use_container_width=True)
                if btn_post_ann:
                    if ann_title.strip() and ann_content.strip():
                        new_id = f"ANN_{len(announcements)+1001}"
                        new_entry = {
                            "id": new_id,
                            "title": ann_title.strip(),
                            "content": ann_content.strip(),
                            "author": ann_author.strip(),
                            "date": datetime.now().strftime("%d/%m/%Y"),
                            "priority": ann_prio
                        }
                        announcements.insert(0, new_entry)
                        save_announcements(announcements)
                        
                        if broadcast_email:
                            with st.spinner("Broadcasting announcement via email to all active staff..."):
                                ok, msg_out = send_announcement_broadcast_smtp(ann_title.strip(), ann_content.strip(), ann_author.strip())
                                if ok:
                                    st.success(f"🎉 Announcement published & email broadcast complete! ({msg_out})")
                                else:
                                    st.warning(f"⚠️ Announcement published to board, but email broadcast failed: {msg_out}")
                        else:
                            st.success("🎉 Store announcement successfully published to board!")
                        st.rerun()
                    else:
                        st.error("❌ Please provide both Title and Message for the announcement.")

    with c_dash2:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #0e2b26 0%, #1a4d43 100%); padding: 12px 18px; border-radius: 10px 10px 0 0; color: #e5a93c !important; font-weight: 800; font-size: 1.15rem; border: 1.5px solid #e5a93c; border-bottom: none;">
            🔔 Notifications & System Alerts
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div style="background: rgba(8, 29, 25, 0.95); border: 1.5px solid #1f5c50; border-radius: 0 0 10px 10px; padding: 14px; margin-bottom: 15px;">
        """, unsafe_allow_html=True)

        if late_count > 0:
            st.warning(f"⚠️ **Attendance Alert**: {late_count} staff late clock-in punch(es) today requiring audit review.")
        else:
            st.success("✅ **Attendance Status**: All shift punches today are on time & verified.")

        unavail_df = load_persisted_df("unavailability.csv")
        u_count = len(unavail_df) if unavail_df is not None else 0
        st.info(f"📩 **Unavailability Records**: {u_count} staff availability constraints logged in database.")

        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("""
        <div style="background: linear-gradient(135deg, #0e2b26 0%, #1a4d43 100%); padding: 12px 18px; border-radius: 10px 10px 0 0; color: #e5a93c !important; font-weight: 800; font-size: 1.15rem; border: 1.5px solid #e5a93c; border-bottom: none;">
            ℹ️ Store Profile & Award Compliance
        </div>
        <div style="background: rgba(8, 29, 25, 0.95); border: 1.5px solid #1f5c50; border-radius: 0 0 10px 10px; padding: 14px;">
            <div style="font-weight: 800; color: #f7d594; font-size: 1.0rem;">🏪 Brumby's Bakery Pakenham</div>
            <div style="color: #cbd5e0; font-size: 0.88rem; margin: 4px 0 8px 0;">
                📍 <b>GPS Location:</b> <code>-38.063557, 145.455262</code><br>
                📜 <b>Award Compliance:</b> General Retail Industry Award 2020 (MA000004)<br>
                🧺 <b>Laundry Allowance:</b> $1.28 / shift worked<br>
                🏦 <b>ATO Super Guarantee:</b> 12.5% SG
            </div>
        </div>
        """, unsafe_allow_html=True)

if is_manager:
    # --- TAB 1: HOME / DASHBOARD ---
    with tab_dash:
        render_home_dashboard()

    # --- TAB 2: ROSTER INSIDE ---
    with tab_roster:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #0e2b26 0%, #1a4d43 100%); padding: 20px; border-radius: 16px; border: 2px solid #e5a93c; box-shadow: 0 8px 30px rgba(0,0,0,0.4); margin-bottom: 25px;">
            <h2 style="color: #f7d594 !important; margin-top: 0; font-size: 1.8rem; font-weight: 800;">📅 Roster Inside — Admin Command Center</h2>
            <p style="color: #ffffff !important; font-size: 1.05rem; margin-bottom: 0;">Select published weekly rosters to view, edit shifts, review real-time payroll breakdowns, and analyze historical financial trends.</p>
        </div>
        """, unsafe_allow_html=True)

        with st.expander("📥 Bulk Upload & System Auto-Scan for Old Rosters", expanded=False):
            col_sc1, col_sc2 = st.columns([1, 1])
            with col_sc1:
                st.markdown("#### 🔍 Auto-Scan System Folders")
                st.write("Scan Downloads, Desktop, and project folders to automatically import all historical roster files.")
                if st.button("🚀 Start Deep System Scan for Rosters", key="btn_deep_scan_rosters", use_container_width=True):
                    num_found = auto_import_reference_rosters(force_scan=True)
                    st.success(f"🎉 Deep system scan complete! Imported / refreshed {num_found} old roster(s).")
                    st.rerun()
            with col_sc2:
                st.markdown("#### 📤 Drag & Drop Old Roster Files")
                st.write("Upload multiple past roster Excel (.xlsx) or CSV files directly to publish them in the app.")
                uploaded_old_files = st.file_uploader(
                    "Upload Old Roster Files",
                    type=["xlsx", "csv"],
                    accept_multiple_files=True,
                    key="bulk_uploader_old_rosters"
                )
                if uploaded_old_files:
                    bulk_count = 0
                    for u_file in uploaded_old_files:
                        u_key = f"bulk_file_done_{u_file.name}_{u_file.size}"
                        if st.session_state.get(u_key) is not True:
                            df_u = read_excel_robust(u_file) if u_file.name.endswith(".xlsx") else pd.read_csv(u_file, dtype=str)
                            if df_u is not None and not df_u.empty:
                                dt_u = extract_date_from_filename(u_file.name)
                                if dt_u is None:
                                    for col in df_u.columns:
                                        dt_cand = extract_date_from_filename(str(col))
                                        if dt_cand:
                                            dt_u = dt_cand
                                            break
                                if dt_u is None:
                                    dt_u = datetime.now().date()
                                save_finalized_roster(df_u, dt_u)
                                st.session_state[u_key] = True
                                bulk_count += 1
                    if bulk_count > 0:
                        st.success(f"🎉 Successfully imported {bulk_count} uploaded roster file(s)!")
                        st.rerun()

        past_rosters = list_finalized_rosters()
        if past_rosters:
            roster_options = {r["label"]: r for r in past_rosters}
            selected_label = st.selectbox("Select Finalized Roster Week to Display & Edit:", list(roster_options.keys()), key="home_select_past_roster")
            selected_info = roster_options[selected_label]
            
            archived_df = load_finalized_roster(selected_info["csv_filename"])
            if archived_df is not None and not archived_df.empty:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 14px 22px; border-radius: 12px 12px 0 0; border: 2px solid #e5a93c; border-bottom: none; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 10px; margin-top: 15px;">
                    <div style="display: flex; align-items: center; gap: 10px; flex-wrap: wrap;">
                        <span style="font-size: 1.3rem;">📅</span>
                        <span style="color: #ffffff !important; font-size: 1.15rem; font-weight: 800; letter-spacing: 0.3px;">Published Roster Schedule for:</span>
                        <span style="color: #f7d594 !important; font-size: 1.1rem; font-weight: 700;">{selected_label}</span>
                    </div>
                    <span style="background: rgba(229, 169, 60, 0.2); color: #f7d594 !important; font-size: 0.78rem; font-weight: 700; padding: 4px 12px; border-radius: 20px; border: 1px solid #e5a93c; letter-spacing: 0.3px;">✏️ Live & Editable</span>
                </div>
                """, unsafe_allow_html=True)
                
                # Live-editable dataframe for displayed roster
                edited_archived_df = st.data_editor(archived_df, num_rows="dynamic", key=f"edit_home_roster_{selected_info['date_str']}")
                
                try:
                    dt = datetime.strptime(selected_info["date_str"], "%Y-%m-%d").date()
                except:
                    dt = datetime.now().date()
                    
                # Action Buttons Row (Save Edits, Download, Delete)
                col_act1, col_act2, col_act3 = st.columns([1.2, 1.2, 1])
                with col_act1:
                    if st.button("💾 SAVE CHANGES TO ROSTER", key=f"btn_save_home_{selected_info['date_str']}", use_container_width=True):
                        save_finalized_roster(edited_archived_df, dt)
                        st.success(f"🎉 Changes to roster for week {selected_info['date_str']} successfully saved to disk!")
                        st.rerun()
                
                with col_act2:
                    archived_excel_bytes = build_roster_excel_bytes(edited_archived_df, dt)
                    past_filename = f"Team_Roster_{dt.strftime('%d.%m.%Y')}.xlsx"
                    st.download_button(
                        label=f"📥 Download {selected_info['date_str']} Roster (.XLSX)",
                        data=archived_excel_bytes,
                        file_name=past_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"btn_dl_home_{selected_info['date_str']}",
                        use_container_width=True
                    )
                    
                with col_act3:
                    if st.button(f"🗑️ Delete Roster", key=f"btn_del_home_{selected_info['date_str']}", use_container_width=True):
                        if delete_finalized_roster(selected_info["date_str"]):
                            st.success(f"🗑️ Finalized Roster for {selected_info['date_str']} permanently deleted!")
                            st.rerun()

                # Upload/Replace file for selected past roster week
                st.markdown("<br>", unsafe_allow_html=True)
                up_past_file = st.file_uploader(f"📤 Replace Roster File for Week {selected_info['date_str']} (.xlsx / .csv)", type=["xlsx", "csv"], key=f"up_home_{selected_info['date_str']}")
                if up_past_file is not None:
                    past_up_key = f"home_up_{up_past_file.name}_{up_past_file.size}_{selected_info['date_str']}"
                    if st.session_state.get("last_home_upload_key") != past_up_key:
                        df_p_up = read_excel_robust(up_past_file)
                        if df_p_up is not None and not df_p_up.empty:
                            save_finalized_roster(df_p_up, dt)
                            st.session_state.last_home_upload_key = past_up_key
                            st.success(f"🎉 Finalized roster for week {selected_info['date_str']} updated via file upload!")
                            st.rerun()

                # Real-Time Wage, Tax & Super Breakdown for Displayed Roster
                st.markdown("<br>", unsafe_allow_html=True)
                wages_summary = calculate_roster_wages(edited_archived_df)
                
                st.markdown("""
                <div style="background: linear-gradient(135deg, #0e2b26 0%, #1a4d43 100%); padding: 14px 20px; border-radius: 12px 12px 0 0; color: #e5a93c !important; font-weight: 800; font-size: 1.25rem; border: 2px solid #e5a93c; border-bottom: none;">
                    💰 Real-Time Wage, Tax & Super Summary
                </div>
                """, unsafe_allow_html=True)
                
                summary_cards_html = f"""
                <div style="background: rgba(8, 29, 25, 0.85); border: 2px solid #e5a93c; border-top: none; border-radius: 0 0 12px 12px; padding: 20px; margin-bottom: 20px;">
                    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 18px;">
                        <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 10px; padding: 14px; text-align: center;">
                            <div style="color: #e5a93c; font-size: 0.82rem; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">💵 Total Gross Payroll</div>
                            <div style="color: #ffffff; font-size: 1.75rem; font-weight: 900; margin-top: 6px;">${wages_summary['total_gross']:,.2f}</div>
                        </div>
                        <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 10px; padding: 14px; text-align: center;">
                            <div style="color: #f7d594; font-size: 0.82rem; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">🏛️ Est. PAYG Tax</div>
                            <div style="color: #ffffff; font-size: 1.75rem; font-weight: 900; margin-top: 6px;">${wages_summary['total_tax']:,.2f}</div>
                        </div>
                        <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 10px; padding: 14px; text-align: center;">
                            <div style="color: #76eec6; font-size: 0.82rem; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">👛 Total Net Take-Home</div>
                            <div style="color: #76eec6; font-size: 1.75rem; font-weight: 900; margin-top: 6px;">${wages_summary['total_net']:,.2f}</div>
                        </div>
                        <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 10px; padding: 14px; text-align: center;">
                            <div style="color: #f7d594; font-size: 0.82rem; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">🏦 Super (12.5% SG)</div>
                            <div style="color: #ffffff; font-size: 1.75rem; font-weight: 900; margin-top: 6px;">${wages_summary['total_super']:,.2f}</div>
                        </div>
                    </div>
                    <div style="background: rgba(229, 169, 60, 0.12); border: 1px solid rgba(229, 169, 60, 0.4); border-radius: 8px; padding: 10px 16px; text-align: center; color: #ffffff; font-size: 1.05rem;">
                        ⏱️ <b>Total Paid Hours:</b> <span style="color:#e5a93c; font-weight:800;">{wages_summary['total_hours']} hrs</span> &nbsp;&nbsp;|&nbsp;&nbsp; 📊 <b>Average Hourly Rate:</b> <span style="color:#e5a93c; font-weight:800;">${wages_summary['avg_hourly_rate']:.2f} / hr</span>
                    </div>
                </div>
                """
                st.markdown(summary_cards_html, unsafe_allow_html=True)
                
                sub_tab1, sub_tab2, sub_tab3 = st.tabs([
                    f"👥 Staff Earnings & Super Breakdown",
                    f"📊 Hour Rate Breakdown ({selected_label})",
                    "📈 Historical Payroll Progress & Trends"
                ])
                
                with sub_tab1:
                    st.markdown(f"#### 👥 Staff Earnings & Super Breakdown Table ({selected_label})")
                    if not wages_summary["breakdown_df"].empty:
                        st.dataframe(wages_summary["breakdown_df"], use_container_width=True, hide_index=True)

                with sub_tab2:
                    st.markdown(f"#### 📊 Hour Rate Breakdown for Selected Roster Week: `{selected_label}`")
                    home_hour_breakdown_df = calculate_weekly_hour_rate_breakdown(edited_archived_df)
                    if not home_hour_breakdown_df.empty:
                        st.dataframe(home_hour_breakdown_df, use_container_width=True, hide_index=True)
                        
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.markdown("""
                        <div style="background: rgba(8, 29, 25, 0.95); border: 2px solid #e5a93c; border-radius: 12px; padding: 14px 20px; margin-top: 10px;">
                            <h4 style="color: #e5a93c; margin: 0 0 6px 0;">⚡ Xero Payroll Automated Input Panel</h4>
                            <p style="color: #d0e6df; font-size: 0.95rem; margin: 0 0 10px 0;">Automatically populate this breakdown into your Xero Pay Run web page. <b>Rule Enforced:</b> If an employee name in the bakery app does not match Xero's web table, their input is safely skipped.</p>
                        </div>
                        """, unsafe_allow_html=True)

                        xero_c1, xero_c2 = st.columns([1.5, 1])
                        with xero_c1:
                            xero_url = st.text_input("Xero Pay Run Target URL", value="https://payroll.xero.com/PayRun/PayRun/Details/74565804?CID=!cSWq8", key="xero_target_url_input")
                            btn_xero_bot = st.button("🚀 Launch Auto-Input Bot to Xero", key="btn_run_xero_bot", use_container_width=True)
                        with xero_c2:
                            st.markdown("<br>", unsafe_allow_html=True)
                            st.markdown("[🔗 Open Xero Pay Run Page](https://payroll.xero.com/PayRun/PayRun/Details/74565804?CID=!cSWq8)")
                            xero_csv_data = generate_xero_timesheet_csv(home_hour_breakdown_df)
                            if xero_csv_data:
                                st.download_button(
                                    label="📥 Download Xero Timesheet CSV",
                                    data=xero_csv_data,
                                    file_name=f"Xero_Payroll_Import_{selected_info['date_str']}.csv",
                                    mime="text/csv",
                                    key="btn_download_xero_csv",
                                    use_container_width=True
                                )

                        if btn_xero_bot:
                            with st.spinner("🤖 Connecting to browser and populating Xero Pay Run..."):
                                try:
                                    success, res_msg = run_xero_playwright_autofill(home_hour_breakdown_df, target_url=xero_url, headless=False)
                                    if success:
                                        st.success(res_msg)
                                    else:
                                        st.warning(res_msg)
                                except Exception as ex:
                                    st.error(f"Xero Bot Launch Error: {ex}")

                        with st.expander("⚡ 1-Click Browser Auto-Fill Script (For Web / Cloud Deployment)"):
                            st.markdown("Copy the script code below, open your Xero Pay Run page in your web browser (`F12` -> `Console`), and paste it to auto-fill all staff hours in 1 second:")
                            js_autofill_code = generate_xero_autofill_js(home_hour_breakdown_df)
                            st.code(js_autofill_code, language="javascript")

                with sub_tab3:
                    st.markdown("#### 📈 Payroll, Tax & Super Progress Over Time (Historical Trend Graph)")
                    trend_df = build_payroll_historical_trend()
                    if not trend_df.empty and len(trend_df) >= 1:
                        st.markdown("Historical trend analysis of **Gross Payroll**, **Est. PAYG Tax**, **Net Take-Home**, and **Super (12.5% SG)** across all finalized weekly rosters (chronologically ordered):")
                        chart_df = trend_df.copy()
                        chart_df["Date"] = pd.to_datetime(chart_df["date"])
                        chart_df = chart_df.sort_values("Date").set_index("Date").drop(columns=["date", "Roster Week"], errors="ignore")
                        st.line_chart(chart_df, use_container_width=True)
                        display_df = trend_df.drop(columns=["date"], errors="ignore")
                        st.dataframe(display_df, use_container_width=True, hide_index=True)
                    else:
                        st.info("ℹ️ Click below to load and analyze historical published rosters.")
                        if st.button("🔄 Load Historical Roster Data & Line Graph", key="btn_load_past_home_2", use_container_width=True):
                            auto_import_reference_rosters()
                            st.rerun()
        else:
            st.info("ℹ️ No finalized rosters displayed yet.")
            if st.button("🔄 Auto-Scan & Restore Published Master Rosters", key="btn_load_past_home_1", use_container_width=True):
                auto_import_reference_rosters()
                st.rerun()

    # --- TAB 2: WEEKLY ROSTER GENERATOR ---
    with tab_gen:
        st.markdown("""
        <div style="background: rgba(9, 32, 28, 0.7); border: 2px solid #e5a93c; border-radius: 16px; padding: 25px; margin-bottom: 25px; box-shadow: 0 8px 30px rgba(0,0,0,0.4);">
            <h2 style="color: #f7d594 !important; margin-top: 0; font-size: 1.8rem; font-weight: 800;">⚡ Weekly Roster Generator</h2>
            <p style="color: #ffffff !important; font-size: 1.05rem; margin-bottom: 0;">Configure your target week period below and hit the <b>Generate Weekly Roster</b> button to instantly build an award-compliant bakery schedule.</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns([1, 1])
        with col1:
            start_date = st.date_input("🗓️ Roster Start Date (Monday)", datetime.now() + timedelta(days=(0 - datetime.now().weekday())), key="gen_start_date")
            
            st.markdown('<div class="hero-generate-btn">', unsafe_allow_html=True)
            if st.button("🚀 GENERATE WEEKLY ROSTER", key="btn_hero_generate"):
                with st.spinner("Calculating optimal bakery roster locally..."):
                    try:
                        emp_data = st.session_state.manual_employees
                        unavail_data = st.session_state.manual_unavailability
                        req_data = st.session_state.manual_requirements
                        fixed_data = st.session_state.manual_fixed
                        
                        roster_out_df = solve_roster(emp_data, unavail_data, req_data, fixed_data, start_date)
                        df_clean = roster_out_df.replace(["off", "Off", "OFF", "None", "none", "nan", "NaN", None], "").fillna("")
                        emp_c = find_column(df_clean, ["employee", "name", "staff"], "Employee")
                        if emp_c in df_clean.columns:
                            df_clean = df_clean[~df_clean[emp_c].astype(str).str.strip().str.lower().isin(["", "none", "nan"])].reset_index(drop=True)
                        df_clean = clean_roster_dataframe(df_clean)
                        df_clean = sort_dataframe_by_team_and_age(df_clean)
                        st.session_state.final_roster_df = df_clean
                        st.success("🎉 Weekly Roster successfully generated!")
                    except Exception as e:
                        st.error(f"Failed to generate roster: {e}")
            
            st.markdown("<br>", unsafe_allow_html=True)
            upload_roster_file = st.file_uploader("📤 OR Upload Existing Roster File (.xlsx / .csv)", type=["xlsx", "csv"], key="upload_roster_selected_week")
            if upload_roster_file is not None:
                file_key = f"roster_up_{upload_roster_file.name}_{upload_roster_file.size}_{start_date}"
                if st.session_state.get("last_uploaded_roster_key") != file_key:
                    df_up = read_excel_robust(upload_roster_file)
                    if df_up is not None and not df_up.empty:
                        df_clean = df_up.replace(["off", "Off", "OFF", "None", "none", "nan", "NaN", None], "").fillna("")
                        emp_c = find_column(df_clean, ["employee", "name", "staff"], "Employee")
                        if emp_c in df_clean.columns:
                            df_clean = df_clean[~df_clean[emp_c].astype(str).str.strip().str.lower().isin(["", "none", "nan"])].reset_index(drop=True)
                        df_clean = clean_roster_dataframe(df_clean)
                        df_clean = sort_dataframe_by_team_and_age(df_clean)
                        st.session_state.final_roster_df = df_clean
                        st.session_state.last_uploaded_roster_key = file_key
                        # Save immediately to disk so it is persisted after turn off / restart
                        save_finalized_roster(df_clean, start_date)
                        st.success(f"🎉 Roster loaded & permanently saved to disk for week starting {start_date.strftime('%d/%m/%Y')}!")

        with col2:
            st.markdown("""
            <div style="background: rgba(9, 32, 28, 0.5); border: 1px solid rgba(229, 169, 60, 0.4); border-radius: 14px; padding: 15px; height: 100%;">
                <h4 style="color: #e5a93c !important; margin-top: 0;">📋 Generator Rules Summary</h4>
                <ul style="margin-bottom: 0; padding-left: 20px; font-size: 0.95rem; color: #ffffff !important;">
                    <li>Respects staff unavailability constraints</li>
                    <li>Fulfills daily shift requirements</li>
                    <li>Ensures mandatory award break times</li>
                    <li>Enforces minimum rest periods between shifts</li>
                    <li>Optimizes total wage costs (junior rate prioritization & penalty minimization)</li>
                    <li>Organizes table layout: Baking Team first, followed by Service Team (Age Descending)</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)

        if 'final_roster_df' in st.session_state and st.session_state.final_roster_df is not None and not st.session_state.final_roster_df.empty:
            st.markdown("<br>", unsafe_allow_html=True)
            
            col_tbl_h1, col_tbl_h2 = st.columns([2, 1])
            with col_tbl_h1:
                st.markdown("""
                <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 12px 20px; border-radius: 12px 12px 0 0; color: #ffffff !important; font-weight: 800; font-size: 1.2rem; letter-spacing: 0.5px; border: 2px solid #e5a93c; border-bottom: none;">
                    📅 Generated Weekly Roster Schedule (Editable)
                </div>
                """, unsafe_allow_html=True)
            with col_tbl_h2:
                show_unavail = st.checkbox("👁️ Show Staff Unavailability", value=True, key="chk_show_unavailability")
            
            # Mobile View Mode & Zoom Level Controls Side-by-Side
            col_mode, col_zoom = st.columns([1.8, 1.2])
            with col_mode:
                roster_view_mode = st.radio(
                    "📱 Mobile Layout Mode:",
                    ["📊 Full 7-Day Table", "📅 Single Day Focus", "🎴 Mobile Staff Cards"],
                    key="roster_view_mode",
                    horizontal=True
                )
            with col_zoom:
                roster_zoom_val = st.select_slider(
                    "🔍 Table Zoom Level:",
                    options=["60%", "65%", "70%", "75%", "80%", "85%", "90%", "95%", "100%", "105%", "110%"],
                    value="100%",
                    key="roster_zoom_slider"
                )

            # Strip out any existing summary row first to get pure staff dataframe
            st.session_state.final_roster_df = strip_daily_gross_row(st.session_state.final_roster_df)
            st.session_state.final_roster_df = sort_dataframe_by_team_and_age(st.session_state.final_roster_df)
            
            # Calculate wages & daily gross breakdown
            wages_summary_gen = calculate_roster_wages(st.session_state.final_roster_df)
            daily_gross_map = wages_summary_gen.get("daily_gross", {})

            # Attach bottom summary row for visual display in data_editor
            df_for_editor = attach_daily_gross_row(st.session_state.final_roster_df, daily_gross_map)

            # Calculate zoom factor and dynamic column width / height scaling
            zoom_pct = int(roster_zoom_val.replace("%", ""))
            zoom_factor = zoom_pct / 100.0
            
            # Calculate pixel widths for Employee and Monday..Sunday columns
            emp_col_width = int(170 * zoom_factor)
            day_col_width = int(135 * zoom_factor)
            
            # Dynamic Column Configuration for Zoom Sizing
            dynamic_col_config = {}
            if df_for_editor is not None and not df_for_editor.empty:
                first_col = df_for_editor.columns[0]
                dynamic_col_config[first_col] = st.column_config.Column(
                    label=first_col,
                    width=emp_col_width
                )
                for c in df_for_editor.columns[1:]:
                    dynamic_col_config[c] = st.column_config.Column(
                        label=c,
                        width=day_col_width
                    )

            # Compute dynamic zoom-scaled table height
            num_roster_rows = len(df_for_editor) if df_for_editor is not None else 0
            roster_table_height = max(int(250 * zoom_factor), int((num_roster_rows + 1) * (38 * zoom_factor) + 30))
            
            # CSS font & canvas variable scaling
            font_size_px = max(9, round(14 * zoom_factor, 1))
            header_font_size_px = max(10, round(15 * zoom_factor, 1))
            
            st.markdown(f"""
            <style>
                div[data-testid="stDataEditor"] {{
                    font-size: {font_size_px}px !important;
                    --gdg-font-size: {font_size_px}px !important;
                }}
                div[data-testid="stDataEditor"] th, 
                div[data-testid="stDataEditor"] div[role="columnheader"] {{
                    font-size: {header_font_size_px}px !important;
                }}
            </style>
            """, unsafe_allow_html=True)

            if roster_view_mode == "📅 Single Day Focus":
                selected_day = st.selectbox("Select Day to Inspect:", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"], key="sel_single_day_focus")
                emp_col = df_for_editor.columns[0]
                day_cols = [emp_col, selected_day] if selected_day in df_for_editor.columns else list(df_for_editor.columns)
                edited_display_df = st.data_editor(
                    df_for_editor[day_cols],
                    num_rows="dynamic",
                    key="edit_generated_roster_single",
                    height=roster_table_height,
                    column_config=dynamic_col_config,
                    use_container_width=(zoom_pct == 100)
                )
                edited_final_df = strip_daily_gross_row(edited_display_df)
                
                day_g = daily_gross_map.get(selected_day, 0.0)
                st.markdown(f"""
                <div style="background: rgba(8, 29, 25, 0.95); border: 2px solid #e5a93c; border-radius: 10px; padding: 12px 18px; margin-top: 8px; text-align: center;">
                    <span style="color: #e5a93c; font-weight: 800; font-size: 1.1rem;">💵 {selected_day} Predicted Gross Payroll: </span>
                    <span style="color: #ffffff; font-weight: 900; font-size: 1.3rem; margin-left: 8px;">${day_g:,.2f}</span>
                </div>
                """, unsafe_allow_html=True)

            elif roster_view_mode == "🎴 Mobile Staff Cards":
                edited_final_df = strip_daily_gross_row(st.session_state.final_roster_df)
                
                # Daily Gross Header Strip for Mobile Staff Cards
                daily_card_strip = "".join([
                    f"<div style='background:#0d332b; border:1px solid #e5a93c; border-radius:8px; padding:6px 10px; text-align:center; min-width:85px; margin:2px;'>"
                    f"<div style='color:#e5a93c; font-size:0.75rem; font-weight:800;'>{d[:3]}</div>"
                    f"<div style='color:#ffffff; font-weight:900; font-size:0.95rem;'>${daily_gross_map.get(d, 0.0):,.2f}</div>"
                    f"</div>"
                    for d in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                ])
                st.markdown(f"""
                <div style="background: rgba(8, 29, 25, 0.95); border: 1.5px solid #e5a93c; border-radius: 12px; padding: 10px 14px; margin-bottom: 15px;">
                    <div style="color: #f7d594; font-weight: 800; font-size: 1.0rem; margin-bottom: 8px;">💵 Daily Gross Payroll Breakdown</div>
                    <div style="display: flex; overflow-x: auto; gap: 4px; padding-bottom: 4px;">{daily_card_strip}</div>
                </div>
                """, unsafe_allow_html=True)

                days_cols = [c for c in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"] if c in edited_final_df.columns]
                emp_col = edited_final_df.columns[0]
                for idx, r_row in edited_final_df.iterrows():
                    emp_n = r_row.get(emp_col, f"Employee #{idx+1}")
                    shifts_html = "".join([f"<span style='background:#0d332b; border:1px solid #e5a93c; border-radius:6px; padding:4px 8px; margin:2px; font-size:0.85rem; display:inline-block;'><b>{d[:3]}:</b> {r_row.get(d, 'OFF')}</span>" for d in days_cols if str(r_row.get(d, '')).strip()])
                    if not shifts_html:
                        shifts_html = "<span style='color:#aaaaaa; font-style:italic;'>No shifts assigned</span>"
                    st.markdown(f"""
                    <div style="background: rgba(8, 29, 25, 0.95); border: 1.5px solid #e5a93c; border-radius: 12px; padding: 12px 16px; margin-bottom: 10px; box-shadow: 0 4px 12px rgba(0,0,0,0.3);">
                        <div style="color: #f7d594; font-weight: 800; font-size: 1.1rem; margin-bottom: 6px;">👤 {emp_n}</div>
                        <div style="display: flex; flex-wrap: wrap; gap: 4px;">{shifts_html}</div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                edited_display_df = st.data_editor(
                    df_for_editor,
                    num_rows="dynamic",
                    key="edit_generated_roster",
                    height=roster_table_height,
                    column_config=dynamic_col_config,
                    use_container_width=(zoom_pct == 100)
                )
                edited_final_df = strip_daily_gross_row(edited_display_df)

            # Real-Time Financial Breakdown for Generated Roster
            wages_summary_gen = calculate_roster_wages(edited_final_df)
            
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("""
            <div style="background: linear-gradient(135deg, #0e2b26 0%, #1a4d43 100%); padding: 14px 20px; border-radius: 12px 12px 0 0; color: #e5a93c !important; font-weight: 800; font-size: 1.25rem; border: 2px solid #e5a93c; border-bottom: none;">
                💰 Real-Time Wage, Tax & Super Summary
            </div>
            """, unsafe_allow_html=True)
            
            gen_summary_cards_html = f"""
            <div style="background: rgba(8, 29, 25, 0.85); border: 2px solid #e5a93c; border-top: none; border-radius: 0 0 12px 12px; padding: 16px; margin-bottom: 20px;">
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(135px, 1fr)); gap: 12px; margin-bottom: 18px;">
                    <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 10px; padding: 14px; text-align: center;">
                        <div style="color: #e5a93c; font-size: 0.82rem; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">💵 Total Gross Payroll</div>
                        <div style="color: #ffffff; font-size: 1.75rem; font-weight: 900; margin-top: 6px;">${wages_summary_gen['total_gross']:,.2f}</div>
                    </div>
                    <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 10px; padding: 14px; text-align: center;">
                        <div style="color: #f7d594; font-size: 0.82rem; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">🏛️ Est. PAYG Tax</div>
                        <div style="color: #ffffff; font-size: 1.75rem; font-weight: 900; margin-top: 6px;">${wages_summary_gen['total_tax']:,.2f}</div>
                    </div>
                    <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 10px; padding: 14px; text-align: center;">
                        <div style="color: #76eec6; font-size: 0.82rem; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">👛 Total Net Take-Home</div>
                        <div style="color: #76eec6; font-size: 1.75rem; font-weight: 900; margin-top: 6px;">${wages_summary_gen['total_net']:,.2f}</div>
                    </div>
                    <div style="background: #0d332b; border: 1.5px solid #e5a93c; border-radius: 10px; padding: 14px; text-align: center;">
                        <div style="color: #f7d594; font-size: 0.82rem; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">🏦 Super (12.5% SG)</div>
                        <div style="color: #ffffff; font-size: 1.75rem; font-weight: 900; margin-top: 6px;">${wages_summary_gen['total_super']:,.2f}</div>
                    </div>
                </div>
                <div style="background: rgba(229, 169, 60, 0.12); border: 1px solid rgba(229, 169, 60, 0.4); border-radius: 8px; padding: 10px 16px; text-align: center; color: #ffffff; font-size: 1.05rem;">
                    ⏱️ <b>Total Paid Hours:</b> <span style="color:#e5a93c; font-weight:800;">{wages_summary_gen['total_hours']} hrs</span> &nbsp;&nbsp;|&nbsp;&nbsp; 📊 <b>Average Hourly Rate:</b> <span style="color:#e5a93c; font-weight:800;">${wages_summary_gen['avg_hourly_rate']:.2f} / hr</span>
                </div>
            </div>
            """
            st.markdown(gen_summary_cards_html, unsafe_allow_html=True)
            
            st.markdown("#### 👥 Staff Earnings & Super Breakdown Table")
            if not wages_summary_gen["breakdown_df"].empty:
                st.dataframe(wages_summary_gen["breakdown_df"], use_container_width=True, hide_index=True)

            gen_hour_breakdown_df = calculate_weekly_hour_rate_breakdown(edited_final_df)
            if not gen_hour_breakdown_df.empty:
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("#### 📊 Hour Rate Breakdown Table (Generated Roster)")
                st.dataframe(gen_hour_breakdown_df, use_container_width=True, hide_index=True)

            # Finalize & Export Section
            st.markdown("<br>", unsafe_allow_html=True)
            col_fin1, col_fin2 = st.columns([1.2, 1])
            with col_fin1:
                if st.button("🔒 FINALIZE WEEKLY ROSTER", key="btn_finalize_roster", use_container_width=True):
                    date_str, xlsx_filename, excel_bytes = save_finalized_roster(edited_final_df, start_date)
                    st.success(f"🎉 Weekly Roster for {start_date.strftime('%d/%m/%Y')} successfully finalized and saved online!")
            
            with col_fin2:
                excel_bytes = build_roster_excel_bytes(edited_final_df, start_date)
                file_name_out = f"Team_Roster_{start_date.strftime('%d.%m.%Y')}.xlsx"
                st.download_button(
                    label="📥 DOWNLOAD CURRENT ROSTER (.XLSX)",
                    data=excel_bytes,
                    file_name=file_name_out,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_export_excel",
                    use_container_width=True
                )
            
            # --- TAB 2: STAFF MEMBERS ---
    with tab_emp:
        st.subheader("Manage Bakery Employees")
        if "email_notice_success" in st.session_state:
            st.success(st.session_state.pop("email_notice_success"))
        if "email_notice_warning" in st.session_state:
            st.warning(st.session_state.pop("email_notice_warning"))

        emp_mode = st.radio("Upload Mode:", ["Replace current data", "Append to current data"], key="emp_upload_mode", horizontal=True)
        upload_emp = st.file_uploader("Upload EMPLOYEE LIST.xlsx (Optional)", type=["xlsx"], key="emp_upload")
        
        if upload_emp is not None:
            file_key = f"processed_{upload_emp.name}_{upload_emp.size}_{emp_mode}"
            if st.session_state.get("last_emp_file") != file_key:
                loaded = read_excel_robust(upload_emp)
                if loaded is not None:
                    loaded = cleanup_duplicate_employee_columns(loaded)
                    if emp_mode == "Replace current data":
                        st.session_state.manual_employees = loaded
                    else:
                        combined = pd.concat([st.session_state.manual_employees, loaded], ignore_index=True).drop_duplicates()
                        st.session_state.manual_employees = cleanup_duplicate_employee_columns(combined)
                    st.session_state.manual_employees = sync_user_profiles_to_employees(st.session_state.manual_employees)
                    st.session_state.last_emp_file = file_key
                    save_persisted_df(st.session_state.manual_employees, "employees.csv")
                    if "edit_employees" in st.session_state:
                        del st.session_state["edit_employees"]
                    st.rerun()
                    
        if st.session_state.manual_employees is not None and not st.session_state.manual_employees.empty:
            st.session_state.manual_employees = sync_user_profiles_to_employees(st.session_state.manual_employees)
        
        # Prepare dataframe for data_editor
        df_for_editor = sort_dataframe_by_team_and_age(st.session_state.manual_employees.copy()) if st.session_state.manual_employees is not None else pd.DataFrame()
        if "Select" in df_for_editor.columns:
            df_for_editor.drop(columns=["Select"], inplace=True)
        if "🗑️ Delete" in df_for_editor.columns:
            df_for_editor.drop(columns=["🗑️ Delete"], inplace=True)

        name_col = find_column(st.session_state.manual_employees, ["name", "employee", "staff"], "NAME") if (st.session_state.manual_employees is not None and not st.session_state.manual_employees.empty) else "NAME"

        # Check if any employee checkbox is currently checked in session state
        selected_names_pre = []
        if "edit_employees" in st.session_state and isinstance(st.session_state.edit_employees, dict):
            edited_rows = st.session_state.edit_employees.get("edited_rows", {})
            for r_idx_str, changes in edited_rows.items():
                if changes.get("Select") is True:
                    try:
                        r_idx = int(r_idx_str)
                        if st.session_state.manual_employees is not None and 0 <= r_idx < len(st.session_state.manual_employees):
                            val = str(st.session_state.manual_employees.iloc[r_idx][name_col]).strip()
                            if val:
                                selected_names_pre.append(val)
                    except:
                        pass

        # UNIFIED TABLE HEADER CONTAINER WITH INTEGRATED DELETE BUTTON
        st.markdown("""
        <style>
        /* Style the stHorizontalBlock to be the single unified header bar */
        div[data-testid="stHorizontalBlock"]:has(#staff-table-hdr-mark) {
            background: linear-gradient(135deg, #081d19 0%, #16443c 100%) !important;
            border: 2px solid #e5a93c !important;
            border-bottom: none !important;
            border-radius: 12px 12px 0 0 !important;
            margin-top: 15px !important;
            padding: 8px 16px !important;
            min-height: 54px !important;
            align-items: center !important;
        }
        #staff-table-hdr-mark {
            color: #ffffff !important;
            font-weight: 800 !important;
            font-size: 1.1rem !important;
            letter-spacing: 0.3px !important;
            display: flex !important;
            align-items: center !important;
        }
        div[data-testid="stHorizontalBlock"]:has(#staff-table-hdr-mark) div[data-testid="stColumn"]:nth-child(2) {
            display: flex !important;
            justify-content: flex-end !important;
            align-items: center !important;
        }
        button[key="btn_header_tiny_trash"] {
            background: linear-gradient(135deg, #ff4d4f 0%, #cf1322 100%) !important;
            color: #ffffff !important;
            border: 1px solid #ff7875 !important;
            border-radius: 6px !important;
            border-radius: 6px !important;
            font-weight: 700 !important;
            font-size: 0.85rem !important;
            padding: 4px 14px !important;
            height: 36px !important;
            margin: 0 !important;
            box-shadow: 0 2px 4px rgba(0,0,0,0.2) !important;
        }
        </style>
        """, unsafe_allow_html=True)

        c_hdr_left, c_hdr_right = st.columns([7.5, 2.5])
        with c_hdr_left:
            st.markdown('<div id="staff-table-hdr-mark">👥 Bakery Staff Members List (Editable Table)</div>', unsafe_allow_html=True)
        with c_hdr_right:
            if selected_names_pre:
                trigger_delete = st.button(f"🗑️ Delete ({len(selected_names_pre)})", key="btn_header_tiny_trash", help=f"Delete selected staff: {', '.join(selected_names_pre)}", type="primary")
            else:
                trigger_delete = False

        if not df_for_editor.empty:
            df_for_editor.insert(0, "Select", False)

        employees_df = st.data_editor(
            df_for_editor,
            num_rows="dynamic",
            hide_index=True,
            key="edit_employees",
            column_config={
                "Select": st.column_config.CheckboxColumn("", help="Check box to select employee for deletion", default=False)
            }
        )

        if employees_df is not None:
            # Check selected names from employees_df
            selected_names = []
            if "Select" in employees_df.columns:
                selected_rows = employees_df[employees_df["Select"] == True]
                if not selected_rows.empty and name_col in selected_rows.columns:
                    selected_names = [str(n).strip() for n in selected_rows[name_col].dropna().tolist() if str(n).strip()]

            # Execute deletion if tiny header trash icon was clicked
            if trigger_delete and (selected_names or selected_names_pre):
                target_del_names = selected_names if selected_names else selected_names_pre
                target_lower = [n.lower() for n in target_del_names]

                # 1. Remove from user_profiles.json (if account exists)
                profiles_to_update = get_active_user_profiles()
                profiles_changed = False
                for del_name in target_del_names:
                    keys_to_del = []
                    for u_k, u_v in profiles_to_update.items():
                        emp_n = u_v.get("employee_name", u_k).strip()
                        full_n = u_v.get("profile", {}).get("full_name", "").strip()
                        if emp_n.lower() == del_name.lower() or full_n.lower() == del_name.lower() or u_k.lower() == del_name.lower():
                            keys_to_del.append(u_k)
                    for k in keys_to_del:
                        del profiles_to_update[k]
                        profiles_changed = True
                if profiles_changed:
                    save_user_profiles(profiles_to_update)

                # 2. Explicitly remove selected employees from manual_employees table (works even if they don't have an account!)
                if st.session_state.manual_employees is not None and not st.session_state.manual_employees.empty:
                    emp_df = st.session_state.manual_employees.copy()
                    n_col = find_column(emp_df, ["name", "employee", "staff"], "NAME")
                    if n_col in emp_df.columns:
                        emp_df = emp_df[~emp_df[n_col].astype(str).str.strip().str.lower().isin(target_lower)].reset_index(drop=True)
                        st.session_state.manual_employees = emp_df
                        save_persisted_df(st.session_state.manual_employees, "employees.csv")

                # 3. Clear widget cache and rerun
                if "edit_employees" in st.session_state:
                    del st.session_state["edit_employees"]

                st.success(f"✅ Deleted selected employee(s): {', '.join(target_del_names)}")
                st.rerun()

            # Normal table edits save
            clean_df = employees_df.drop(columns=["Select"], errors="ignore") if "Select" in employees_df.columns else employees_df
            st.session_state.manual_employees = cleanup_duplicate_employee_columns(clean_df)
            save_persisted_df(st.session_state.manual_employees, "employees.csv")

        # --- ➕ NEW EMPLOYEE ACCOUNT CREATION FORM & MISSING ACCOUNTS TOOL ---
        with st.expander("➕ Add New Staff Account / Auto-Create Missing Logins", expanded=False):
            # Check for employees in table without an account
            if st.session_state.manual_employees is not None and not st.session_state.manual_employees.empty:
                n_col = find_column(st.session_state.manual_employees, ["name", "employee", "staff"], "NAME")
                if n_col in st.session_state.manual_employees.columns:
                    all_names = [str(n).strip() for n in st.session_state.manual_employees[n_col].dropna().tolist() if str(n).strip()]
                    current_profiles = get_active_user_profiles()
                    account_names = set()
                    for u_k, u_v in current_profiles.items():
                        if u_v.get("role") == "Employee":
                            account_names.add(u_v.get("employee_name", u_k).strip().lower())
                            account_names.add(u_v.get("profile", {}).get("full_name", "").strip().lower())
                            account_names.add(u_k.lower())

                    unlinked = [name for name in set(all_names) if name.lower() not in account_names and "demo" not in name.lower()]
                    if unlinked:
                        st.info(f"💡 Found **{len(unlinked)} staff member(s)** in the table without a login account: **{', '.join(unlinked)}**")
                        if st.button("⚡ Auto-Create Login Accounts for All Missing Staff", key="btn_autocreate_missing_logins"):
                            created_count = 0
                            for un_name in unlinked:
                                base_user = un_name.lower().replace(" ", ".")
                                u_username = base_user
                                counter = 1
                                while u_username in current_profiles:
                                    u_username = f"{base_user}{counter}"
                                    counter += 1
                                
                                current_profiles[u_username] = {
                                    "username": u_username,
                                    "password": "TempPass123!",
                                    "role": "Employee",
                                    "employee_name": un_name,
                                    "profile": {
                                        "full_name": un_name,
                                        "email": "",
                                        "store": "Brumby's Pakenham",
                                        "classification": "Casual",
                                        "employment_level": "Service Staff",
                                        "commencement_date": datetime.now().strftime("%Y-%m-%d")
                                    }
                                }
                                created_count += 1
                            
                            save_user_profiles(current_profiles)
                            st.session_state.manual_employees = sync_user_profiles_to_employees(st.session_state.manual_employees)
                            save_persisted_df(st.session_state.manual_employees, "employees.csv")
                            if "edit_employees" in st.session_state:
                                del st.session_state["edit_employees"]
                            st.success(f"🎉 Successfully created {created_count} user login accounts! Initial Password: `TempPass123!`")
                            st.rerun()

            st.markdown("---")
            send_email_chk = st.checkbox("☑️ Send Welcome Email automatically", value=True, key="chk_send_welcome_email")
            
            emp_email_val = ""
            if send_email_chk:
                emp_email_val = st.text_input("Employee Email Address", placeholder="e.g. jack.smith@outlook.com", key="input_emp_email").strip()

            with st.form(key="form_create_new_employee_account"):
                st.markdown("#### 👤 New Employee Credentials & Information")
                c1, c2 = st.columns(2)
                with c1:
                    new_name = st.text_input("Employee Full Name", placeholder="e.g. Jack Smith").strip()
                    new_user = st.text_input("Username for Login", placeholder="e.g. jack.smith or jack").strip().lower()
                    new_pass = st.text_input("Initial Password", value="TempPass123!", type="password")
                with c2:
                    new_role_level = st.text_input("Role / Position", value="Junior Team Member")
                    new_emp_type = st.selectbox("Employment Classification", ["Casual", "Part-Time", "Full-Time"], index=0)
                    new_age = st.number_input("Age", min_value=14, max_value=80, value=18)

                submit_new_emp = st.form_submit_button("🚀 Create Employee Account")

                if submit_new_emp:
                    if not new_name:
                        st.error("❌ Employee name cannot be empty.")
                    elif not new_user:
                        st.error("❌ Username cannot be empty.")
                    elif new_user in user_profiles:
                        st.error(f"❌ Username '{new_user}' already exists. Please choose a different username.")
                    elif send_email_chk and not emp_email_val:
                        st.error("❌ Please enter the Employee Email Address or uncheck 'Send Welcome Email automatically'.")
                    else:
                        user_profiles[new_user] = {
                            "username": new_user,
                            "password": new_pass if new_pass else "TempPass123!",
                            "role": "Employee",
                            "employee_name": new_name,
                            "profile": {
                                "full_name": new_name,
                                "email": emp_email_val,
                                "store": "Brumby's Pakenham",
                                "classification": new_emp_type,
                                "employment_level": new_role_level,
                                "commencement_date": datetime.now().strftime("%Y-%m-%d")
                            }
                        }
                        save_user_profiles(user_profiles)

                        st.session_state.manual_employees = sync_user_profiles_to_employees(st.session_state.manual_employees)
                        save_persisted_df(st.session_state.manual_employees, "employees.csv")

                        if "edit_employees" in st.session_state:
                            del st.session_state["edit_employees"]

                        subj, body_text = build_welcome_email_content(new_name, new_user, new_pass)
                        
                        email_sent = False
                        email_msg = ""
                        if send_email_chk and emp_email_val:
                            email_sent, email_msg = send_welcome_email_smtp(emp_email_val, new_name, new_user, new_pass)

                        if email_sent:
                            st.session_state["email_notice_success"] = f"🎉 Account created for **{new_name}**! {email_msg}"
                        else:
                            if send_email_chk and emp_email_val:
                                st.session_state["email_notice_warning"] = f"🎉 Account created for **{new_name}**! Username: `{new_user}` | Initial Password: `{new_pass}`\n\n⚠️ Email Status: {email_msg}"
                            else:
                                st.session_state["email_notice_success"] = f"🎉 Account created for **{new_name}**! Username: `{new_user}` | Initial Password: `{new_pass}`"

                        st.rerun()

        # --- INTEGRATED CONFIDENTIAL EMPLOYEE PROFILE VIEWER ---
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("""
        <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 10px 18px; border-radius: 12px 12px 0 0; color: #ffffff !important; font-weight: 800; font-size: 1.1rem; letter-spacing: 0.3px; border: 2px solid #e5a93c; border-bottom: none; margin-top: 20px;">
            📜 Click / Select Employee to Open Confidential Profile Form
        </div>
        """, unsafe_allow_html=True)

        emp_options = {k: v.get("employee_name", k) for k, v in user_profiles.items() if v.get("role") == "Employee"}
        if emp_options:
            selected_user_key = st.selectbox(
                "👇 Select staff member from table above to open their confidential profile page:",
                options=list(emp_options.keys()),
                format_func=lambda x: f"👤 {emp_options[x]} (Username: {x})",
                key="select_emp_profile_in_tab"
            )
            
            if selected_user_key:
                render_confidential_profile_form(selected_user_key, is_admin=True)
                
                c_pw, c_del = st.columns(2)
                with c_pw:
                    with st.expander(f"🔑 Reset Password for {emp_options[selected_user_key]}"):
                        new_pw = st.text_input("New Password", type="password", key=f"tab_reset_pw_{selected_user_key}")
                        if st.button("Update Employee Password", key=f"tab_btn_reset_pw_{selected_user_key}"):
                            if new_pw.strip():
                                user_profiles[selected_user_key]["password"] = new_pw.strip()
                                save_user_profiles(user_profiles)
                                st.success(f"✅ Password for {emp_options[selected_user_key]} updated successfully!")
                            else:
                                st.error("Password cannot be empty.")

                with c_del:
                    with st.expander(f"🗑️ Delete Account for {emp_options[selected_user_key]}"):
                        st.warning(f"⚠️ Deleting this account will permanently remove **{emp_options[selected_user_key]}**'s login credentials, confidential profile, and roster records.")
                        confirm_del = st.checkbox(f"I understand, delete account for {emp_options[selected_user_key]}", key=f"chk_del_{selected_user_key}")
                        if st.button(f"🚨 Permanently Delete {emp_options[selected_user_key]}", key=f"btn_del_emp_{selected_user_key}"):
                            if confirm_del:
                                target_name = emp_options[selected_user_key]
                                # 1. Remove from user_profiles.json
                                if selected_user_key in user_profiles:
                                    del user_profiles[selected_user_key]
                                    save_user_profiles(user_profiles)
                                
                                # 2. Remove from manual_employees table
                                if st.session_state.manual_employees is not None and not st.session_state.manual_employees.empty:
                                    emp_df = st.session_state.manual_employees.copy()
                                    name_col = find_column(emp_df, ["name", "employee", "staff"], "NAME")
                                    if name_col in emp_df.columns:
                                        emp_df = emp_df[emp_df[name_col].astype(str).str.strip().str.lower() != target_name.strip().lower()].reset_index(drop=True)
                                        st.session_state.manual_employees = emp_df
                                        save_persisted_df(st.session_state.manual_employees, "employees.csv")
                                
                                # 3. Clear widget cache
                                if "edit_employees" in st.session_state:
                                    del st.session_state["edit_employees"]
                                    
                                st.success(f"✅ Account for **{target_name}** has been permanently deleted.")
                                st.rerun()
                            else:
                                st.error("Please check the confirmation box first.")

    # --- TAB 3: UNAVAILABILITY ---
    with tab_unavail:
        # Bakery Team Monthly Calendar Grid & Monthly Unavailability Breakdown
        render_team_monthly_calendar_grid()

    # --- TAB 4: DAILY REQUIREMENTS ---
    with tab_req:
        col_hdr_r1, col_hdr_r2 = st.columns([3, 1])
        with col_hdr_r1:
            st.subheader("Daily Bakery Shift Requirements")
        with col_hdr_r2:
            if st.button("🔄 Reset Master Requirements Data", key="btn_reset_req_master"):
                st.session_state.manual_requirements = load_persisted_df("requirements.csv", default_req)
                if "edit_requirements" in st.session_state:
                    del st.session_state["edit_requirements"]
                st.success("✅ Requirements reloaded from master data!")
                st.rerun()

        if st.session_state.manual_requirements is None or len(st.session_state.manual_requirements) <= 2:
            st.session_state.manual_requirements = load_persisted_df("requirements.csv", default_req)

        req_mode = st.radio("Upload Mode:", ["Replace current data", "Append to current data"], key="req_upload_mode", horizontal=True)
        upload_req = st.file_uploader("Upload Daily Shift personel requirement.xlsx (Optional)", type=["xlsx"], key="req_upload")
        
        if upload_req is not None:
            file_key = f"processed_{upload_req.name}_{upload_req.size}_{req_mode}"
            if st.session_state.get("last_req_file") != file_key:
                loaded = read_excel_robust(upload_req)
                if loaded is not None:
                    if req_mode == "Replace current data":
                        st.session_state.manual_requirements = loaded
                    else:
                        combined = pd.concat([st.session_state.manual_requirements, loaded], ignore_index=True).drop_duplicates()
                        st.session_state.manual_requirements = combined
                    st.session_state.last_req_file = file_key
                    save_persisted_df(st.session_state.manual_requirements, "requirements.csv")
                    st.rerun()
                    
        st.markdown("""
        <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 10px 18px; border-radius: 12px 12px 0 0; color: #ffffff !important; font-weight: 800; font-size: 1.1rem; letter-spacing: 0.3px; border: 2px solid #e5a93c; border-bottom: none; margin-top: 15px;">
            📋 Daily Shift Coverage Requirements (Mon-Sun)
        </div>
        """, unsafe_allow_html=True)
        requirements_df = st.data_editor(st.session_state.manual_requirements, num_rows="dynamic", key="edit_requirements_v2")
        st.session_state.manual_requirements = requirements_df
        save_persisted_df(requirements_df, "requirements.csv")

    # --- TAB 5: FIXED SHIFTS ---
    with tab_fixed:
        col_hdr_f1, col_hdr_f2 = st.columns([3, 1])
        with col_hdr_f1:
            st.subheader("Fixed Baseline Shifts")
        with col_hdr_f2:
            if st.button("🔄 Reset Master Fixed Shifts Data", key="btn_reset_fixed_master"):
                st.session_state.manual_fixed = sort_dataframe_by_team_and_age(load_persisted_df("fixed.csv", default_fixed))
                if "edit_fixed" in st.session_state:
                    del st.session_state["edit_fixed"]
                st.success("✅ Fixed shifts reloaded from master data!")
                st.rerun()

        if st.session_state.manual_fixed is None or len(st.session_state.manual_fixed) <= 2:
            st.session_state.manual_fixed = sort_dataframe_by_team_and_age(load_persisted_df("fixed.csv", default_fixed))

        fixed_mode = st.radio("Upload Mode:", ["Replace current data", "Append to current data"], key="fixed_upload_mode", horizontal=True)
        upload_fixed = st.file_uploader("Upload Roster fixed - dont change.xlsx (Optional)", type=["xlsx"], key="fixed_upload")
        
        if upload_fixed is not None:
            file_key = f"processed_{upload_fixed.name}_{upload_fixed.size}_{fixed_mode}"
            if st.session_state.get("last_fixed_file") != file_key:
                loaded = read_excel_robust(upload_fixed)
                if loaded is not None:
                    if fixed_mode == "Replace current data":
                        st.session_state.manual_fixed = loaded
                    else:
                        combined = pd.concat([st.session_state.manual_fixed, loaded], ignore_index=True).drop_duplicates()
                        st.session_state.manual_fixed = combined
                    st.session_state.last_fixed_file = file_key
                    save_persisted_df(st.session_state.manual_fixed, "fixed.csv")
                    st.rerun()
                    
        st.markdown("""
        <div style="background: linear-gradient(135deg, #081d19 0%, #16443c 100%); padding: 10px 18px; border-radius: 12px 12px 0 0; color: #ffffff !important; font-weight: 800; font-size: 1.1rem; letter-spacing: 0.3px; border: 2px solid #e5a93c; border-bottom: none; margin-top: 15px;">
            📌 Fixed Baseline Staff Shifts
        </div>
        """, unsafe_allow_html=True)
        if st.session_state.manual_fixed is not None and not st.session_state.manual_fixed.empty:
            st.session_state.manual_fixed = sort_dataframe_by_team_and_age(st.session_state.manual_fixed)
        fixed_df = st.data_editor(st.session_state.manual_fixed, num_rows="dynamic", key="edit_fixed_v2")
        st.session_state.manual_fixed = fixed_df
        save_persisted_df(fixed_df, "fixed.csv")

    # --- TAB 7: SHIFT TIMESHEET AUDIT & LIVE ATTENDANCE ---
    with tab_timesheets:
        render_manager_timesheet_audit_dashboard()
